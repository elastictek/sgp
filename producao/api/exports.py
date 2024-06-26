import base64
from operator import eq
from pyexpat import features
import re
from typing import List
from wsgiref.util import FileWrapper
from rest_framework.generics import ListAPIView, RetrieveAPIView, CreateAPIView
from rest_framework.views import APIView
from django.http import Http404, request, StreamingHttpResponse
from rest_framework.response import Response
from django.http.response import HttpResponse
from django.http import FileResponse
from producao.models import Artigo, Palete, Bobine, Emenda, Bobinagem, Cliente, Encomenda, Carga
from .serializers import ArtigoDetailSerializer, PaleteStockSerializer, PaleteListSerializer, PaleteDetailSerializer, CargaListSerializer, PaletesCargaSerializer, CargasEncomendaSerializer, CargaDetailSerializer, BobineSerializer, EncomendaListSerializer, BobinagemCreateSerializer, BobinesDmSerializer, BobinesPaleteDmSerializer, EmendaSerializer, EmendaCreateSerializer, BobinagemListSerializer, BobineListAllSerializer, ClienteSerializer, BobinagemBobinesSerializer, PaleteDmSerializer
from django.contrib.auth.mixins import LoginRequiredMixin
from rest_framework import status
import mimetypes
from datetime import datetime, timedelta
from io import BytesIO
# import cups
import os, tempfile
from producao.api import reports

from pyodbc import Cursor, Error, connect, lowercase
from datetime import datetime
from django.http.response import JsonResponse
from rest_framework.decorators import api_view, authentication_classes, permission_classes, renderer_classes
from django.db import connections, transaction
from support.database import encloseColumn, Filters, DBSql, TypeDml, fetchall, Check
from support.myUtils import  ifNull

from rest_framework.renderers import JSONRenderer, MultiPartRenderer, BaseRenderer
from rest_framework.utils import encoders, json
from rest_framework.authentication import SessionAuthentication, BasicAuthentication
from rest_framework.permissions import IsAuthenticated
import collections
import hmac
import hashlib
import math
from django.core.files.storage import FileSystemStorage
from sistema.settings.appSettings import AppSettings
import time
import requests
import psycopg2
import pandas as pd
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill

def max_length(worksheet):
    column_widths = {}
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value:
                if cell.column in column_widths:
                    column_widths[cell.column_letter] = max(column_widths[cell.column_letter], len(str(cell.value)))
                else:
                    column_widths[cell.column_letter] = len(str(cell.value))
    for i, column_width in column_widths.items():
        worksheet.column_dimensions[i].width = column_width + 10

def getKeyInsensitive(dictionary, key):
    for dict_key in dictionary.keys():
        if dict_key.lower() == key.lower():
            return dict_key
    return None  # Return None if the key is not found

def getCols(req):
    get_string_after_last_dot = lambda input_string: input_string.split('.')[-1] if '.' in input_string else input_string
    req_cols = {}
    for key, value in req["cols"].items():
        if "reportField" in value and value["reportField"] is not None:
            k=value["reportField"]
        elif "field" in value and value["field"] is not None:
            k=value["field"]
        else:
            k = get_string_after_last_dot(key)
        if "visible" in value and value["visible"]==False:
            continue
        req_cols[k] = req["cols"][key]
    return req_cols

def exportRunxlslist(req,dbi,conn):
    try:
        response = dbi.executeSimpleList(req["sql"], conn, req["data"])
        df = pd.DataFrame(response["rows"])
        #lowercase_reqcols = [column.lower() for column in req["cols"].keys()]
        #lowercase_columns = [column.lower() for column in df.columns.values.tolist()]
        #cols = [column for column in df.columns.values.tolist() if column.lower() in lowercase_reqcols]
        
        req_cols = getCols(req)

        #lowercase_reqcols = [column.lower() for column in req["cols"].keys()]
        lowercase_columns = [column.lower() for column in df.columns.values.tolist()]
        cols = [column for column in req_cols if column.lower() in lowercase_columns]
        _cols = lowercase_columns if "notListed" in req and req["notListed"]==True else cols
        rcols = []
        for v in _cols:
            df_k = getKeyInsensitive(df,v)
            rq_k = getKeyInsensitive(req_cols,v)
            if (rq_k in req_cols and "format" in req_cols[rq_k]):
                if req_cols[rq_k]["format"]=='0':
                    df[df_k] = df[df_k].fillna(0).astype(int)
                elif req_cols[rq_k]["format"]=='0.0':
                    df[df_k] = df[df_k].fillna(0).astype(float).round(1)
                elif req_cols[rq_k]["format"]=='0.00':
                    df[df_k] = df[df_k].fillna(0).astype(float).round(2)
                elif req_cols[rq_k]["format"]=='0.000':
                    df[df_k] = df[df_k].fillna(0).astype(float).round(3)
                elif req_cols[rq_k]["format"]=='0.0000':
                    df[df_k] = df[df_k].fillna(0).astype(float).round(4)
            if rq_k in req_cols and "title" in req_cols[rq_k]:
                df.rename(columns = {df_k:req_cols[rq_k]["title"]}, inplace = True)
                rcols.append(req_cols[rq_k]["title"])
            else:
                rcols.append(df_k)

        #print(req["cols"].keys())
        #df = df.loc[:, df.columns.isin(req["cols"].keys())]
        # Create a new Excel file and add the dataframe to it
        #book = Workbook()
        excel_file  = BytesIO()
        #writer = pd.ExcelWriter(sio, engine='openpyxl') 
        #writer.book = book

        df.to_excel(excel_file, engine='openpyxl', index=False,columns=rcols)
        
        excel_file.seek(0)
        wb = load_workbook(filename=excel_file)
        ws = wb.active
        header = ws[1]
        
        # for cell in header:
        #     if "format" in req["cols"][cell.value]:
        #         print(cell.column_letter)
        #         print(req["cols"][cell.value]["format"])
        #         cell.number_format = req["cols"][cell.value]["format"]
        #     cell.value = req["cols"][cell.value]["title"]
        #     cell.fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type = "solid")


        excel_file.close()
        excel_file  = BytesIO()
        wb.save(excel_file)
        #sio.seek(0)
        #wb = load_workbook(sio)
        #ws = wb.active


        # Get the worksheet and format the header
        #worksheet = writer.book['Sheet1']
        #header = worksheet[1] 
        #for cell in header:
        #    cell.fill = openpyxl.styles.PatternFill(start_color="0066CC", end_color="0066CC", fill_type = "solid")

        # Save the changes
        #x  = BytesIO()
        #wb.save(excel_file)
        excel_file.seek(0)
        response =  HttpResponse(excel_file.read(), content_type="application/ms-excel")
        response['Content-Disposition'] = 'attachment; filename=list.xlsx'
        return response
        

    except Exception as error:
        print(str(error))
        return Response({"status": "error", "title": str(error)})

def sqlCols(cols):
    column_names = []
    for key, value in cols.items():
        if "field" in value:
            column_names.append(f"""{key} {value["field"]}""")
        else:
            column_names.append(key)
    return ",".join(column_names)

def export(sql, db_parameters, parameters,conn_name,dbi=None,conn=None):
    if ("export" in parameters and parameters["export"] is not None):
        _sql = None
        if callable(sql):
            _sql=sql(lambda v:v,lambda v:v,lambda v:v)
        else:
            _sql=sql
        dbparams={}
        for key, value in db_parameters.items():
            if f"%({key})s" not in _sql: 
                continue
            dbparams[key] = value
        if parameters["export"] == "clean-excel" or parameters["export"] not in ["pdf", "excel", "word"]:
            req = {
                "sql":_sql,
                "data":dbparams,
                "cols":parameters["cols"],
                "notListed":parameters.get("notListed")
            }
            return exportRunxlslist(req,dbi,conn)

        markPattern = r'\[#MARK-[A-Z]+-\d+#\]'
        _sql = re.sub(markPattern,'',_sql)
        for key, value in db_parameters.items():
            _sql = _sql.replace(f"%({key})s",f""":{key.replace(".","_")}""")        
        hash = base64.b64encode(hmac.new(bytes("SA;PA#Jct\"#f.+%UxT[vf5B)XW`mssr$" , 'utf-8'), msg = bytes(_sql , 'utf-8'), digestmod = hashlib.sha256).hexdigest().upper().encode()).decode()
        req = {            
            "conn-name":conn_name,
            "sql":_sql,
            "hash":hash,
            "data":{key.replace('.', '_'): value for key, value in db_parameters.items()},
            **parameters,
            "cols":getCols(parameters)
        }
        wService = "runxlslist" if parameters["export"] == "clean-excel" else "runlist"
        fstream = requests.post(f'{AppSettings.reportServer["default"]}/{wService}', json=req)
        if (fstream.status_code==200):
            resp =  HttpResponse(fstream.content, content_type=fstream.headers["Content-Type"])
            if (parameters["export"] == "pdf"):
                resp['Content-Disposition'] = "inline; filename=list.pdf"
            elif (parameters["export"] == "excel"):
                resp['Content-Disposition'] = "inline; filename=list.xlsx"
            elif (parameters["export"] == "word"):
                resp['Content-Disposition'] = "inline; filename=list.docx"
            if (parameters["export"] == "csv"):
                resp['Content-Disposition'] = "inline; filename=list.csv"
            return resp