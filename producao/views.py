import os
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse, reverse_lazy, resolve
from django.forms.models import modelformset_factory, inlineformset_factory
from django import forms
from django.shortcuts import render, get_object_or_404, HttpResponseRedirect, redirect, reverse, HttpResponse
from django.views.generic import CreateView
from django.views.generic import ListView, DetailView, CreateView, TemplateView, View, FormView, UpdateView
# from .forms import CreateNonwovenManual, SearchBobinagem, PerfilDMForm, SearchPerfil, PerfilLinhaForm, ImprimirEtiquetaFinalPalete, ImprimirEtiquetaPalete, ImprimirEtiquetaBobine, PicagemBobines, PerfilCreateForm, ClassificacaoBobines, LarguraForm, BobinagemCreateForm, BobineStatus, AcompanhamentoDiarioSearchForm, ConfirmReciclarForm, RetrabalhoFormEmendas, PaleteCreateForm, SelecaoPaleteForm, AddPalateStockForm, PaletePesagemForm, RetrabalhoCreateForm, CargaCreateForm, EmendasCreateForm, ClienteCreateForm, UpdateBobineForm, PaleteRetrabalhoForm, OrdenarBobines, ClassificacaoBobines, RetrabalhoForm, EncomendaCreateForm
from .forms import *
from .models import Cartao, Especificacoes, TrasporteArtigoCliente, Transporte, Mdf, Artigo, Cinta, CoreLargura, Filme, LinhaEncomenda, PaleteEmb, PerfilEmbalamento, ProdutoGranulado, Reciclado, MovimentoMP, EtiquetaReciclado, MovimentosBobines, InventarioBobinesDM, InventarioPaletesCliente, Nonwoven, ConsumoNonwoven, EtiquetaFinal, Largura, Perfil, Bobinagem, Bobine, Palete, Emenda, Cliente, EtiquetaRetrabalho, Encomenda, EtiquetaPalete, ArtigoCliente, Rececao, ArtigoNW
from django.utils import timezone
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import JsonResponse, Http404, HttpResponse, HttpResponseRedirect
from django.db.models.signals import pre_save, post_save
from django.contrib import messages
from time import gmtime, strftime
import datetime
import time
from .funcs import *
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.db.models import Q
from django.forms import formset_factory
from django.http import HttpResponse
from datetime import datetime,timedelta
from urllib.error import HTTPError
import django.core.exceptions
import csv
import xlsxwriter
import io
from calendar import monthrange
from datetime import date
from collections import defaultdict
from django.contrib.sessions.models import Session
from planeamento.models import *
from django.template.loader import get_template
from xhtml2pdf import pisa
from django.conf import settings
from django.contrib.staticfiles import finders
import pyodbc
import openpyxl
from django.db import connections, transaction
from producao.api.views import addToReciclado


# class CreatePerfil(LoginRequiredMixin, CreateView):
#     template_name = 'perfil/perfil_create.html'
#     form_class = PerfilCreateForm
#     success_url = '/producao/perfil/{id}/'

#     def form_valid(self, form):
#         form.instance.user = self.request.user
#         return super().form_valid(form)

@login_required
def create_perfil(request):

    template_name = 'perfil/perfil_create.html'
    form = PerfilCreateForm(request.POST or None)

    if form.is_valid():
        instance = form.save(commit=False)
        instance.user = request.user
        instance.save()

        for i in range(instance.num_bobines):
            lar = Largura.objects.create(
                perfil=instance, num_bobine=i+1, designacao_prod=instance.produto)
            lar.save()

        return redirect('producao:perfil_details', pk=instance.pk)

    context = {
        "form": form
    }

    return render(request, template_name, context)


@login_required
def create_bobinagem(request):

    
    template_name = 'producao/bobinagem_create.html'
    form = BobinagemCreateForm(request.POST or None)

    # form = BobinagemCreateForm(initial=num)

    if form.is_valid():
        instance = form.save(commit=False)
        instance.user = request.user

        sup = instance.lotenwsup
        inf = instance.lotenwinf

        metros_nwsup = instance.nwsup
        metros_nwinf = instance.nwinf

        instance.lotenwsup = sup.replace(" ", "")
        instance.lotenwinf = inf.replace(" ", "")

        if (instance.estado == 'R' and instance.obs == ''):
            messages.error(
                request, 'Para Rejeitar a bobinagem é necessario indicar motivo nas observações.')
        else:
            nonwovensup = Bobinagem.objects.filter(lotenwsup=sup)
            nonwoveninf = Bobinagem.objects.filter(lotenwinf=inf)

            total_sup = instance.nwsup
            total_inf = instance.nwinf
            for ns in nonwovensup:
                total_sup += ns.nwsup
            for ni in nonwoveninf:
                total_inf += ni.nwinf

            # if (total_sup > 7500):
            #     messages.error(request, 'A soms total de metros do lote de Nonwoven superior "' + sup + '" excede o limite establecido de 7500. Por favor verifique o valor introduzido.')
            # if (total_inf > 7500):

            if (total_inf > 16000 or total_sup > 16000):
                messages.error(
                    request, 'A soma total de metros dos lotes de Nonwoven excedem o limite establecido. Por favor verifique os valores introduzidos.')
            else:
                instance.save()
                bobinagem_create(instance.pk)

                if not instance.estado == 'LAB' or instance.estado == 'HOLD':
                    areas(instance.pk)

                if instance.estado == 'BA' or instance.num_bobinagem == 1 or multipleOf10(instance.num_bobinagem) == 0:
                    return redirect('producao:bobines_larguras_reais', pk=instance.pk)
                else:
                    return redirect('producao:etiqueta_retrabalho', pk=instance.pk)

    context = {
        "form": form
    }

    return render(request, template_name, context)


def perfil_list(request):
    perfil = Perfil.objects.all()
    # paginator = Paginator(perfil, 15)
    # page = request.GET.get('page')
    template_name = 'perfil/perfil_home.html'

    # try:
    #     perfil = paginator.page(page)
    # except PageNotAnInteger:
    #     perfil = paginator.page(1)
    # except EmptyPage:
    #     perfil = paginator.page(paginator.num_pages)

    context = {
        "perfil": perfil,
    }
    return render(request, template_name, context)


@login_required
def perfil_detail(request, pk):
    perfil = Perfil.objects.get(pk=pk)
    largura = Largura.objects.filter(perfil=pk)
    template_name = 'perfil/perfil_detail.html'
    context = {
        'perfil': perfil,
        'largura': largura,
    }
    return render(request, template_name, context)


# class BobinagemListView(LoginRequiredMixin, ListView):
#     model = Bobinagem
#     template_name = 'producao/bobinagem_home.html'

#     def get_context_data(self, **kwargs):
#         context = super().get_context_data(**kwargs)
#         context['now'] = timezone.now()
#         return context

def bobinagem_list(request):
    now = datetime.datetime.now()
    bobinagem = Bobinagem.objects.all()

    template_name = 'producao/bobinagem_home.html'
    bobine = Bobine.objects.all()

    context = {
        "bobinagem": bobinagem,
        "bobine": bobine,
        "now": now

    }
    return render(request, template_name, context)


def bobinagem_historico(request):
    bobinagem = Bobinagem.objects.all()
    s = request.GET.get("s")
    if s:
        bobinagem = Bobinagem.objects.filter(
            Q(nome__icontains=s) | Q(data__icontains=s))

    paginator = Paginator(bobinagem, 17)
    page = request.GET.get('page')
    template_name = 'producao/bobinagem_all.html'
    bobine = Bobine.objects.all()

    try:
        bobinagem = paginator.page(page)
    except PageNotAnInteger:
        bobinagem = paginator.page(1)
    except EmptyPage:
        bobinagem = paginator.page(paginator.num_pages)

    context = {
        "bobinagem": bobinagem,
        "bobine": bobine,


    }
    return render(request, template_name, context)


@login_required
def bobinagem_status(request, pk):
    template_name = 'producao/bobinagem_details.html'
    bobinagem = Bobinagem.objects.get(pk=pk)
    bobine = Bobine.objects.filter(bobinagem=pk)
    emenda = Emenda.objects.filter(bobinagem=pk)
    etiquetas = EtiquetaRetrabalho.objects.filter(bobinagem=pk)
    #etiquetas_all = EtiquetaRetrabalho.objects.all()
    estado_impressao = False
    form = ImprimirEtiquetaBobine(request.POST or None)
    if form.is_valid():
        impressora = form['impressora'].value()
        num_copias = int(form['num_copias'].value())

        #for etiqueta in etiquetas_all:
        #    if etiqueta.estado_impressao == 1:
        #        estado_impressao = True
        #        break

        if estado_impressao == False:
            for etiqueta in etiquetas:
                etiqueta.impressora = impressora
                etiqueta.num_copias = num_copias
                etiqueta.estado_impressao = True
                etiqueta.save()
                # messages.warning(request, 'SUCESSO')
        else:
            messages.warning(
                request, 'Impressão em curso noutro posto. Tente de novo em 10 segundos.')

    context = {
        "bobinagem": bobinagem,
        "bobine": bobine,
        "emenda": emenda,
        "form": form
    }

    return render(request, template_name, context)


class LarguraUpdate(LoginRequiredMixin, UpdateView):
    model = Largura
    fields = ['largura', 'designacao_prod', 'gsm']
    template_name = 'perfil/largura_update.html'


def update_bobine(request, pk=None):

    instance = get_object_or_404(Bobine, pk=pk)
    template_name = 'producao/bobine_update.html'
    estado_anterior = instance.estado
    form = UpdateBobineForm(request.POST or None, instance=instance)
    context = {
        "form": form,
        "instance": instance,
        "title": instance.nome
    }
    if form.is_valid():
        instance = form.save(commit=False)

        if (instance.estado == 'DM'):
            if (instance.con == False and instance.furos == False and instance.descen == False and instance.esp == False and instance.presa == False and instance.troca_nw == False and instance.diam_insuf == False and instance.buraco == False and instance.outros == False and instance.nok == False):
                messages.error(
                    request, 'Para classificar a bobine como estado DM, é necessário atribuir-lhe pelo menos um motivo.')
            elif (instance.outros == True and instance.obs == ''):
                messages.error(
                    request, 'Obrigatório escrever motivo nas observações')
            else:
                update_areas_bobine(instance.pk, estado_anterior)
                instance.save()
                return redirect('producao:bobinestatus', pk=instance.bobinagem.pk)
        elif (instance.estado == 'R' and instance.obs == ''):
            messages.error(
                request, 'Obrigatório escrever o motivo da rejeição nas observações')
        else:
            instance.save()
            update_areas_bobine(instance.pk, estado_anterior)
            return redirect('producao:bobinestatus', pk=instance.bobinagem.pk)

    return render(request, template_name, context)


class BobinagemUpdate(LoginRequiredMixin, UpdateView):
    model = Bobinagem
    fields = ['tiponwsup', 'tiponwinf', 'lotenwsup', 'lotenwinf',
              'nwsup', 'nwinf', 'comp', 'comp_par', 'diam', 'inico', 'fim']
    template_name = 'producao/bobinagem_update.html'


# class PaleteListView(LoginRequiredMixin, ListView):
#     model = Palete
#     template_name = 'palete/palete_home.html'

#     def get_context_data(self, **kwargs):
#         context = super().get_context_data(**kwargs)
#         context['now'] = timezone.now()
#         return context

@login_required
def pelete_list(request):
    palete = Palete.objects.all()
    e_p = EtiquetaPalete.objects.all()
    # paginator = Paginator(palete, 15)
    # page = request.GET.get('page')
    template_name = 'palete/palete_home.html'

    # try:
    #     palete = paginator.page(page)
    # except PageNotAnInteger:
    #     palete = paginator.page(1)
    # except EmptyPage:
    #     palete = paginator.page(paginator.num_pages)

    context = {
        "palete": palete,
        "e_p": e_p
    }
    return render(request, template_name, context)


# class PaleteCreateView(LoginRequiredMixin, CreateView):
#     form_class = PaleteCreateForm
#     template_name = "palete/palete_create.html"
#     success_url = "/producao/palete/{id}"

#     def form_valid(self, form):
#         form.instance.user = self.request.user
#         return super().form_valid(form)

def create_palete(request):
    form = PaleteCreateForm(request.POST or None)
    template_name = "palete/palete_create.html"

    if form.is_valid():
        instance = form.save(commit=False)
        nome_s_r = ''
        nome_c_r = ''
        ano = instance.data_pal
        ano = ano.strftime('%Y')
        instance.user = request.user
        instance.estado = 'G'
        instance.area = 0
        instance.comp_total = 0

        if instance.num < 10:
            nome_s_r = 'P000%s-%s' % (instance.num, ano)
            nome_c_r = 'R000%s-%s' % (instance.num, ano)
        elif instance.num < 100:
            nome_s_r = 'P00%s-%s' % (instance.num, ano)
            nome_c_r = 'R00%s-%s' % (instance.num, ano)
        elif instance.num < 1000:
            nome_s_r = 'P0%s-%s' % (instance.num, ano)
            nome_c_r = 'R0%s-%s' % (instance.num, ano)
        else:
            nome_s_r = 'P%s-%s' % (instance.num, ano)
            nome_c_r = 'R%s-%s' % (instance.num, ano)

        if Palete.objects.filter(nome=nome_s_r).exists() or Palete.objects.filter(nome=nome_c_r).exists():
            messages.error(request, 'A palete nº' + str(instance.num) + ' de ' + str(ano) + ' já existe.')
        elif instance.perfil_embalamento == None:
            messages.error(request, 'O perfil de embalamento é de preenchimento obrigatório.')
        else:
            if instance.ordem != None:
                ordem = OrdemProducao.objects.get(pk=instance.ordem.pk)
                instance.ordem_original = ordem.op
                instance.ordem_id_original = ordem.pk
                
                try:
                    encomenda = Encomenda.objects.get(pk=ordem.enc.pk)
                    num_paletes_ordem = Palete.objects.filter(ordem=instance.ordem).count()
                    ordem.num_paletes_produzidas += 1
                    encomenda.num_paletes_actual += 1
                    if num_paletes_ordem == 0:
                        instance.num_palete_ordem = 1
                    else:
                        try:
                            last_palete = Palete.objects.filter(ordem=instance.ordem).latest('num_palete_ordem')
                            instance.num_palete_ordem = last_palete.num_palete_ordem + 1
                        except:
                            instance.num_palete_ordem = 1
                    instance.cliente = encomenda.cliente
                    instance.destino = encomenda.cliente.nome + ' L' + str(int(instance.largura_bobines)) + ' ' + str(instance.num_palete_ordem)
                    ordem.save()
                    encomenda.save()
                    instance.save()
                    palete_nome(instance.pk)
                    return redirect('producao:palete_picagem_v3', pk=instance.pk)
                    
                except:
                    ordem = OrdemProducao.objects.get(pk=instance.ordem.pk)
                    num_paletes_ordem = Palete.objects.filter(ordem=instance.ordem).count()
                    instance.ordem_original = ordem.op
                    instance.ordem_id_original = ordem.pk
                    instance.ordem_original_stock = True
                    ordem.num_paletes_produzidas += 1
                    if num_paletes_ordem == 0:
                        instance.num_palete_ordem = 1
                    else:
                        last_palete = Palete.objects.filter(ordem=instance.ordem).latest('num_palete_ordem')
                        instance.num_palete_ordem = last_palete.num_palete_ordem + 1
                    instance.destino = ordem.cliente.nome + ' STOCK L' + str(int(instance.largura_bobines)) + ' ' + str(instance.num_palete_ordem)
                    instance.cliente = ordem.cliente
                    ordem.save()
                    instance.save()
                    instance.stock = True
                    instance.save()
                    palete_nome(instance.pk)
                    return redirect('producao:palete_picagem_v3', pk=instance.pk)
            else:
                messages.error(
                    request, 'Não é possivel criar uma palete sem ordem de produção associada. Por favor selecione uma Ordem de produção em progresso.')

    context = {
        "form": form
    }

    return render(request, template_name, context)


def ordenar_bobines_op(request, pk, operation):
    bobine = Bobine.objects.get(pk=pk)
    palete = Palete.objects.get(pk=bobine.palete.pk)
    pos = bobine.posicao_palete
    # pos_up = 0
    # pos_down = 0
    # bobine_up = 0
    # bobine_down = 0

    if Bobine.objects.filter(posicao_palete=pos-1, palete=palete).exists():
        bobine_up = Bobine.objects.get(posicao_palete=pos-1, palete=palete)
        pos_up = bobine_up.posicao_palete

    if Bobine.objects.filter(posicao_palete=pos+1, palete=palete).exists():
        bobine_down = Bobine.objects.get(posicao_palete=pos+1, palete=palete)
        pos_down = bobine_down.posicao_palete

    if operation == 'up':
        bobine.posicao_palete = pos - 1
        bobine_up.posicao_palete = pos_up + 1
        bobine.save()
        bobine_up.save()
    elif operation == 'down':
        bobine.posicao_palete = pos + 1
        bobine_down.posicao_palete = pos_down - 1
        bobine.save()
        bobine_down.save()

    return redirect('producao:addbobinepalete', pk=palete.pk)


def create_palete_retrabalho(request):

    form = PaleteRetrabalhoForm(request.POST or None)
    template_name = "retrabalho/palete_retrabalho_create.html"

    if form.is_valid():
        instance = form.save(commit=False)
        nome = ''
        ano = instance.data_pal
        ano = ano.strftime('%Y')
        instance.user = request.user
        instance.retrabalhada = True
        instance.estado = 'DM'
        instance.num_bobines_act = 0
        instance.largura_bobines = 0
        instance.area = 0
        instance.comp_total = 0

        if instance.num < 10:
            nome = 'DM000%s-%s' % (instance.num, ano)
        elif instance.num < 100:
            nome = 'DM00%s-%s' % (instance.num, ano)
        elif instance.num < 1000:
            nome = 'DM0%s-%s' % (instance.num, ano)
        else:
            nome = 'DM%s-%s' % (instance.num, ano)

        # print(nome, ano, instance.num)
        if Palete.objects.filter(nome=nome).exists():
            messages.error(request, 'A palete nº' + str(instance.num) + ' de ' + str(ano) + ' já existe.')
        elif instance.perfil_embalamento == None:
            messages.error(request, 'O perfil de embalamento é de preenchimento obrigatório.')
        else:
            instance.save()
            palete_nome(instance.pk)

            if EtiquetaPalete.objects.filter(palete=instance).exists():
                return redirect('producao:palete_picagem_dm', pk=instance.pk)
            else:
                e_p = EtiquetaPalete.objects.create(palete=instance, palete_nome=instance.nome, largura_bobine=instance.largura_bobines)
                e_p.save()
                return redirect('producao:palete_picagem_dm', pk=instance.pk)

            

    context = {
        "form": form,
    }
    return render(request, template_name, context)


@login_required
def add_bobine_palete(request, pk):
    template_name = 'palete/add_bobine_palete.html'
    palete = Palete.objects.get(pk=pk)
    estado = palete.estado
    ordem_palete_total = 0
    if estado == "G":
        ordemid = palete.ordem_id
        if ordemid is not None:
            ordem_palete_total = OrdemProducao.objects.get(pk=ordemid).num_paletes_total
        bobinagem = Bobinagem.objects.filter(diam=palete.diametro)
        bobine = Bobine.objects.all().order_by('posicao_palete')
        bobines = Bobine.objects.filter(palete=palete).order_by('posicao_palete')
        movimentos_bobines = MovimentosBobines.objects.filter(palete=palete)
        has_carga = False
    else:
        bobinagem = Bobinagem.objects.filter(diam=palete.diametro)
        bobine = Bobine.objects.all().order_by('posicao_palete')
        bobines = Bobine.objects.filter(palete=palete).order_by('posicao_palete')
        movimentos_bobines = MovimentosBobines.objects.filter(palete=palete)
        has_carga = False
    try:
        carga = Carga.objects.get(id=palete.carga.id)
        has_carga = True
    except:
        pass

    form = ImprimirEtiquetaPalete(request.POST or None)

    if form.is_valid():
        etiqueta = EtiquetaPalete.objects.get(palete=palete)
        impressora = form['impressora'].value()
        num_copias = int(form['num_copias'].value())
        etiqueta.impressora = impressora
        etiqueta.num_copias = num_copias
        etiqueta.estado_impressao = True
        etiqueta.save()

    context = {
        "has_carga": has_carga,
        "movimentos_bobines": movimentos_bobines,
        "palete": palete,
        "bobines": bobines,
        "bobinagem": bobinagem,
        "form": form,
        "totalordem": ordem_palete_total
    }
    return render(request, template_name, context)


@login_required
def palete_change(request, operation, pk_bobine, pk_palete):

    palete = Palete.objects.get(pk=pk_palete)
    bobine = Bobine.objects.get(pk=pk_bobine)

    if operation == 'add':
        Bobine.add_bobine(palete.pk, bobine.pk)
    elif operation == 'remove':
        Bobine.remove_bobine(palete.pk, bobine.pk)

    return redirect('producao:addbobinepalete', pk=palete.pk)


class RetrabalhoCreateView(LoginRequiredMixin, CreateView):
    form_class = RetrabalhoCreateForm
    template_name = 'retrabalho/retrabalho_create.html'
    success_url = "/producao/retrabalho/filter/{id}"

    def form_valid(self, form):
        form.instance.user = self.request.user
        return super().form_valid(form)


@login_required
def create_bobinagem_retrabalho(request):
    template_name = 'retrabalho/retrabalho_create.html'
    form = RetrabalhoCreateForm(request.POST or None)
    if form.is_valid():
        instance = form.save(commit=False)

        bobinagem_nome = bobinagem_retrabalho_nome(
            instance.data, instance.num_bobinagem)

        if Bobinagem.objects.filter(nome=bobinagem_nome[0]).exists() or Bobinagem.objects.filter(nome=bobinagem_nome[1]).exists():
            messages.error(
                request, 'A bobinagem que deseja criar já existe. por favor verifique o numero da bobinagem.')
        else:
            instance.user = request.user
            instance.nome = bobinagem_nome[0]
            instance.save()
            area_bobinagem(instance.pk)
            create_bobine(instance.pk)
            return redirect('producao:retrabalho_v2', pk=instance.pk)

    context = {
        "form": form
    }

    return render(request, template_name, context)


@login_required
def picagem_retrabalho(request, pk):

    bobinagem = Bobinagem.objects.get(pk=pk)
    palete = Palete.objects.filter(estado="DM")
    bobine = Bobine.objects.all()
    emenda = Emenda.objects.filter(bobinagem=pk)

    q_largura = request.GET.get("l")

    if q_largura:
        palete = palete.filter(largura_bobines__gte=q_largura)

    template_name = 'retrabalho/retrabalho_inicio.html'
    context = {

        "palete": palete,
        "bobine": bobine,
        "bobinagem": bobinagem,
        "emenda": emenda

    }

    return render(request, template_name, context)


@login_required
def picagem_retrabalho_add(request, pk):

    bobinagem = Bobinagem.objects.get(pk=pk)

    q_bobine = request.POST.get('b')
    q_metros = int(request.POST.get('m'))
    # else:
    #     messages.error(request, 'Por favor introduza a bobine desejada e os metros a retrabalhar.')
    #     return redirect('producao:retrabalho_filter', pk=bobinagem.pk)

    try:
        bobine = Bobine.objects.get(nome=q_bobine)
    except:
        messages.error(request, 'A bobine selecionada não existe.')
        return redirect('producao:retrabalho_filter', pk=bobinagem.pk)

    if bobine:
        if bobine.estado != 'DM':
            messages.error(
                request, 'O estado da bobine selecionada não permite efectuar esta operação.')
        elif q_metros > bobine.comp_actual:
            messages.error(
                request, 'A bobine selecionada não tem comprimento suficiente para efectuar esta operação.')
        else:
            emenda = Emenda.objects.create(
                bobinagem=bobinagem, bobine=bobine, metros=q_metros)
            bobine.comp_actual -= emenda.metros
            bobine.save()
            emenda.num_emenda = bobinagem.num_emendas + 1
            emenda.save()
            emenda_num = emenda.num_emenda
            if emenda_num == 0:
                emenda.emenda = 0
                emenda.save()
            elif emenda_num > 0:
                if emenda_num == 1:
                    emenda.emenda = emenda.metros
                    emenda.save()
                elif emenda_num == 2:
                    emenda_ul_bob = Emenda.objects.get(
                        bobinagem=bobinagem, num_emenda=1)
                    emenda.emenda = emenda_ul_bob.emenda + emenda.metros
                    emenda.save()
                elif emenda_num == 3:
                    emenda_ul_bob = Emenda.objects.get(
                        bobinagem=bobinagem, num_emenda=2)
                    emenda.emenda = emenda_ul_bob.emenda + emenda.metros
                    emenda.save()

            bobinagem.num_emendas += 1
            data = bobinagem.data
            data = data.strftime('%Y%m%d')
            map(int, data)
            if bobinagem.perfil.retrabalho == True and bobinagem.num_emendas > 1:
                if bobinagem.num_bobinagem < 10:
                    bobinagem.nome = '3%s-0%s' % (data[1:],
                                                  bobinagem.num_bobinagem)
                    bobinagem_pk = bobinagem.pk
                    bobine_nome(bobinagem_pk)
                else:
                    bobinagem.nome = '3%s-%s' % (data[1:],
                                                 bobinagem.num_bobinagem)
            elif bobinagem.perfil.retrabalho == True and bobinagem.num_emendas == 0:
                if bobinagem.num_bobinagem < 10:
                    bobinagem.nome = '4%s-0%s' % (data[1:],
                                                  bobinagem.num_bobinagem)
                else:
                    bobinagem.nome = '4%s-%s' % (data[1:],
                                                 bobinagem.num_bobinagem)

            bobinagem.comp = bobinagem.comp + q_metros
            bobinagem.save()

        return redirect('producao:retrabalho_filter', pk=bobinagem.pk)


class ClienteCreateView(LoginRequiredMixin, CreateView):
    form_class = ClienteCreateForm
    template_name = 'cliente/cliente_create.html'
    success_url = "/producao/clientes/"


@login_required
def picagem(request, pk):
    palete = Palete.objects.get(pk=pk)
    bob = request.POST.get('q')
    try:
        bobine = Bobine.objects.get(nome=bob)
    except:
        messages.error(request, 'A bobine selecionada não existe.')
        return redirect('producao:addbobinepalete', pk=palete.pk)
    if bobine:
        if bobine.palete:
            if bobine.palete == palete:
                messages.error(request, 'A bobine já faz parte desta palete.')
                # erro = 4
                # return redirect('producao:addbobinepaleteerro', pk=palete.pk, e=erro)
                return redirect('producao:addbobinepalete', pk=palete.pk)
            else:
                messages.error(
                    request, 'A bobine já faz parte de outra palate.')
                # erro = 5
                # return redirect('producao:addbobinepaleteerro', pk=palete.pk, e=erro)
                return redirect('producao:addbobinepalete', pk=palete.pk)
        else:
            if (bobine.estado == 'G' or bobine.estado == 'LAB') and palete.estado != 'DM':
                if palete.num_bobines_act == palete.num_bobines:
                    messages.error(
                        request, 'A palete já se encontra completa.')
                    #  erro = 3
                    #  return redirect('producao:addbobinepaleteerro', pk=palete.pk, e=erro)
                    return redirect('producao:addbobinepalete', pk=palete.pk)
                else:
                    if (bobine.bobinagem.diam == palete.diametro or palete.cliente.limsup >= bobine.bobinagem.diam >= palete.cliente.liminf) and bobine.bobinagem.perfil.core == palete.core_bobines and bobine.largura.largura == palete.largura_bobines:
                        Bobine.add_bobine(palete.pk, bobine.pk)
                        #  etiqueta_add_bobine(palete.pk, bobine.pk)
                        return redirect('producao:addbobinepalete', pk=palete.pk)
                    else:
                        messages.error(
                            request, 'A bobine selecionada está fora de especificações.')
                        #  erro = 1
                        #  return redirect('producao:addbobinepaleteerro', pk=palete.pk, e=erro)
                        return redirect('producao:addbobinepalete', pk=palete.pk)
            elif (bobine.estado == 'DM' or bobine.estado == 'G' or bobine.estado == 'IND' or bobine.estado == 'HOLD') and palete.estado == 'DM':
                Bobine.add_bobine(palete.pk, bobine.pk)
                #   etiqueta_add_bobine(palete.pk, bobine.pk)
                return redirect('producao:addbobinepalete', pk=palete.pk)
            else:
                messages.error(
                    request, 'A bobine selecionada está fora de especificações.')
                return redirect('producao:addbobinepalete', pk=palete.pk)

    else:
        messages.error(request, 'A bobine selecionada não existe.')
        #  erro = 2
        #  return redirect('producao:addbobinepaleteerro', pk=palete.pk, e=erro)
        return redirect('producao:addbobinepalete', pk=palete.pk)


@login_required
def reabrir_palete(request, pk):
    palete = get_object_or_404(Palete, pk=pk)
    etiqueta = EtiquetaPalete.objects.get(palete=palete)

    etiqueta.produto = ""
    etiqueta.diam_max = 0
    etiqueta.diam_min = 0
    etiqueta.save()

    return redirect('producao:addbobinepalete', pk=palete.pk)


@login_required
def add_bobine_palete_erro(request, pk, e):
    template_name = 'palete/add_bobine_erro.html'
    palete = Palete.objects.get(pk=pk)

    erro = e

    context = {"palete": palete, "erro": erro}

    return render(request, template_name, context)


@login_required
def perfil_delete(request, pk):
    obj = get_object_or_404(Perfil, pk=pk)
    bobinagem = Bobinagem.objects.filter(perfil=pk)
    if request.method == "POST":
        obj.delete()
        return redirect('producao:perfil')

    context = {
        "object": obj,
    }
    return render(request, "perfil/perfil_delete.html", context)


@login_required
def bobinagem_delete(request, pk):
    obj = get_object_or_404(Bobinagem, pk=pk)
    bobine = Bobine.objects.filter(bobinagem=obj)
    template_name = "producao/bobinagem_delete.html"
    emenda = Emenda.objects.filter(bobinagem=obj)
    pal = False
    if request.method == "POST":
        for b in bobine:
            if b.palete != None:
                pal = True
                break
            else:
                pal = False

        if pal == False:
            if obj.perfil.retrabalho == True:
                # emenda = Emenda.objects.filter(bobinagem=obj)
                for e in emenda:
                    bobine = Bobine.objects.get(pk=e.bobine.pk)
                    bobine.comp_actual += e.metros
                    if bobine.recycle == True:
                        bobine.recycle = False
                    bobine.save()
                    e.delete()
                etiquetas = EtiquetaRetrabalho.objects.filter(bobinagem=obj)
                for eti in etiquetas:
                    eti.delete()
                obj.delete()
                if obj.perfil.retrabalho == False:
                    return redirect('producao:bobinagem_list_v3')
                else:
                    return redirect('producao:bobinagem_retrabalho_list_v2')
            else:
                etiquetas = EtiquetaRetrabalho.objects.filter(bobinagem=obj)
                for eti in etiquetas:
                    eti.delete()
                obj.delete()
                if obj.perfil.retrabalho == False:
                    return redirect('producao:bobinagem_list_v3')
                else:
                    return redirect('producao:bobinagem_retrabalho_list_v2')

        else:
            messages.error(
                request, 'A bobinagem não pode ser apagada porque existem bobines atribuidas a paletes.')

    context = {
        "object": obj,
        "bobine": bobine,
        "emenda": emenda
    }
    return render(request, template_name, context)


@login_required
def palete_delete(request, pk):
    obj = get_object_or_404(Palete, pk=pk)
    bobine = Bobine.objects.filter(palete=obj)
    if EtiquetaPalete.objects.get(palete=obj):
        e_p = EtiquetaPalete.objects.get(palete=obj)

    if request.method == "POST":
        num_bobines = Bobine.objects.filter(palete=obj).count()
        if obj.estado == 'G':
            if num_bobines != 0:
                messages.error(
                    request, 'A Palete não pode ser apagada porque já lhe foram atribuidas bobines. Para apagar esta palete, é necessário refazer a mesma.')
            elif obj.ordem != None and obj.carga == None:
                ordem = OrdemProducao.objects.get(pk=obj.ordem.pk)
                try:
                    enc = Encomenda.objects.get(pk=ordem.enc.pk)
                    enc.num_paletes_actual -= 1
                    ordem.num_paletes_produzidas -= 1
                    if ordem.ativa == False and ordem.completa == True:
                        ordem.ativa = True
                        ordem.completa = False
                    enc.save()
                    ordem.save()
                    obj.delete()
                    if e_p:
                        e_p.delete()
                    return redirect('producao:palete_list_all')
                except:
                    ordem.num_paletes_produzidas -= 1
                    if ordem.ativa == False and ordem.completa == True:
                        ordem.ativa = True
                        ordem.completa = False
                    ordem.save()
                    obj.delete()
                    if e_p:
                        e_p.delete()    
                    return redirect('producao:palete_list_all')
            elif obj.ordem == None and obj.carga == None:
                obj.delete()
                if e_p:
                    e_p.delete()
                return redirect('producao:palete_list_all')
        else:
            if num_bobines != 0:
                messages.error(
                    request, 'A Palete não pode ser apagada porque já lhe foram atribuidas bobines. Para apagar esta palete, é necessário refazer a mesma.')
            else:
                obj.delete()
                if e_p:
                    e_p.delete()
                return redirect('producao:paletes_retrabalho')

    context = {
        "object": obj,
        "bobine": bobine
    }
    return render(request, "palete/palete_delete.html", context)


@login_required
def status_bobinagem(request, operation, pk):
    bobinagem = Bobinagem.objects.get(pk=pk)
    if operation == 'ap':
        bobinagem.estado = 'G'
        bobinagem.save()
        num = 1
        for i in range(bobinagem.perfil.num_bobines):
            largura = Largura.objects.get(
                perfil=bobinagem.perfil, num_bobine=num)
            bobine = Bobine.objects.get(bobinagem=bobinagem, largura=largura)
            if bobine.estado == 'LAB' or bobine.estado == 'HOLD':
                bobine.estado = 'G'
                bobine.save()
            num += 1
        areas(pk)
    elif operation == 'rej':
        bobinagem.estado = 'R'
        bobinagem.save()
        num = 1
        for i in range(bobinagem.perfil.num_bobines):
            largura = Largura.objects.get(
                perfil=bobinagem.perfil, num_bobine=num)
            bobine = Bobine.objects.get(bobinagem=bobinagem, largura=largura)
            if bobine.estado == 'LAB' or bobine.estado == 'HOLD':
                bobine.estado = 'R'
                bobine.save()
            num += 1
        areas(pk)
    elif operation == 'dm':
        bobinagem.estado = 'DM'
        bobinagem.save()
        num = 1
        for i in range(bobinagem.perfil.num_bobines):
            largura = Largura.objects.get(
                perfil=bobinagem.perfil, num_bobine=num)
            bobine = Bobine.objects.get(bobinagem=bobinagem, largura=largura)
            if bobine.estado == 'LAB' or bobine.estado == 'HOLD':
                bobine.estado = 'DM'
                bobine.save()
            num += 1
        areas(pk)
    elif operation == 'hold':
        bobinagem.estado = 'hold'
        bobinagem.save()
        num = 1
        for i in range(bobinagem.perfil.num_bobines):
            largura = Largura.objects.get(
                perfil=bobinagem.perfil, num_bobine=num)
            bobine = Bobine.objects.get(bobinagem=bobinagem, largura=largura)
            if bobine.estado == 'LAB' or bobine.estado == 'DM':
                bobine.estado = 'HOLD'
                bobine.save()
            num += 1
        areas(pk)

    return redirect('producao:bobinestatus', pk=bobinagem.pk)


@login_required
def palete_retrabalho(request):
    palete_list = Palete.objects.filter(
        estado='DM').order_by('-data_pal', '-num')
    template_name = 'palete/palete_retrabalho.html'

    query = ""
    if request.GET:
        query = request.GET.get('q', '')
        # print(query)
        palete_list = Palete.objects.filter(
            nome__icontains=query, estado='DM').order_by('-data_pal', '-num')

    paginator = Paginator(palete_list, 11)
    page = request.GET.get('page')

    try:
        palete = paginator.page(page)
    except PageNotAnInteger:
        palete = paginator.page(1)
    except EmptyPage:
        palete = paginator.page(paginator.num_pages)

    context = {
        "palete": palete,
        "query": query,

    }
    return render(request, template_name, context)


@login_required
def palete_create_retrabalho(request):
    palete = Palete.objects.create(
        estado="DM", num_bobines=0, largura_bobines=0, diametro=0, core_bobines='3', user=request.user)
    palete.save()
    return redirect('producao:addbobinepalete', pk=palete.pk)


@login_required
def retrabalho_home(request):
    palete = Palete.objects.filter(estado='DM')
    bobine = Bobine.objects.all()
    bobinagem = Bobinagem.objects.all()
    template_name = 'retrabalho/retrabalho_home.html'

    context = {
        "palete": palete,
        "bobine": bobine,
        "bobinagem": bobinagem,
    }
    return render(request, template_name, context)


@login_required
def retrabalho_filter(request, pk):
    # palete = Palete.objects.filter(estado="DM")
    # bobine = Bobine.objects.all()
    bobinagem = Bobinagem.objects.get(pk=pk)
    # q_largura = request.GET.get("l")
    # if q_largura:
    #     palete = palete.filter(largura_bobines__gte=q_largura)
    # bobinagem = Bobinagem.objects.get(pk=pk)
    form = EmendasCreateForm

    if request.method == "POST":
        form = EmendasCreateForm(request.POST)

        if form.is_valid():
            bobine_ori = int(form['bobine'].value())
            comp_bobine_ori = int(form['metros'].value())
            bobine = Bobine.objects.get(pk=bobine_ori)
            if comp_bobine_ori > bobine.comp_actual:
                messages.error(
                    request, 'A bobine selecionada não tem comprimento suficiente para efectuar esta operação.')
                # redirect('producao:retrabalho_filter', pk=bobinagem.pk)
            else:
                emenda = Emenda.objects.create(
                    **form.cleaned_data, bobinagem=bobinagem)
                bobine.comp_actual = bobine.comp_actual - emenda.metros
                bobine.save()
                emenda.num_emenda = bobinagem.num_emendas + 1
                emenda.save()
                emenda_num = emenda.num_emenda
                print(emenda_num)
                if emenda_num == 0:
                    emenda.emenda = 0
                    emenda.save()
                elif emenda_num > 0:
                    if emenda_num == 1:
                        emenda.emenda = emenda.metros
                        emenda.save()
                    elif emenda_num == 2:
                        emenda_ul_bob = Emenda.objects.get(
                            bobinagem=bobinagem, num_emenda=1)
                        emenda.emenda = emenda_ul_bob.emenda + emenda.metros
                        emenda.save()
                    elif emenda_num == 3:
                        emenda_ul_bob = Emenda.objects.get(
                            bobinagem=bobinagem, num_emenda=2)
                        emenda.emenda = emenda_ul_bob.emenda + emenda.metros
                        emenda.save()

                bobinagem.num_emendas += 1
                data = bobinagem.data
                data = data.strftime('%Y%m%d')
                map(int, data)
                if bobinagem.perfil.retrabalho == True and bobinagem.num_emendas > 1:
                    if bobinagem.num_bobinagem < 10:
                        # instance.nome = '3%s-0%s' % (data, instance.num_bobinagem)
                        bobinagem.nome = '3%s-0%s' % (
                            data[1:], bobinagem.num_bobinagem)
                        bobinagem_pk = bobinagem.pk
                        bobine_nome(bobinagem_pk)
                    else:
                        bobinagem.nome = '3%s-0%s' % (
                            data[1:], bobinagem.num_bobinagem)
                elif bobinagem.perfil.retrabalho == True and bobinagem.num_emendas == 0:
                    if bobinagem.num_bobinagem < 10:
                        bobinagem.nome = '4%s-0%s' % (
                            data[1:], bobinagem.num_bobinagem)
                    else:
                        bobinagem.nome = '4%s-%s' % (data[1:],
                                                     bobinagem.num_bobinagem)

                comp_parcial = int(form['metros'].value())
                bobinagem.comp = bobinagem.comp + comp_parcial
                bobinagem.save()

        return redirect('producao:retrabalho_filter', pk=bobinagem.pk)

    palete = Palete.objects.filter(estado="DM")
    bobine = Bobine.objects.all()
    emenda = Emenda.objects.filter(bobinagem=pk)

    q_largura = request.GET.get("l")

    if q_largura:
        palete = palete.filter(largura_bobines__gte=q_largura)

    template_name = 'retrabalho/retrabalho_inicio.html'
    context = {

        "palete": palete,
        "bobine": bobine,
        "form": form,
        "bobinagem": bobinagem,
        "emenda": emenda

    }

    return render(request, template_name, context)


# Funções de apoio

def bobine_nome(pk):
    bobinagem = Bobinagem.objects.get(pk=pk)
    bobine = Bobine.objects.filter(bobinagem=bobinagem)
    for b in bobine:
        data = bobinagem.data
        data = data.strftime('%Y%m%d')
        map(int, data)
        if bobinagem.num_bobinagem < 10 and b.largura.num_bobine < 10:
            b.nome = '3%s-0%s-0%s' % (data[1:],
                                      bobinagem.num_bobinagem, b.largura.num_bobine)
            b.save()
        elif bobinagem.num_bobinagem >= 10 and b.largura.num_bobine < 10:
            b.nome = '3%s-%s-0%s' % (data[1:],
                                     bobinagem.num_bobinagem, b.largura.num_bobine)
            b.save()
        elif bobinagem.num_bobinagem > 10 and b.largura.num_bobine >= 10:
            b.nome = '3%s-%s-%s' % (data[1:],
                                    bobinagem.num_bobinagem, b.largura.num_bobine)
            b.save()


def comprimento_bobine_original(pk):
    pass

# def finalizar_retrabalho(request, pk):
#     bobinagem = get_object_or_404(Bobinagem, pk=pk)
#     form = BobinagemCreateForm(request.POST or None, instance=bobinagem)
#     if form.is_valid():
#         bobinagem = form.save(commit=False)
#         bobinagem.save()
#         return redirect('producao:retrabalho_home')

#     template_name = 'retrabalho/retrabalho_finalizar.html'
#     context = {
#         "bobinagem": bobinagem,
#         "form": form
#     }

#     return render(request, template_name, context)
@login_required
def bobinagem_retrabalho_finalizar(request, pk):
    bobinagem = get_object_or_404(Bobinagem, pk=pk)
    template_name = 'retrabalho/retrabalho_finalizar.html'

    form = BobinagemFinalizarForm(request.POST or None, instance=bobinagem)

    if form.is_valid():
        instance = form.save(commit=False)

        fim = instance.fim
        fim = fim.strftime('%H:%M')
        inico = instance.inico
        inico = inico.strftime('%H:%M')
        (hf, mf) = fim.split(':')
        (hi, mi) = inico.split(':')
        if hf < hi:
            result = (int(hf) * 3600 + int(mf) * 60) - \
                (int(hi) * 3600 + int(mi) * 60) + 86400
        else:
            result = (int(hf) * 3600 + int(mf) * 60) - \
                (int(hi) * 3600 + int(mi) * 60)

        result_str = strftime("%H:%M", gmtime(result))
        instance.duracao = result_str
        instance.save()

        bobines = Bobine.objects.filter(bobinagem=bobinagem)
        for bob in bobines:
            bob.diam = instance.diam
            bob.save()

        return redirect('producao:etiqueta_retrabalho', pk=bobinagem.pk)

    context = {
        "bobinagem": bobinagem,
        "form": form
    }

    return render(request, template_name, context)


class BobinagemRetrabalhoFinalizar(LoginRequiredMixin, UpdateView):
    model = Bobinagem
    fields = ['fim', 'diam']
    template_name = 'retrabalho/retrabalho_finalizar.html'
    success_url = '/producao/etiqueta/retrabalho/{id}/'

    def form_valid(self, form):
        b = form.save(commit=False)
        bobinagem = Bobinagem.objects.get(pk=self.kwargs['pk'])
        diam = b.diam
        bobines = Bobine.objects.filter(bobinagem=bobinagem)
        for bob in bobines:
            bob.diam = diam
            bob.save()
        fim = b.fim
        fim = fim.strftime('%H:%M')
        inico = b.inico
        inico = inico.strftime('%H:%M')
        (hf, mf) = fim.split(':')
        (hi, mi) = inico.split(':')
        if hf < hi:
            result = (int(hf) * 3600 + int(mf) * 60) - \
                (int(hi) * 3600 + int(mi) * 60) + 86400
        else:
            result = (int(hf) * 3600 + int(mf) * 60) - \
                (int(hi) * 3600 + int(mi) * 60)

        result_str = strftime("%H:%M", gmtime(result))
        b.duracao = result_str

        # b.comp_cli = b.comp
        # bobine = Bobine.objects.filter(bobinagem=b.pk)
        # for bob in bobine:
        #     largura = bob.largura.largura
        #     bob.comp_actual = b.comp_cli
        #     largura = b.perfil.largura_bobinagem / 1000
        #     bob.area = largura * b.comp_cli
        #     cont = 0
        #     cont += largura
        #     bob.save()

        # b.area = b.comp_cli * cont

        b.save()

        return super(BobinagemRetrabalhoFinalizar, self).form_valid(form)


def finalizar_retarbalho(request, pk):
    bobinagem = get_object_or_404(Bobinagem, pk=pk)
    form = BobinagemCreateForm(request.POST or None, instance=bobinagem)
    if form.is_valid():
        bobinagem = form.save(commit=False)
        bobinagem.save()
        # tempo_duracao(bobinagem.pk)
        return redirect('producao:etiqueta_retrabalho', pk=bobinagem.pk)

    template_name = 'retrabalho/retrabalho_finalizar.html'
    context = {
        "bobinagem": bobinagem,
        "form": form
    }

    return render(request, template_name, context)


@login_required
def emenda_delete(request, pk):
    emenda = get_object_or_404(Emenda, pk=pk)
    bobine = get_object_or_404(Bobine, pk=emenda.bobine.pk)
    bobinagem = get_object_or_404(Bobinagem, pk=emenda.bobinagem.pk)
    bob = Bobine.objects.filter(bobinagem=bobinagem)
    if request.method == "POST":
        bobinagem.num_emendas -= 1
        bobinagem.comp -= emenda.metros
        bobine.comp_actual += emenda.metros
        data = bobinagem.data
        data = data.strftime('%Y%m%d')
        map(int, data)
        if bobinagem.perfil.retrabalho == True and bobinagem.num_emendas > 1:
            if bobinagem.num_bobinagem < 10:
                # instance.nome = '3%s-0%s' % (data, instance.num_bobinagem)
                bobinagem.nome = '3%s-0%s' % (data[1:],
                                              bobinagem.num_bobinagem)
                for b in bob:
                    if b.largura.num_bobine < 10:
                        b.nome = '%s-0%s' % (bobinagem.nome,
                                             b.largura.num_bobine)
                        b.save()
                    else:
                        b.nome = '%s-%s' % (bobinagem.nome,
                                            b.largura.num_bobine)
                        b.save()
            else:
                bobinagem.nome = '3%s-0%s' % (data[1:],
                                              bobinagem.num_bobinagem)
                for b in bob:
                    if b.largura.num_bobine < 10:
                        b.nome = '%s-0%s' % (bobinagem.nome,
                                             b.largura.num_bobine)
                        b.save()
                    else:
                        b.nome = '%s-%s' % (bobinagem.nome,
                                            b.largura.num_bobine)
                        b.save()

        elif bobinagem.perfil.retrabalho == True and (bobinagem.num_emendas == 0 or bobinagem.num_emendas == 1):
            if bobinagem.num_bobinagem < 10:
                bobinagem.nome = '4%s-0%s' % (data[1:],
                                              bobinagem.num_bobinagem)
                for b in bob:
                    if b.largura.num_bobine < 10:
                        b.nome = '%s-0%s' % (bobinagem.nome,
                                             b.largura.num_bobine)
                        b.save()
                    else:
                        b.nome = '%s-%s' % (bobinagem.nome,
                                            b.largura.num_bobine)
                        b.save()

            else:
                bobinagem.nome = '4%s-%s' % (data[1:], bobinagem.num_bobinagem)
                for b in bob:
                    if b.largura.num_bobine < 10:
                        b.nome = '%s-0%s' % (bobinagem.nome,
                                             b.largura.num_bobine)
                        b.save()
                    else:
                        b.nome = '%s-%s' % (bobinagem.nome,
                                            b.largura.num_bobine)
                        b.save()

        bobinagem.save()
        # bobinagem_pk = bobinagem.pk
        # bobine_nome(bobinagem_pk)
        bobine.save()
        emenda.delete()
        return redirect('producao:retrabalho_filter', pk=emenda.bobinagem.pk)

    template_name = 'retrabalho/emenda_delete.html'
    context = {
        "emenda": emenda,
        "bobinagem": bobinagem,
        "bobine": bobine
    }
    return render(request, template_name, context)


@login_required
def cliente_home(request):

    cliente = Cliente.objects.all().order_by('-timestamp')
    template_name = 'cliente/cliente_home.html'

    # query = ""
    # if request.GET:
    #     query = request.GET.get('q', '')
    #     # print(query)
    #     cliente_list = Cliente.objects.filter(
    #         nome__icontains=query).order_by('-timestamp')

    # paginator = Paginator(cliente_list, 12)
    # page = request.GET.get('page')

    # try:
    #     cliente = paginator.page(page)
    # except PageNotAnInteger:
    #     cliente = paginator.page(1)
    # except EmptyPage:
    #     cliente = paginator.page(paginator.num_pages)

    context = {
        "cliente": cliente,
        # "query": query,

    }
    return render(request, template_name, context)

    context = {
        "cliente": cliente
    }
    return render(request, template_name, context)


@login_required
def producao_home(request):
    template_name = 'producao/producao_home.html'
    context = {}

    return render(request, template_name, context)

@login_required
def planeamento_home(request):
    template_name = 'producao/planeamento_home.html'
    context = {}

    return render(request, template_name, context)


@login_required
def bobine_details(request, pk):
    bobine = get_object_or_404(Bobine, pk=pk)
    form = ImprimirEtiquetaBobine(request.POST or None)
    emenda = Emenda.objects.filter(bobinagem=bobine.bobinagem)
    etiqueta = get_object_or_404(EtiquetaRetrabalho, bobine=bobine.nome)
    movimentos_bobine = MovimentosBobines.objects.filter(bobine=bobine)

    if form.is_valid():
        impressora = form['impressora'].value()
        num_copias = int(form['num_copias'].value())
        etiqueta.impressora = impressora
        etiqueta.num_copias = num_copias
        etiqueta.estado_impressao = True
        etiqueta.save()

    template_name = 'producao/bobine_details.html'
    context = {
        "movimentos_bobine": movimentos_bobine,
        "bobine": bobine,
        "emenda": emenda,
        "form": form
    }
    return render(request, template_name, context)


@login_required
def relatorio_diario(request):
    inicio_data = request.GET.get("id")
    fim_data = request.GET.get("fd")
    inicio_hora = request.GET.get("fd")
    fim_hora = request.GET.get("fd")

    bobinagem = []

    area_g = 0
    area_r = 0
    area_dm = 0
    area_ind = 0
    area_ba = 0

    if inicio_data and fim_data:
        bobinagem = Bobinagem.objects.filter(
            data__range=(inicio_data, fim_data))
        for bob in bobinagem:
            area_g += bob.area_g
            area_r += bob.area_r
            area_dm += bob.area_dm
            area_ind += bob.area_ind
            area_ba += bob.area_ba

    template_name = 'relatorio/relatorio_diario_linha.html'
    context = {
        "bobinagem": bobinagem,
        "area_g": area_g,
        "area_dm": area_dm,
        "area_r": area_r,
        "area_ba": area_ba,
        "area_ind": area_ind,

    }

    return render(request, template_name, context)


@login_required
def relatorio_consumos(request):
    inicio_data = request.GET.get("id")
    fim_data = request.GET.get("fd")
    inicio_hora = request.GET.get("fd")
    fim_hora = request.GET.get("fd")

    bobinagem = []
    c_sup_total = 0
    c_inf_total = 0
    c_sup = 0
    c_inf = 0

    if inicio_data and fim_data:
        bobinagem = Bobinagem.objects.filter(
            data__range=(inicio_data, fim_data))

        for bob in bobinagem:
            if bob.perfil.retrabalho == False:
                c_sup = bob.nwsup
                c_inf = bob.nwinf
                c_sup_total += c_sup
                c_inf_total += c_inf

    template_name = 'relatorio/relatorio_consumos.html'
    context = {
        "bobinagem": bobinagem,
        "c_sup_total": c_sup_total,
        "c_inf_total": c_inf_total,

    }

    return render(request, template_name, context)


@login_required
def relatorio_paletes(request):
    inicio_data = request.GET.get("id")
    fim_data = request.GET.get("fd")
    inicio_hora = request.GET.get("fd")
    fim_hora = request.GET.get("fd")

    palete = []
    num_paletes = 0
    area_total = 0

    if inicio_data and fim_data:
        palete = Palete.objects.filter(
            data_pal__range=(inicio_data, fim_data), estado='G')

        for p in palete:
            area_total += p.area
            num_paletes += 1

    template_name = 'relatorio/relatorio_paletes.html'
    context = {
        "palete": palete,
        "num_paletes": num_paletes,
        "area_total": area_total,


    }

    return render(request, template_name, context)


@login_required
def relatorio_home(request):
    template_name = 'relatorio/relatorio_home.html'
    context = {}

    return render(request, template_name, context)


@login_required
def etiqueta_retrabalho(request, pk):
    bobinagem = Bobinagem.objects.get(pk=pk)
    bobine = Bobine.objects.filter(bobinagem=bobinagem)

    if EtiquetaRetrabalho.objects.filter(bobinagem=bobinagem).exists():

        return redirect('producao:bobinestatus', pk=bobinagem.pk)

    else:
        for b in bobine:
            artigo_cliente = ArtigoCliente.objects.get(
                cliente=b.largura.cliente, artigo=b.artigo)
            e_r = EtiquetaRetrabalho.objects.create(bobinagem=bobinagem, bobine=b.nome, data=bobinagem.data, produto=b.largura.designacao_prod if artigo_cliente.produto=='' or artigo_cliente.produto is None else artigo_cliente.produto, largura_bobinagem=bobinagem.perfil.largura_bobinagem,
                                                    largura_bobine=b.largura.largura, diam=bobinagem.diam, comp_total=bobinagem.comp_cli, area=b.area, cod_cliente=artigo_cliente.cod_client, artigo=b.artigo.des)
            if Emenda.objects.filter(bobinagem=bobinagem).exists():
                emenda = Emenda.objects.filter(bobinagem=bobinagem)
                for e in emenda:
                    if e.num_emenda == 1:
                        e_r.bobine_original_1 = e.bobine.nome
                        e_r.emenda1 = e.emenda
                        e_r.metros1 = e.metros
                    elif e.num_emenda == 2:
                        e_r.bobine_original_2 = e.bobine.nome
                        e_r.emenda2 = e.emenda
                        e_r.metros2 = e.metros
                    elif e.num_emenda == 3:
                        e_r.bobine_original_3 = e.bobine.nome
                        e_r.emenda3 = e.emenda
                        e_r.metros3 = e.metros

                e_r.save()
        if bobinagem.perfil.retrabalho == True:
            return redirect('producao:bobinestatus', pk=bobinagem.pk)
        else:
            return redirect('producao:bobinagem_list_v3')


@login_required
def etiqueta_palete(request, pk):
    palete = Palete.objects.get(pk=pk)
    bobine = Bobine.objects.filter(palete=palete)
    e_p = EtiquetaPalete.objects.get(palete=palete)
    d_min = 0
    d_max = 0
    e_p.produto = bobine[0].bobinagem.perfil.produto
    bobines = [None] * 61

    for b in bobine:
        pos = b.posicao_palete
        bobines[pos] = b.nome
        d = b.bobinagem.diam
        if d_max == 0:
            d_max = d
        elif d > d_max:
            d_max = d
        elif d_min == 0:
            d_min = d
        elif d < d_min:
            d_min = d

    e_p.bobine1 = bobines[1]
    e_p.bobine2 = bobines[2]
    e_p.bobine3 = bobines[3]
    e_p.bobine4 = bobines[4]
    e_p.bobine5 = bobines[5]
    e_p.bobine6 = bobines[6]
    e_p.bobine7 = bobines[7]
    e_p.bobine8 = bobines[8]
    e_p.bobine9 = bobines[9]
    e_p.bobine10 = bobines[10]
    e_p.bobine11 = bobines[11]
    e_p.bobine12 = bobines[12]
    e_p.bobine13 = bobines[13]
    e_p.bobine14 = bobines[14]
    e_p.bobine15 = bobines[15]
    e_p.bobine16 = bobines[16]
    e_p.bobine17 = bobines[17]
    e_p.bobine18 = bobines[18]
    e_p.bobine19 = bobines[19]
    e_p.bobine20 = bobines[20]
    e_p.bobine21 = bobines[21]
    e_p.bobine22 = bobines[22]
    e_p.bobine23 = bobines[23]
    e_p.bobine24 = bobines[24]
    e_p.bobine25 = bobines[25]
    e_p.bobine26 = bobines[26]
    e_p.bobine27 = bobines[27]
    e_p.bobine28 = bobines[28]
    e_p.bobine29 = bobines[29]
    e_p.bobine30 = bobines[30]
    e_p.bobine31 = bobines[31]
    e_p.bobine32 = bobines[32]
    e_p.bobine33 = bobines[33]
    e_p.bobine34 = bobines[34]
    e_p.bobine35 = bobines[35]
    e_p.bobine36 = bobines[36]
    e_p.bobine37 = bobines[37]
    e_p.bobine38 = bobines[38]
    e_p.bobine39 = bobines[39]
    e_p.bobine40 = bobines[40]
    e_p.bobine41 = bobines[41]
    e_p.bobine42 = bobines[42]
    e_p.bobine43 = bobines[43]
    e_p.bobine44 = bobines[44]
    e_p.bobine45 = bobines[45]
    e_p.bobine46 = bobines[46]
    e_p.bobine47 = bobines[47]
    e_p.bobine48 = bobines[48]
    e_p.bobine49 = bobines[49]
    e_p.bobine50 = bobines[50]
    e_p.bobine51 = bobines[51]
    e_p.bobine52 = bobines[52]
    e_p.bobine53 = bobines[53]
    e_p.bobine54 = bobines[54]
    e_p.bobine55 = bobines[55]
    e_p.bobine56 = bobines[56]
    e_p.bobine57 = bobines[57]
    e_p.bobine58 = bobines[58]
    e_p.bobine59 = bobines[59]
    e_p.bobine60 = bobines[60]
    e_p.diam_min = d_min
    e_p.diam_max = d_max
    e_p.save()

    return redirect('producao:addbobinepalete', pk=palete.pk)


@login_required
def error_500(request):
    data = {}
    return render(request, 'error/error_500.html', data)


@login_required
def ordenar_bobines(request, pk):
    template_name = "palete/ordenar_bobines_formset.html"
    OrdenarBobinesFormSet = modelformset_factory(
        Bobine, form=OrdenarBobines, extra=0)
    palete = Palete.objects.get(pk=pk)
    bobines = Bobine.objects.filter(palete=palete)
    formset = OrdenarBobinesFormSet(request.POST or None, queryset=bobines)
    if formset.is_valid():
        bobs = formset.save(commit=False)
        for bob in bobs:
            bob.save()

    context = {
        "formset": formset,
        "bobines": bobines
    }

    return render(request, template_name, context)


@login_required
def c_bobines(request, pk):
    template_name = "producao/classificacao_bobines_formset.html"
    ClassificacaoBobinesFormSet = modelformset_factory(
        Bobine, form=ClassificacaoBobines, extra=0)
    bobinagem = Bobinagem.objects.get(pk=pk)
    bobines = Bobine.objects.filter(bobinagem=bobinagem)
    formset = ClassificacaoBobinesFormSet(
        request.POST or None, queryset=bobines)
    if formset.is_valid():
        bobs = formset.save(commit=False)
        for bob in bobs:
            if bob.estado == 'G':
                print('ok')
            # bob.save()

    context = {
        "formset": formset,
        "bobines": bobines,
        "bobinagem": bobinagem

    }

    return render(request, template_name, context)


@login_required
def retrabalho(request):
    template_name = 'retrabalho/retrabalho_formset.html'
    RetrabalhoFormSet = modelformset_factory(
        Emenda, form=RetrabalhoForm, extra=3)
    formset = RetrabalhoFormSet(request.POST or None)

    for f in formset:
        bobine = f['bobine'].value()
        print(bobine)
        bobine = Bobine.objects.get(nome=bobine)
        f.bobine = bobine.id

        print(f.bobine)

    if formset.is_valid():
        retrabalho = formset.save(commit=False)
        for r in retrabalho:
            r.save()

    context = {
        "formset": formset,

    }

    return render(request, template_name, context)


def destruir_bobine(request, pk_bobinagem, pk_bobine):
    bobine = Bobine.objects.get(pk=pk_bobine)
    bobinagem = Bobinagem.objects.get(pk=pk_bobinagem)

    margem_min = bobine.bobinagem.comp_cli * Decimal('0.20')
    print(margem_min)
    if bobine.comp_actual < margem_min:
        bobine.recycle = True
        bobine.save()
        return redirect('producao:retrabalho_filter', pk=bobinagem.pk)
    else:
        return redirect('producao:retrabalho_filter', pk=bobinagem.pk)


@login_required
def bobinagem_list_all(request):

    bobinagem = Bobinagem.objects.all()

    template_name = 'producao/bobinagem_list_all.html'
    context = {

        "bobinagem": bobinagem,

    }

    return render(request, template_name, context)


@login_required
def bobinagem_list_all_historico(request):

    bobinagem = Bobinagem.objects.all()

    template_name = 'producao/bobinagem_list_all_historico.html'

    context = {

        "bobinagem": bobinagem,

    }

    return render(request, template_name, context)


@login_required
def palete_list_all(request):
    palete_list = Palete.objects.filter(
        estado='G').order_by('-data_pal', '-num')
    template_name = 'palete/palete_list.html'

    query = ""
    if request.GET:
        query = request.GET.get('q', '')
        # print(query)
        palete_list = Palete.objects.filter(
            nome__icontains=query, estado='G').order_by('-data_pal', '-num')

    paginator = Paginator(palete_list, 12)
    page = request.GET.get('page')

    try:
        palete = paginator.page(page)
    except PageNotAnInteger:
        palete = paginator.page(1)
    except EmptyPage:
        palete = paginator.page(paginator.num_pages)

    context = {
        "palete": palete,
        "query": query,

    }
    return render(request, template_name, context)


@login_required
def palete_list_all_historico(request):

    palete = Palete.objects.all()

    template_name = 'palete/palete_list_all_historico.html'
    context = {

        "palete": palete,

    }

    return render(request, template_name, context)


@login_required
def palete_details(request, pk):
    palete = Palete.objects.get(pk=pk)

    template_name = 'palete/palete_details.html'

    context = {
        "palete": palete,
    }

    return render(request, template_name, context)


@login_required
def palete_confirmation(request, pk, id_bobines):
    palete = Palete.objects.get(pk=pk)
    e_p = EtiquetaPalete.objects.get(palete=palete)
    bobines = id_bobines
    num = 1
    comp = 0
    area = 0
    bobines_array = bobines.split("-")
    del bobines_array[-1]
    for x in bobines_array:
        id_b = int(x)
        bobine = Bobine.objects.get(pk=id_b)
        if bobine.palete != None:
            palete_o = Palete.objects.get(id=bobine.palete.id)
            palete_o.area -= bobine.area
            palete_o.comp_total -= bobine.bobinagem.comp
            palete_o.num_bobines_act -= 1
            palete_o.num_bobines -= 1
            palete_o.save()
        bobine.palete = palete
        bobine.posicao_palete = num
        comp += bobine.comp_actual
        area += bobine.area
        bobine.save()
        num += 1

    palete.num_bobines_act = num - 1
    palete.area = area
    palete.comp_total = comp
    palete.save()

    d_min = 0
    d_max = 0

    b = int(bobines_array[0])
    bobine_produto = Bobine.objects.get(pk=b)
    e_p.produto = bobine_produto.largura.designacao_prod
    bobine_posicao = [None] * 61

    for x in bobines_array:
        id_b = int(x)
        bobine = Bobine.objects.get(pk=id_b)
        pos = bobine.posicao_palete
        bobine_posicao[pos] = bobine.nome
        d = bobine.bobinagem.diam

        if d_min == 0 and d_max == 0:
            d_min = d
            d_max = d
        elif d <= d_min:
            d_min = d
        elif d >= d_max:
            d_max = d

    e_p.bobine1 = bobine_posicao[1]
    e_p.bobine2 = bobine_posicao[2]
    e_p.bobine3 = bobine_posicao[3]
    e_p.bobine4 = bobine_posicao[4]
    e_p.bobine5 = bobine_posicao[5]
    e_p.bobine6 = bobine_posicao[6]
    e_p.bobine7 = bobine_posicao[7]
    e_p.bobine8 = bobine_posicao[8]
    e_p.bobine9 = bobine_posicao[9]
    e_p.bobine10 = bobine_posicao[10]
    e_p.bobine11 = bobine_posicao[11]
    e_p.bobine12 = bobine_posicao[12]
    e_p.bobine13 = bobine_posicao[13]
    e_p.bobine14 = bobine_posicao[14]
    e_p.bobine15 = bobine_posicao[15]
    e_p.bobine16 = bobine_posicao[16]
    e_p.bobine17 = bobine_posicao[17]
    e_p.bobine18 = bobine_posicao[18]
    e_p.bobine19 = bobine_posicao[19]
    e_p.bobine20 = bobine_posicao[20]
    e_p.bobine21 = bobine_posicao[21]
    e_p.bobine22 = bobine_posicao[22]
    e_p.bobine23 = bobine_posicao[23]
    e_p.bobine24 = bobine_posicao[24]
    e_p.bobine25 = bobine_posicao[25]
    e_p.bobine26 = bobine_posicao[26]
    e_p.bobine27 = bobine_posicao[27]
    e_p.bobine28 = bobine_posicao[28]
    e_p.bobine29 = bobine_posicao[29]
    e_p.bobine30 = bobine_posicao[30]
    e_p.bobine31 = bobine_posicao[31]
    e_p.bobine32 = bobine_posicao[32]
    e_p.bobine33 = bobine_posicao[33]
    e_p.bobine34 = bobine_posicao[34]
    e_p.bobine35 = bobine_posicao[35]
    e_p.bobine36 = bobine_posicao[36]
    e_p.bobine37 = bobine_posicao[37]
    e_p.bobine38 = bobine_posicao[38]
    e_p.bobine39 = bobine_posicao[39]
    e_p.bobine40 = bobine_posicao[40]
    e_p.bobine41 = bobine_posicao[41]
    e_p.bobine42 = bobine_posicao[42]
    e_p.bobine43 = bobine_posicao[43]
    e_p.bobine44 = bobine_posicao[44]
    e_p.bobine45 = bobine_posicao[45]
    e_p.bobine46 = bobine_posicao[46]
    e_p.bobine47 = bobine_posicao[47]
    e_p.bobine48 = bobine_posicao[48]
    e_p.bobine49 = bobine_posicao[49]
    e_p.bobine50 = bobine_posicao[50]
    e_p.bobine51 = bobine_posicao[51]
    e_p.bobine52 = bobine_posicao[52]
    e_p.bobine53 = bobine_posicao[53]
    e_p.bobine54 = bobine_posicao[54]
    e_p.bobine55 = bobine_posicao[55]
    e_p.bobine56 = bobine_posicao[56]
    e_p.bobine57 = bobine_posicao[57]
    e_p.bobine58 = bobine_posicao[58]
    e_p.bobine59 = bobine_posicao[59]
    e_p.bobine60 = bobine_posicao[60]
    e_p.diam_min = d_min
    e_p.diam_max = d_max
    e_p.save()

    for x in bobines_array:
        id_b = int(x)
        bobine = Bobine.objects.get(pk=id_b)
        if bobine.bobinagem.perfil.retrabalho == True:

            ano = palete.data_pal
            ano = ano.strftime('%Y')
            num = palete.num
            if num < 10:
                palete.nome = 'R000%s-%s' % (num, ano)
                e_p.palete_nome = palete.nome
            elif num < 100:
                palete.nome = 'R00%s-%s' % (num, ano)
                e_p.palete_nome = palete.nome
            elif num < 1000:
                palete.nome = 'R0%s-%s' % (num, ano)
                e_p.palete_nome = palete.nome
            else:
                palete.nome = 'R%s-%s' % (num, ano)
                e_p.palete_nome = palete.nome

            palete.save()
            e_p.save()
            break

    add_artigo_to_bobine(pk)

    return redirect('producao:addbobinepalete', pk=palete.pk)


@login_required
def palete_rabrir(request, pk):
    palete = Palete.objects.get(pk=pk)

    if palete.estado == 'G':
        bobines = Bobine.objects.filter(palete=palete)
        e_p = EtiquetaPalete.objects.get(palete=palete)
        e_p.diam_min = 0
        e_p.diam_max = 0
        if palete.estado == 'G':
            ano = palete.data_pal
            ano = ano.strftime('%Y')
            num = palete.num
            if num < 10:
                palete.nome = 'P000%s-%s' % (num, ano)
                e_p.palete_nome = palete.nome
            elif num < 100:
                palete.nome = 'P00%s-%s' % (num, ano)
                e_p.palete_nome = palete.nome
            elif num < 1000:
                palete.nome = 'P0%s-%s' % (num, ano)
                e_p.palete_nome = palete.nome
            else:
                palete.nome = 'P%s-%s' % (num, ano)
                e_p.palete_nome = palete.nome

        palete.num_bobines_act = 0
        palete.save()
        e_p.save()
        for b in bobines:
            b.palete = None
            b.save()

        return redirect('producao:palete_picagem_v3', pk=palete.pk)

    elif palete.estado == 'DM':
        bobines = Bobine.objects.filter(palete=palete)
        e_p = EtiquetaPalete.objects.get(palete=palete)
        for b in bobines:
            b.palete == None
            b.save()

        # e_p.delete()
        # palete.delete()

        return redirect('producao:palete_create_retrabalho')

# @login_required
# def picagem_palete_dm(request, pk):
#     palete = Palete.objects.get(pk=pk)
#     e_p = EtiquetaPalete.objects.get(palete=palete)

#     template_name = 'palete/palete_picagem_dm.html'


#     context = {
#         "palete":palete,
#         "e_p":e_p,
#     }

#     return render(request, template_name, context)

@login_required
def palete_picagem_dm(request, pk):
    palete = get_object_or_404(Palete, pk=pk)
    template_name = 'palete/palete_picagem_dm_v3.html'
    num_bobines = palete.num_bobines
    peso_cores = 0
    PicagemBobinesFormSet = formset_factory(PicagemBobines, extra=num_bobines)
    form = PaletePesagemDMForm(request.POST or None, instance=palete)

    if request.method == 'POST':
        formset = PicagemBobinesFormSet(request.POST)
        if formset.is_valid() and form.is_valid():
            instance = form.save(commit=False)
            count = 0
            array_bobines = []
            validation = True

            for f in formset:
                count += 1
                cd = f.cleaned_data
                b = cd.get('bobine')
                if b is not None:
                    try:
                        bobine = get_object_or_404(Bobine, nome=b)
                        array_bobines.append(bobine)

                        if bobine.palete:
                            if bobine.palete.estado == 'G':
                                messages.error(request, 'A bobine ' + b + ' encontra-se numa palete para cliente. Por favor verifique.')
                            elif bobine.palete.estado == 'DM':
                                validation = True

                    except:
                        messages.error(request, '(' + str(count) + ') A bobine ' + b + ' não existe.')
                        validation = False

                else:
                    messages.error(
                        request, '(' + str(count) + ') Por favor preencha a campo nº ' + str(count) + '.')
                    validation = False

            if len(array_bobines) == palete.num_bobines and validation == True:
                if len(array_bobines) > len(set(array_bobines)):
                    messages.error(request, 'A picagem contem bobines repetidas.')
                elif instance.peso_bruto <= 0 or instance.peso_bruto == None:
                    messages.error(request, 'O peso Bruto não pode ser igual ou inferior a 0.')
                elif instance.peso_palete == None:
                    messages.error(request, 'O peso da palete é de preenchimento obrigatório.')
                else:
                    c = 0
                    area_sum = 0
                    comp_total = 0

                    for y in array_bobines:
                        y_b = get_object_or_404(Bobine, nome=y)
                        if y_b.palete:
                            y_p = get_object_or_404(Palete, pk=y_b.palete.pk)
                            if y_p.estado == 'DM':
                                y_p.area -= Decimal(y_b.area)
                                y_p.comp_total -= Decimal(
                                    y_b.bobinagem.comp_cli)
                                y_p.num_bobines -= 1
                                y_p.num_bobines_act -= 1
                                y_p.save()

                    for ab in array_bobines:
                        c += 1
                        bob = get_object_or_404(Bobine, nome=ab)
                        bob.palete = palete
                        bob.posicao_palete = c
                        area_sum += bob.area
                        comp_total += bob.bobinagem.comp_cli
                        bob.save()
                        movimento_bobine = MovimentosBobines.objects.create(bobine=bob, palete=palete, timestamp=palete.timestamp, destino=bob.destino)

                    e_p = EtiquetaPalete.objects.get(palete=palete)
                    

                    palete.num_bobines_act = c
                    palete.area = area_sum
                    palete.comp_total = comp_total
                    palete.save()

                    bobines = Bobine.objects.filter(palete=palete)
                    bobines_nome = []
                    d_min = 0
                    d_max = 0
                    bobine_posicao = [None] * 61

                    for bn in bobines:
                        bobines_nome.append(bn.nome)
                        d = bn.bobinagem.diam
                        pos = bn.posicao_palete
                        bobine_posicao[pos] = bn.nome
                        if d_min == 0 and d_max == 0:
                            d_min = d
                            d_max = d
                        elif d <= d_min:
                            d_min = d
                        elif d >= d_max:
                            d_max = d

                    # bobine_produto = Bobine.objects.get(nome=bobines_nome[0])
                    # e_p.produto = bobine_produto.largura.designacao_prod

                    e_p.bobine1 = bobine_posicao[1]
                    e_p.bobine2 = bobine_posicao[2]
                    e_p.bobine3 = bobine_posicao[3]
                    e_p.bobine4 = bobine_posicao[4]
                    e_p.bobine5 = bobine_posicao[5]
                    e_p.bobine6 = bobine_posicao[6]
                    e_p.bobine7 = bobine_posicao[7]
                    e_p.bobine8 = bobine_posicao[8]
                    e_p.bobine9 = bobine_posicao[9]
                    e_p.bobine10 = bobine_posicao[10]
                    e_p.bobine11 = bobine_posicao[11]
                    e_p.bobine12 = bobine_posicao[12]
                    e_p.bobine13 = bobine_posicao[13]
                    e_p.bobine14 = bobine_posicao[14]
                    e_p.bobine15 = bobine_posicao[15]
                    e_p.bobine16 = bobine_posicao[16]
                    e_p.bobine17 = bobine_posicao[17]
                    e_p.bobine18 = bobine_posicao[18]
                    e_p.bobine19 = bobine_posicao[19]
                    e_p.bobine20 = bobine_posicao[20]
                    e_p.bobine21 = bobine_posicao[21]
                    e_p.bobine22 = bobine_posicao[22]
                    e_p.bobine23 = bobine_posicao[23]
                    e_p.bobine24 = bobine_posicao[24]
                    e_p.bobine25 = bobine_posicao[25]
                    e_p.bobine26 = bobine_posicao[26]
                    e_p.bobine27 = bobine_posicao[27]
                    e_p.bobine28 = bobine_posicao[28]
                    e_p.bobine29 = bobine_posicao[29]
                    e_p.bobine30 = bobine_posicao[30]
                    e_p.bobine31 = bobine_posicao[31]
                    e_p.bobine32 = bobine_posicao[32]
                    e_p.bobine33 = bobine_posicao[33]
                    e_p.bobine34 = bobine_posicao[34]
                    e_p.bobine35 = bobine_posicao[35]
                    e_p.bobine36 = bobine_posicao[36]
                    e_p.bobine37 = bobine_posicao[37]
                    e_p.bobine38 = bobine_posicao[38]
                    e_p.bobine39 = bobine_posicao[39]
                    e_p.bobine40 = bobine_posicao[40]
                    e_p.bobine41 = bobine_posicao[41]
                    e_p.bobine42 = bobine_posicao[42]
                    e_p.bobine43 = bobine_posicao[43]
                    e_p.bobine44 = bobine_posicao[44]
                    e_p.bobine45 = bobine_posicao[45]
                    e_p.bobine46 = bobine_posicao[46]
                    e_p.bobine47 = bobine_posicao[47]
                    e_p.bobine48 = bobine_posicao[48]
                    e_p.bobine49 = bobine_posicao[49]
                    e_p.bobine50 = bobine_posicao[50]
                    e_p.bobine51 = bobine_posicao[51]
                    e_p.bobine52 = bobine_posicao[52]
                    e_p.bobine53 = bobine_posicao[53]
                    e_p.bobine54 = bobine_posicao[54]
                    e_p.bobine55 = bobine_posicao[55]
                    e_p.bobine56 = bobine_posicao[56]
                    e_p.bobine57 = bobine_posicao[57]
                    e_p.bobine58 = bobine_posicao[58]
                    e_p.bobine59 = bobine_posicao[59]
                    e_p.bobine60 = bobine_posicao[60]
                    e_p.diam_min = d_min
                    e_p.diam_max = d_max
                    e_p.save()

                    lots = Bobine.objects.filter(palete=palete)
                    perfil_embalamento = get_object_or_404(PerfilEmbalamento, pk=instance.perfil_embalamento.pk)
                    for lot in lots:
                        core_largura = CoreLargura.objects.get(core=lot.largura.perfil.core, largura=lot.largura.largura)
                        peso_cores += core_largura.peso

                    instance.peso_liquido = instance.peso_bruto - int(instance.peso_palete) - perfil_embalamento.peso - peso_cores
                    instance.save()

                    return redirect('producao:addbobinepalete', pk=palete.pk)

    else:
        formset = PicagemBobinesFormSet()

    context = {
        "palete": palete,
        "formset": formset,
        "form":form
    }

    return render(request, template_name, context)


@login_required
def validate_palete_dm(request, pk, id_bobines):
    palete = Palete.objects.get(pk=pk)
    e_p = EtiquetaPalete.objects.get(palete=palete)
    bobines = id_bobines
    num = 1
    comp = 0
    area = 0
    bobines_array = bobines.split("-")
    del bobines_array[-1]
    for x in bobines_array:
        id_b = int(x)
        bobine = Bobine.objects.get(pk=id_b)
        bobine.palete = palete
        bobine.posicao_palete = num
        comp += bobine.comp_actual

        area += bobine.area
        bobine.save()
        num += 1

    palete.num_bobines_act = num - 1
    if palete.retrabalhada == True:
        palete.num_bobines = palete.num_bobines_act
    palete.area = area
    palete.comp_total = comp
    palete.save()

    d_min = 0
    d_max = 0

    b = int(bobines_array[0])
    bobine_produto = Bobine.objects.get(pk=b)
    e_p.produto = bobine_produto.bobinagem.perfil.produto
    bobine_posicao = [None] * 61

    for x in bobines_array:
        id_b = int(x)
        bobine = Bobine.objects.get(pk=id_b)
        pos = bobine.posicao_palete
        bobine_posicao[pos] = bobine.nome
        d = bobine.bobinagem.diam

        if d_min == 0 and d_max == 0:
            d_min = d
            d_max = d
        elif d <= d_min:
            d_min = d
        elif d >= d_max:
            d_max = d

    e_p.bobine1 = bobine_posicao[1]
    e_p.bobine2 = bobine_posicao[2]
    e_p.bobine3 = bobine_posicao[3]
    e_p.bobine4 = bobine_posicao[4]
    e_p.bobine5 = bobine_posicao[5]
    e_p.bobine6 = bobine_posicao[6]
    e_p.bobine7 = bobine_posicao[7]
    e_p.bobine8 = bobine_posicao[8]
    e_p.bobine9 = bobine_posicao[9]
    e_p.bobine10 = bobine_posicao[10]
    e_p.bobine11 = bobine_posicao[11]
    e_p.bobine12 = bobine_posicao[12]
    e_p.bobine13 = bobine_posicao[13]
    e_p.bobine14 = bobine_posicao[14]
    e_p.bobine15 = bobine_posicao[15]
    e_p.bobine16 = bobine_posicao[16]
    e_p.bobine17 = bobine_posicao[17]
    e_p.bobine18 = bobine_posicao[18]
    e_p.bobine19 = bobine_posicao[19]
    e_p.bobine20 = bobine_posicao[20]
    e_p.bobine21 = bobine_posicao[21]
    e_p.bobine22 = bobine_posicao[22]
    e_p.bobine23 = bobine_posicao[23]
    e_p.bobine24 = bobine_posicao[24]
    e_p.bobine25 = bobine_posicao[25]
    e_p.bobine26 = bobine_posicao[26]
    e_p.bobine27 = bobine_posicao[27]
    e_p.bobine28 = bobine_posicao[28]
    e_p.bobine29 = bobine_posicao[29]
    e_p.bobine30 = bobine_posicao[30]
    e_p.bobine31 = bobine_posicao[31]
    e_p.bobine32 = bobine_posicao[32]
    e_p.bobine33 = bobine_posicao[33]
    e_p.bobine34 = bobine_posicao[34]
    e_p.bobine35 = bobine_posicao[35]
    e_p.bobine36 = bobine_posicao[36]
    e_p.bobine37 = bobine_posicao[37]
    e_p.bobine38 = bobine_posicao[38]
    e_p.bobine39 = bobine_posicao[39]
    e_p.bobine40 = bobine_posicao[40]
    e_p.bobine41 = bobine_posicao[41]
    e_p.bobine42 = bobine_posicao[42]
    e_p.bobine43 = bobine_posicao[43]
    e_p.bobine44 = bobine_posicao[44]
    e_p.bobine45 = bobine_posicao[45]
    e_p.bobine46 = bobine_posicao[46]
    e_p.bobine47 = bobine_posicao[47]
    e_p.bobine48 = bobine_posicao[48]
    e_p.bobine49 = bobine_posicao[49]
    e_p.bobine50 = bobine_posicao[50]
    e_p.bobine51 = bobine_posicao[51]
    e_p.bobine52 = bobine_posicao[52]
    e_p.bobine53 = bobine_posicao[53]
    e_p.bobine54 = bobine_posicao[54]
    e_p.bobine55 = bobine_posicao[55]
    e_p.bobine56 = bobine_posicao[56]
    e_p.bobine57 = bobine_posicao[57]
    e_p.bobine58 = bobine_posicao[58]
    e_p.bobine59 = bobine_posicao[59]
    e_p.bobine60 = bobine_posicao[60]
    e_p.diam_min = d_min
    e_p.diam_max = d_max
    e_p.save()

    return redirect('producao:addbobinepalete', pk=palete.pk)


@login_required
def retrabalho_dm(request, pk):
    bobinagem = Bobinagem.objects.get(pk=pk)

    template_name = 'retrabalho/retrabalho_dm.html'

    context = {
        "bobinagem": bobinagem,

    }

    return render(request, template_name, context)


@login_required
def validate_bobinagem_dm(request, pk, id_bobines, metros, recycle):

    bobinagem = Bobinagem.objects.get(pk=pk)
    bobines = Bobine.objects.filter(bobinagem=bobinagem)

    bobines_originais = id_bobines
    metros_bobines = metros
    recycle_bobines = recycle
    comp_total = 0
    emendas = 0
    cont = 0
    recycle_value = 0
    bobines_array = bobines_originais.split("--")
    metros_array = metros_bobines.split("-")
    recycle_array = recycle_bobines.split("-")
    bobines_length = len(bobines_array)

    for x in bobines_array:
        bobine = Bobine.objects.get(nome=x)
        comp_actual = bobine.comp_actual
        comp_actual -= Decimal(metros_array[cont])
        bobine.comp_actual = comp_actual
        comp_total += Decimal(metros_array[cont])
        bobine.recycle = recycle_array[recycle_value].capitalize()
        bobine.save()
        recycle_value += 1
        cont += 1
        emendas += 1
        emenda = Emenda.objects.create(bobinagem=bobinagem, bobine=bobine, metros=Decimal(
            metros_array[cont-1]), emenda=comp_total, num_emenda=cont)
        emenda.save()

    bobinagem.num_emendas = emendas - 1
    bobinagem.comp = comp_total
    bobinagem.comp_cli = comp_total
    bobinagem.estado = 'G'
    largura_bobinagem = bobinagem.perfil.largura_bobinagem / 1000
    bobinagem.area = bobinagem.comp_cli * largura_bobinagem
    bobinagem.save()
    print(emendas)

    for b in bobines:
        b.comp_actual = bobinagem.comp_cli
        b.estado = 'G'
        largura = b.largura.largura / 1000
        area = largura * b.bobinagem.comp_cli
        b.area = area
        b.save()
        if emendas == 1 or emendas == 0:
            pass
        elif emendas > 1:
            data = bobinagem.data
            data = data.strftime('%Y%m%d')
            map(int, data)
            if bobinagem.num_bobinagem < 10:
                bobinagem.nome = '3%s-0%s' % (data[1:],
                                              bobinagem.num_bobinagem)
                bobinagem_pk = bobinagem.pk
                bobine_nome(bobinagem_pk)
                bobinagem.save()
            else:
                bobinagem.nome = '3%s-%s' % (data[1:], bobinagem.num_bobinagem)
                bobinagem_pk = bobinagem.pk
                bobine_nome(bobinagem_pk)
                bobinagem.save()

    return redirect('producao:finalizar_retrabalho', pk=bobinagem.pk)


@login_required
def refazer_bobinagem_dm(request, pk):
    bobinagem = Bobinagem.objects.get(pk=pk)
    bobines = Bobine.objects.filter(bobinagem=bobinagem)
    emendas = Emenda.objects.filter(bobinagem=bobinagem)

    data = bobinagem.data
    data = data.strftime('%Y%m%d')

    if bobinagem.num_bobinagem < 10:
        bobinagem.nome = '4%s-0%s' % (data[1:], bobinagem.num_bobinagem)
    else:
        bobinagem.nome = '4%s-%s' % (data[1:], bobinagem.num_bobinagem)

    num = 1
    for b in bobines:
        if num < 10:
            b.nome = '%s-0%s' % (bobinagem.nome, num)
        else:
            b.nome = '%s-%s' % (bobinagem.nome, num)
        num += 1
        b.area = 0
        b.comp_actual = 0
        b.save()

    for e in emendas:
        bobine_original = Bobine.objects.get(pk=e.bobine.pk)
        bobine_original.comp_actual += e.metros
        if bobine_original.recycle == True:
            bobine_original.recycle = False
        bobine_original.save()
        e.delete()

    etiquetas = EtiquetaRetrabalho.objects.filter(bobinagem=bobinagem)
    for eti in etiquetas:
        eti.delete()

    bobinagem.num_emendas = 0
    bobinagem.comp = 0
    bobinagem.comp_par = 0
    bobinagem.comp_cli = 0
    bobinagem.diam = 0
    bobinagem.area = 0
    bobinagem.fim = None
    bobinagem.duracao = None
    bobinagem.area_g = 0
    bobinagem.area_dm = 0
    bobinagem.area_r = 0
    bobinagem.area_ind = 0
    bobinagem.area_ba = 0
    bobinagem.save()
    return redirect('producao:retrabalho_v2', pk=bobinagem.pk)


@login_required
def delete_bobinagem_dm(request, pk):
    bobinagem = Bobinagem.objects.get(pk=pk)
    bobines = Bobines.objects.filter(bobinagem=bobinagem)
    emendas = Emenda.objects.filter(bobinagem=bobinagem)

    for b in bobines:
        b.delete()

    bobinagem.delete()
    pass


@login_required
def encomenda_list(request):

    #server = 'SRV-SAGE\SAGEX3' 
    #database = 'sagex3' 
    #username = 'X3_ELASTICTEK' 
    #password = '%ElAsTicT3k@2021!RePoRt' 
    #conn = pyodbc.connect('DRIVER={sql server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)

    server = 'SRV-SAGE-V12\SAGEX3' 
    #database = 'sagex3'
    database = 'sagex3'
    username = 'sa' 
    password = '#ElasticTek#2022#' 
    #username = 'X3_ELASTICTEK' 
    #password = '%ElAsTicT3k@2021!RePoRt' 
    conn = pyodbc.connect('DRIVER={/opt/microsoft/msodbcsql17/lib64/libmsodbcsql-17.9.so.1.1};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)

    cursor=conn.cursor()
    cursor.execute("select distinct e.ROWID, e.SOHNUM_0, e.ORDDAT_0, e.DEMDLVDAT_0, e.SHIDAT_0, e.EXTDLVDAT_0, e.ITMREF_0, a.ITMDES1_0, e.QTY_0, e.BPCORD_0, c.BPCNAM_0, b.GROPRI_0 from sagex3.ELASTICTEK.SORDERQ as e left join sagex3.ELASTICTEK.SORDERP as b on e.SOHNUM_0 = b.SOHNUM_0 left join sagex3.ELASTICTEK.ITMMASTER as a on e.ITMREF_0 = a.ITMREF_0 left join sagex3.ELASTICTEK.BPCUSTOMER as c on e.BPCORD_0 = c.BPCNUM_0 where e.ORDDAT_0 > '2021-01-01' order by e.ROWID desc;")
    result = cursor.fetchall()

    for r in result:
        if Encomenda.objects.filter(eef=r[1]).exists():
            encomenda = get_object_or_404(Encomenda, eef=r[1])
            
            if encomenda.data_encomenda == r[2] and encomenda.data_solicitada == r[3] and encomenda.data_expedicao == r[4] and encomenda.data_prevista_expedicao == r[5]:
                pass
            else:
                encomenda.data_encomenda = r[2] 
                encomenda.data_solicitada = r[3] 
                encomenda.data_expedicao = r[4] 
                encomenda.data_prevista_expedicao = r[5]
                encomenda.save()

            linhas = conn.cursor()
            linhas.execute("select distinct e.ROWID, e.SOHNUM_0, e.ORDDAT_0, e.DEMDLVDAT_0, e.SHIDAT_0, e.EXTDLVDAT_0, e.ITMREF_0, a.ITMDES1_0, e.QTY_0, e.BPCORD_0, c.BPCNAM_0, b.GROPRI_0 from sagex3.ELASTICTEK.SORDERQ as e left join sagex3.ELASTICTEK.SORDERP as b on e.SOHNUM_0 = b.SOHNUM_0 left join sagex3.ELASTICTEK.ITMMASTER as a on e.ITMREF_0 = a.ITMREF_0 left join sagex3.ELASTICTEK.BPCUSTOMER as c on e.BPCORD_0 = c.BPCNUM_0 where e.SOHNUM_0 = '" + r[1] + "' order by e.ROWID desc;")
            result_linhas = linhas.fetchall()

            for linha in result_linhas:
                try:
                    artigo = get_object_or_404(Artigo, cod=linha[6])
                    linha_encomenda = LinhaEncomenda.objects.get(encomenda=encomenda, artigo=artigo)
                    if linha_encomenda.qtd == linha[8] and linha_encomenda.prc == linha[11]:
                        pass
                    else:
                        encomenda.sqm -= linha_encomenda.qtd
                        linha_encomenda.qtd = linha[8] 
                        linha_encomenda.prc = linha[11]
                        linha_encomenda.save()
                        encomenda.sqm += linha_encomenda.qtd
                        encomenda.save()
                except:
                    pass
            
        else:
            try:
                linha_artigo = 1        
                cliente = get_object_or_404(Cliente, cod=r[9])
                nova_encomenda = Encomenda.objects.create(user=request.user, eef=r[1], data_encomenda=r[2], data_solicitada=r[3], data_expedicao=r[4], data_prevista_expedicao=r[5], cliente=cliente, sqm=0)
                prf = conn.cursor()
                prf.execute("select distinct o.SOHNUM_0, s.SOHNUM_0, s.PRFNUM_0 from sagex3.ELASTICTEK.SORDERQ as o left join sagex3.ELASTICTEK.SORDER as s on o.SOHNUM_0 = s.SOHNUM_0 where o.SOHNUM_0 = '" + nova_encomenda.eef + "';")
                result_prf = prf.fetchall()
                nova_encomenda.prf = result_prf[0][2]
                nova_encomenda.save()
                linhas = conn.cursor()
                linhas.execute("select distinct e.ROWID, e.SOHNUM_0, e.ORDDAT_0, e.DEMDLVDAT_0, e.SHIDAT_0, e.EXTDLVDAT_0, e.ITMREF_0, a.ITMDES1_0, e.QTY_0, e.BPCORD_0, c.BPCNAM_0, b.GROPRI_0 from sagex3.ELASTICTEK.SORDERQ as e left join sagex3.ELASTICTEK.SORDERP as b on e.SOHNUM_0 = b.SOHNUM_0 left join sagex3.ELASTICTEK.ITMMASTER as a on e.ITMREF_0 = a.ITMREF_0 left join sagex3.ELASTICTEK.BPCUSTOMER as c on e.BPCORD_0 = c.BPCNUM_0 where e.SOHNUM_0 = '" + r[1] + "' order by e.ROWID desc;")
                result_linhas = linhas.fetchall()
                for linha in result_linhas:
                    artigo = get_object_or_404(Artigo, cod=linha[6])
                    nova_linha = LinhaEncomenda.objects.create(encomenda=nova_encomenda, artigo=artigo, linha=linha_artigo, qtd = linha[8], prc=linha[11])
                    linha_artigo += 1
                    nova_encomenda.sqm += linha[8]
                    nova_encomenda.save()

                paletes = nova_encomenda.sqm / 5400
                paletes = round(paletes, 0) + 1
                if nova_encomenda.num_paletes == 0:
                    nova_encomenda.num_paletes = paletes
                    nova_encomenda.save()
            except:
                pass
        

    encomenda_list = Encomenda.objects.all().order_by('-eef')
    template_name = 'encomenda/encomenda_list.html'

    query = ""
    if request.GET:
        query = request.GET.get('q', '')
        encomenda_list = encomenda_list.filter(Q(eef__icontains=query) | Q(cliente__nome__icontains=query)).order_by('-eef')

    paginator = Paginator(encomenda_list, 12)
    page = request.GET.get('page')

    try:
        encomenda = paginator.page(page)
    except PageNotAnInteger:
        encomenda = paginator.page(1)
    except EmptyPage:
        encomenda = paginator.page(paginator.num_pages)

    context = {
        "encomenda": encomenda,
        "query": query,

    }
    return render(request, template_name, context)


@login_required
def encomenda_create(request):

    template_name = 'encomenda/encomenda_create.html'
    form = EncomendaCreateForm(request.POST or None)

    if form.is_valid():
        instance = form.save(commit=False)
        instance.user = request.user
        instance.save()

        return redirect('producao:encomenda_list')

    context = {
        "form": form
    }

    return render(request, template_name, context)


@login_required
def encomenda_detail(request, pk):
    enc = get_object_or_404(Encomenda, pk=pk)
    cargas = Carga.objects.filter(enc=enc)
    paletes = Palete.objects.filter(ordem__enc=enc)
    linhas = LinhaEncomenda.objects.filter(encomenda=enc)

    template_name = 'encomenda/encomenda_detail.html'

    context = {
        "enc": enc,
        "cargas": cargas,
        "paletes": paletes,
        "linhas": linhas
    }

    return render(request, template_name, context)


@login_required
def carga_list(request):
    carga_list = Carga.objects.raw(f"""
    SELECT pc.*,(select round((sum(nbobines_emendas)/sum(nbobines_real))*100,2) from producao_palete pl where pl.carga_id=pc.id) nbobines_emendas 
    FROM producao_carga pc
    where estado = %s order by data desc
    """,['I'])
    #carga_list = Carga.objects.filter(estado='I').order_by('-data')
    template_name = 'carga/carga_list.html'
    for carga in carga_list:
        num_paletes = Palete.objects.filter(carga=carga).count()
        if carga.num_paletes_actual != num_paletes:
            carga.num_paletes_actual = num_paletes
            carga.save()

    query = ""
    if request.GET:
        query = request.GET.get('q', '')
        carga_list = Carga.objects.raw(f"""
        SELECT pc.*,(select round((sum(nbobines_emendas)/sum(nbobines_real))*100,2) from producao_palete pl where pl.carga_id=pc.id) nbobines_emendas 
        FROM producao_carga pc
        where estado = %s { f"and carga like '%%{query}%%'" } order by data desc
        """,['I'])
        #carga_list = Carga.objects.filter(carga__icontains=query, estado='I').order_by('-data',)

    paginator = Paginator(carga_list, 12)
    page = request.GET.get('page')

    try:
        carga = paginator.page(page)
    except PageNotAnInteger:
        carga = paginator.page(1)
    except EmptyPage:
        carga = paginator.page(paginator.num_pages)

    context = {
        "carga": carga,
        "query": query,

    }
    return render(request, template_name, context)


@login_required
def carga_list_completa(request):
    carga_list = Carga.objects.raw(f"""
    SELECT pc.*,(select round((sum(nbobines_emendas)/sum(nbobines_real))*100,2) from producao_palete pl where pl.carga_id=pc.id) nbobines_emendas 
    FROM producao_carga pc
    where estado = %s order by data desc
    """,['C'])
    template_name = 'carga/carga_list_complete.html'

    query = ""
    if request.GET:
        query = request.GET.get('q', '')
        # print(query)
        carga_list = Carga.objects.raw(f"""
        SELECT pc.*,(select round((sum(nbobines_emendas)/sum(nbobines_real))*100,2) from producao_palete pl where pl.carga_id=pc.id) nbobines_emendas 
        FROM producao_carga pc
        where estado = %s { f"and carga like '%%{query}%%'" } order by data desc
        """,['C'])

    paginator = Paginator(carga_list, 12)
    page = request.GET.get('page')

    try:
        carga = paginator.page(page)
    except PageNotAnInteger:
        carga = paginator.page(1)
    except EmptyPage:
        carga = paginator.page(paginator.num_pages)

    context = {
        "carga": carga,
        "query": query,

    }
    return render(request, template_name, context)


@login_required
def carga_create(request):

    template_name = 'carga/carga_create.html'
    form = CargaCreateForm(request.POST or None)

    if form.is_valid():
        instance = form.save(commit=False)
        instance.user = request.user
        instance.sqm = 0
        num_carga = form.cleaned_data['num_carga']
        tipo = form.cleaned_data['tipo']
        enc = str(form.cleaned_data['enc'])
        encomenda = get_object_or_404(Encomenda, eef=enc)
        cargas = Carga.objects.all()
        prf = encomenda.prf

        if num_carga < 10:
            if tipo == "CONTENTOR":
                num_carga = str(num_carga)
                carga = prf + "-CON00" + num_carga + \
                    "-" + str(encomenda.cliente.abv)
            else:
                num_carga = str(num_carga)
                carga = prf + "-CAM00" + num_carga + \
                    "-" + str(encomenda.cliente.abv)
        elif num_carga < 100:
            if tipo == "CONTENTOR":
                num_carga = str(num_carga)
                carga = prf + "-CON0" + num_carga + \
                    "-" + str(encomenda.cliente.abv)
            else:
                num_carga = str(num_carga)
                carga = prf + "-CAM0" + num_carga + \
                    "-" + str(encomenda.cliente.abv)
        elif num_carga < 1000:
            if tipo == "CONTENTOR":
                num_carga = str(num_carga)
                carga = prf + "-CON" + num_carga + \
                    "-" + str(encomenda.cliente.abv)
            else:
                num_carga = str(num_carga)
                carga = prf + "-CAM" + num_carga + \
                    "-" + str(encomenda.cliente.abv)

        instance.carga = carga
        instance.eef = enc
        instance.save()

        encomenda.num_cargas_actual += 1
        encomenda.num_cargas += 1
        encomenda.save()

        return redirect('producao:carga_list')

    context = {
        "form": form
    }

    return render(request, template_name, context)

@login_required
def carga_reabir(request, pk):
    carga = get_object_or_404(Carga, pk=pk)
    if carga.estado == 'C':
        carga.estado = 'I'
        carga.save()

    return redirect('producao:carga_detail', pk=pk)


@login_required
def carga_detail(request, pk):
    carga = get_object_or_404(Carga, pk=pk)
    paletes = Palete.objects.filter(carga=carga).order_by('-num_palete_carga')

    relacoes = []
    relacoes_area = []
    realcoes_comp = []
    num = 1
    som = 0
    som_comp = 0
    som_area = 0
    for pal in paletes:
        rel = round((Decimal(pal.peso_liquido) / (Decimal(pal.area / 10))) * 100, 2)
        relacoes.append(rel)
        rel_area = round((Decimal(pal.peso_liquido) * 1000) / 100, 0)
        relacoes_area.append(rel_area)
        rel_comp = round((Decimal(rel_area) / ((Decimal(pal.num_bobines) * Decimal(pal.largura_bobines)) * Decimal(0.001))) * Decimal(pal.num_bobines), 2)
        realcoes_comp.append(rel_comp)

    form = ImprimirEtiquetaFinalPalete(request.POST or None)
    if form.is_valid():
        num_copias = int(form['num_copias'].value())
        for p in paletes:
            etiqueta = EtiquetaFinal.objects.get(palete=p, activa=True)
            etiqueta.impressora = 'ARMAZEM_CAB_SQUIX_6.3_200'
            etiqueta.num_copias = num_copias
            etiqueta.estado_impressao = True
            etiqueta.save()

    for p in paletes:
        som += (p.peso_liquido/(p.area/10))*100
        area = (p.peso_liquido * 1000) / 100
        som_area += area
        som_comp += (Decimal(area) / ((Decimal(p.num_bobines) *
                                       Decimal(p.largura_bobines)) * Decimal(0.001))) * (Decimal(p.num_bobines))

    if len(paletes) == 0:
        som = 0
    else:
        som = som / len(paletes)

    som = round(som, 2)
    som_comp = round(som_comp, 2)
    som_area = round(som_area, 2)

    data_inicial = 0
    data_final = 0
    l_real_max = 0
    l_real_min = 0
    l_real = 0

    for p in paletes:
        bobines = Bobine.objects.filter(palete=p)

        for b in bobines:
            data = b.bobinagem.data
            if data_inicial == 0 and data_final == 0:
                data_inicial = data
                data_final = data
            elif data <= data_inicial:
                data_inicial = data
            elif data >= data_final:
                data_final = data

            if b.l_real == None:
                l_real = 0
            else:
                l_real = b.l_real

            if l_real_max == 0 and l_real_min == 0:
                l_real_min = l_real
                l_real_max = l_real
            if l_real == 0:
                pass
            elif l_real <= l_real_min:
                l_real_min = l_real
            elif l_real >= l_real_max:
                l_real_max = l_real

    template_name = 'carga/carga_detail.html'
    context = {
        "carga": carga,
        "data_inicial": data_inicial,
        "data_final": data_final,
        "som": som,
        "som_comp": som_comp,
        "som_area": som_area,
        "form": form,
        "paletes": paletes,
        "relacoes": relacoes,
        "relacoes_area": relacoes_area,
        "realcoes_comp": realcoes_comp,
        "l_real_min": l_real_min,
        "l_real_max": l_real_max
    }

    return render(request, template_name, context)


@login_required
def carga_etiqueta_palete(request, pk):
    palete = get_object_or_404(Palete, pk=pk)
    etiqueta = EtiquetaFinal.objects.get(palete=palete, activa=True)
    template_name = 'producao/carga_etiqueta_palete.html'

    etiqueta.impressora = 'ARMAZEM_CAB_SQUIX_6.3_200'
    etiqueta.num_copias = 1
    etiqueta.estado_impressao = True
    etiqueta.save()

    context = {}

    return redirect('producao:carga_detail', pk=palete.carga_id)


@login_required
def carga_edit(request, pk):
    pass


@login_required
def armazem_home(request):
    template_name = 'producao/armazem_home.html'
    context = {}

    return render(request, template_name, context)


@login_required
def palete_selecao(request):
    template_name = 'palete/palete_selecao.html'
    form = SelecaoPaleteForm(request.POST or None)

    if request.method == 'POST':
        form = SelecaoPaleteForm(request.POST or None)
        if form.is_valid:
            palete_nome = form['palete'].value()
            palete = get_object_or_404(Palete, nome=palete_nome)
            return redirect('producao:palete_pesagem', pk=palete.pk)

    context = {
        "form": form
    }

    return render(request, template_name, context)


@login_required
def palete_pesagem(request, pk=None):
    instance = get_object_or_404(Palete, pk=pk)
    template_name = 'palete/palete_pesagem.html'
    palete = get_object_or_404(Palete, pk=pk)
    bobines = Bobine.objects.filter(palete=palete)
    form = PaletePesagemForm(request.POST or None, instance=instance)
    
    tem_artigo = True

    for b in bobines:
        if b.artigo == None:
            tem_artigo = False
            break
        else:
            tem_artigo = True

    if form.is_valid():
        if tem_artigo == False:
            messages.error(request, 'As bobines da palete ' + palete.nome +' não têm artigo atribuido. Por favor contacte o Administrador.')
        else:
            try:
                instance = form.save(commit=False)
                perfil_embalamento = get_object_or_404(PerfilEmbalamento, pk=instance.perfil_embalamento.pk)
                if instance.peso_bruto == 0:
                    messages.error(request, 'A palete ' + palete.nome + ' não pode ter peso bruto de 0Kg.')
                else:
                    core_largura = CoreLargura.objects.get(core=palete.core_bobines, largura=palete.largura_bobines)
                    peso_cores = core_largura.peso * int(palete.num_bobines)
                    instance.peso_liquido = instance.peso_bruto - int(instance.peso_palete) - perfil_embalamento.peso - peso_cores
                    instance.save()
                    return redirect('producao:palete_selecao')
            except:
                messages.error(request, 'A palete ' + palete.nome + ' não têm perfil de embalaemnto atribuido. Por favor contacte o Administrador.')

    context = {
        "form": form,
        "instance": instance,
        "temartigo": tem_artigo
    }

    return render(request, template_name, context)


@login_required
def palete_details_armazem(request, pk):

    palete = get_object_or_404(Palete, pk=pk)
    template_name = 'palete/palete_details_armazem.html'

    context = {
        "palete": palete,


    }

    return render(request, template_name, context)


@login_required
def stock_list(request):
    palete_list = Palete.objects.filter(
        stock=True).order_by('-data_pal', '-num')
    template_name = 'stock/stock_list.html'

    query = ""
    if request.GET:
        query = request.GET.get('q', '')
        # print(query)
        palete_list = Palete.objects.filter(
            nome__icontains=query, stock=True).order_by('-data_pal', '-num')

    paginator = Paginator(palete_list, 12)
    page = request.GET.get('page')

    try:
        palete = paginator.page(page)
    except PageNotAnInteger:
        palete = paginator.page(1)
    except EmptyPage:
        palete = paginator.page(paginator.num_pages)

    context = {
        "palete": palete,
        "query": query,

    }
    return render(request, template_name, context)
    paletes = Palete.objects.filter(stock=True)

    template_name = 'stock/stock_list.html'

    context = {
        'paletes': paletes,
    }

    return render(request, template_name, context)


@login_required
def stock_add_to_carga(request, pk=None):
    instance = get_object_or_404(Palete, pk=pk)
    template_name = 'stock/stock_add_carga.html'

    form = AddPalateStockForm(request.POST or None, instance=instance)
    if form.is_valid():
        instance = form.save(commit=False)
        instance.stock = False
        carga = Carga.objects.get(carga=instance.carga)
        carga.num_paletes_actual += 1

        paletes_carga_1 = Palete.objects.filter(carga=carga)
        cont = 0
        array_num_palete = []

        for p1 in paletes_carga_1:
            array_num_palete.append(p1.num_palete_carga)
            # array_num_palete[cont] = p1.num_palete_carga
            # cont += 1

        if len(array_num_palete) == 0:
            instance.num_palete_carga = 1
        else:
            array_num_palete.sort()
            cont2 = 0
            for a in array_num_palete:
                if a != cont2 + 1:
                    instance.num_palete_carga = cont2 + 1
                    break
                elif len(array_num_palete) == cont2 + 1:
                    instance.num_palete_carga = cont2 + 2
                    break
                cont2 += 1

        carga.sqm += instance.area
        carga.metros += instance.comp_total
        if carga.num_paletes_actual == carga.num_paletes:
            carga.estado = 'C'

        carga.save()
        instance.save()
        gerar_etiqueta_final(instance.pk)

        return redirect('producao:carga_detail', pk=carga.pk)

    context = {
        "form": form,
        "instance": instance,
    }

    return render(request, template_name, context)


@login_required
def acompanhamento_diario(request):

    form = AcompanhamentoDiarioSearchForm(request.POST or None)

    template_name = 'lab/acompanhamento_diario.html'

    if form.is_valid():
        instance = form.save(commit=False)

    context = {
        "form": form,
    }

    return render(request, template_name, context)

# @login_required
# def acompanhamento_diario_filter(request):


#     form = AcompanhamentoDiarioSearchForm(request.POST or None)

#     template_name = 'lab/acompanhamento_diario.html'

#     if form.is_valid():
#         instance = form.save(commit=False)
#         data_i = instance.data_inicio
#         hora_i = instance.hora_inicio
#         data_f = instance.data_fim
#         hora_f = instance.hora_fim
#         return redirect('producao:qualidade_home')

#     context = {
#         "form": form,
#     }


#     return render(request, template_name, context)


@login_required
def qualidade_home(request):
    template_name = 'lab/qualidade_home.html'
    context = {}

    return render(request, template_name, context)


@login_required
def retrabalho_v2(request, pk):
    instance = get_object_or_404(Bobinagem, pk=pk)
    form = RetrabalhoFormEmendas(request.POST or None)
    template_name = "retrabalho/retrabalho_create_v2.html"

    print("######################")
    print(form.is_valid())
    if form.is_valid():
        b_1 = form.cleaned_data['bobine_1']
        b_2 = form.cleaned_data['bobine_2']
        b_3 = form.cleaned_data['bobine_3']
        m_b_1 = form.cleaned_data['m_bobine_1']
        m_b_2 = form.cleaned_data['m_bobine_2']
        m_b_3 = form.cleaned_data['m_bobine_3']

        if b_1 and b_2 and b_3 and m_b_1 and m_b_2 and m_b_3:
            if Bobine.objects.filter(nome=b_1) and Bobine.objects.filter(nome=b_2) and Bobine.objects.filter(nome=b_3):
                b_1 = get_object_or_404(Bobine, nome=b_1)
                b_2 = get_object_or_404(Bobine, nome=b_2)
                b_3 = get_object_or_404(Bobine, nome=b_3)
                comp_actual_1 = b_1.comp_actual
                comp_actual_2 = b_2.comp_actual
                comp_actual_3 = b_3.comp_actual

                if comp_actual_1 >= m_b_1 and comp_actual_2 >= m_b_2 and comp_actual_3 >= m_b_3 and b_1.estado == 'DM' and b_2.estado == 'DM' and b_3.estado == 'DM':
                    if b_1 == b_2 == b_3:
                        comp_total = m_b_1 + m_b_2 + m_b_3
                        dif = comp_total - b_1.comp_actual
                        if comp_total > b_1.comp_actual:
                            messages.error(request, 'A bobine ' + b_1.nome +
                                           ' não tem metros suficentes para efectuar esta operação. O valor que introduziu excede o comprimento atual da bobine em ' + str(dif) + ' m.')
                        else:
                            return redirect('producao:retrabalho_confirmacao', pk=pk, b1=b_1.pk, m1=m_b_1, b2=b_2.pk, m2=m_b_2, b3=b_3.pk, m3=m_b_3)

                    elif b_1 == b_2:
                        comp_total = m_b_1 + m_b_2
                        dif = comp_total - b_1.comp_actual
                        if comp_total > b_1.comp_actual:
                            messages.error(request, 'A bobine ' + b_1.nome +
                                           ' não tem metros suficentes para efectuar esta operação. O valor que introduziu excede o comprimento atual da bobine em ' + str(dif) + ' m.')
                        else:
                            return redirect('producao:retrabalho_confirmacao', pk=pk, b1=b_1.pk, m1=m_b_1, b2=b_2.pk, m2=m_b_2, b3=b_3.pk, m3=m_b_3)

                    elif b_1 == b_3:
                        comp_total = m_b_1 + m_b_3
                        dif = comp_total - b_1.comp_actual
                        if comp_total > b_1.comp_actual:
                            messages.error(request, 'A bobine ' + b_1.nome +
                                           ' não tem metros suficentes para efectuar esta operação. O valor que introduziu excede o comprimento atual da bobine em ' + str(dif) + ' m.')
                        else:
                            return redirect('producao:retrabalho_confirmacao', pk=pk, b1=b_1.pk, m1=m_b_1, b2=b_2.pk, m2=m_b_2, b3=b_3.pk, m3=m_b_3)

                    elif b_2 == b_3:
                        comp_total = m_b_2 + m_b_3
                        dif = comp_total - b_2.comp_actual
                        if comp_total > b_2.comp_actual:
                            messages.error(request, 'A bobine ' + b_2.nome +
                                           ' não tem metros suficentes para efectuar esta operação. O valor que introduziu excede o comprimento atual da bobine em ' + str(dif) + ' m.')
                        else:
                            return redirect('producao:retrabalho_confirmacao', pk=pk, b1=b_1.pk, m1=m_b_1, b2=b_2.pk, m2=m_b_2, b3=b_3.pk, m3=m_b_3)
                    else:
                        return redirect('producao:retrabalho_confirmacao', pk=pk, b1=b_1.pk, m1=m_b_1, b2=b_2.pk, m2=m_b_2, b3=b_3.pk, m3=m_b_3)

                else:
                    if b_1.estado != 'DM':
                        messages.error(request, 'A bobine ' + b_1.nome + ' não pode ser usada para retrabalho porque o seu estado é ' +
                                       b_1.estado + '. Para poder ser retrabalhada o seu estado deverá ser DM.')

                    if b_2.estado != 'DM':
                        messages.error(request, 'A bobine ' + b_2.nome + ' não pode ser usada para retrabalho porque o seu estado é ' +
                                       b_2.estado + '. Para poder ser retrabalhada o seu estado deverá ser DM.')

                    if b_3.estado != 'DM':
                        messages.error(request, 'A bobine ' + b_3.nome + ' não pode ser usada para retrabalho porque o seu estado é ' +
                                       b_3.estado + '. Para poder ser retrabalhada o seu estado deverá ser DM.')

                    if comp_actual_1 < m_b_1:
                        dif = m_b_1 - comp_actual_1
                        messages.error(request, 'A bobine ' + b_1.nome +
                                       ' não tem metros suficentes para efectuar esta operação. O valor que introduziu excede o comprimento atual da bobine em ' + str(dif) + ' m.')

                    if comp_actual_2 < m_b_2:
                        dif = m_b_2 - comp_actual_2
                        messages.error(request, 'A bobine ' + b_2.nome +
                                       ' não tem metros suficentes para efectuar esta operação. O valor que introduziu excede o comprimento atual da bobine em ' + str(dif) + ' m.')

                    if comp_actual_3 < m_b_3:
                        dif = m_b_3 - comp_actual_3
                        messages.error(request, 'A bobine ' + b_3.nome +
                                       ' não tem metros suficentes para efectuar esta operação. O valor que introduziu excede o comprimento atual da bobine em ' + str(dif) + ' m.')

            else:
                if not Bobine.objects.filter(nome=b_1):
                    messages.error(request, 'A bobine ' + b_1 + ' não existe.')
                if not Bobine.objects.filter(nome=b_2):
                    messages.error(request, 'A bobine ' + b_2 + ' não existe.')
                if not Bobine.objects.filter(nome=b_3):
                    messages.error(request, 'A bobine ' + b_3 + ' não existe.')

        elif b_1 and b_2 and m_b_1 and m_b_2:
            if Bobine.objects.filter(nome=b_1) and Bobine.objects.filter(nome=b_2):
                b_1 = get_object_or_404(Bobine, nome=b_1)
                b_2 = get_object_or_404(Bobine, nome=b_2)

                comp_actual_1 = b_1.comp_actual
                comp_actual_2 = b_2.comp_actual

                if comp_actual_1 >= m_b_1 and comp_actual_2 >= m_b_2 and b_1.estado == 'DM' and b_2.estado == 'DM':
                    if b_1 == b_2:
                        comp_total = m_b_1 + m_b_2
                        dif = comp_total - b_1.comp_actual
                        if comp_total > b_1.comp_actual:
                            messages.error(request, 'A bobine ' + b_1.nome +
                                           ' não tem metros suficentes para efectuar esta operação. O valor que introduziu excede o comprimento atual da bobine em ' + str(dif) + ' m.')
                        else:
                            return redirect('producao:retrabalho_confirmacao', pk=pk, b1=b_1.pk, m1=m_b_1, b2=b_2.pk, m2=m_b_2, b3=None, m3=None)
                    else:
                        return redirect('producao:retrabalho_confirmacao', pk=pk, b1=b_1.pk, m1=m_b_1, b2=b_2.pk, m2=m_b_2, b3=None, m3=None)

                else:
                    if b_1.estado != 'DM':
                        messages.error(request, 'A bobine ' + b_1.nome + ' não pode ser usada para retrabalho porque o seu estado é ' +
                                       b_1.estado + '. Para poder ser retrabalhada o seu estado deverá ser DM.')

                    if b_2.estado != 'DM':
                        messages.error(request, 'A bobine ' + b_2.nome + ' não pode ser usada para retrabalho porque o seu estado é ' +
                                       b_2.estado + '. Para poder ser retrabalhada o seu estado deverá ser DM.')

                    if comp_actual_1 < m_b_1:
                        dif = m_b_1 - comp_actual_1
                        messages.error(request, 'A bobine ' + b_1.nome +
                                       ' não tem metros suficentes para efectuar esta operação. O valor que introduziu excede o comprimento atual da bobine em ' + str(dif) + ' m.')

                    if comp_actual_2 < m_b_2:
                        dif = m_b_2 - comp_actual_2
                        messages.error(request, 'A bobine ' + b_2.nome +
                                       ' não tem metros suficentes para efectuar esta operação. O valor que introduziu excede o comprimento atual da bobine em ' + str(dif) + ' m.')

            else:
                if not Bobine.objects.filter(nome=b_1):
                    messages.error(request, 'A bobine ' + b_1 + ' não existe.')
                if not Bobine.objects.filter(nome=b_2):
                    messages.error(request, 'A bobine ' + b_2 + ' não existe.')

        elif b_1 and m_b_1:
            if Bobine.objects.filter(nome=b_1):
                b_1 = get_object_or_404(Bobine, nome=b_1)
                comp_actual_1 = b_1.comp_actual

                if comp_actual_1 >= m_b_1 and b_1.estado == 'DM':
                    return redirect('producao:retrabalho_confirmacao', pk=pk, b1=b_1.pk, m1=m_b_1, b2=None, m2=None, b3=None, m3=None)

                else:
                    if b_1.estado != 'DM':
                        messages.error(request, 'A bobine ' + b_1.nome + ' não pode ser usada para retrabalho porque o seu estado é ' +
                                       b_1.estado + '. Para poder ser retrabalhada o seu estado deverá ser DM.')

                    if comp_actual_1 < m_b_1:
                        dif = m_b_1 - comp_actual_1
                        messages.error(request, 'A bobine ' + b_1.nome +
                                       ' não tem metros suficentes para efectuar esta operação. O valor que introduziu excede o comprimento atual da bobine em ' + str(dif) + ' m.')

            else:
                if not Bobine.objects.filter(nome=b_1):
                    messages.error(request, 'A bobine ' + b_1 + ' não existe.')

    context = {
        "form": form,
        "instance": instance,
    }
    return render(request, template_name, context)


@login_required
def retrabalho_confirmacao(request, pk, b1, m1, b2=None, m2=None, b3=None, m3=None):
    bobinagem = get_object_or_404(Bobinagem, pk=pk)
    bobines = Bobine.objects.filter(bobinagem=bobinagem)
    form = ConfirmReciclarForm(request.POST or None)
    m_1 = m1
    m_2 = m2
    m_3 = m3

    try:
        b_3 = Bobine.objects.get(pk=b3)
    except:
        b_3 = "N/A"
        m_3 = "N/A"

    try:
        b_2 = Bobine.objects.get(pk=b2)
    except:
        b_2 = "N/A"
        m_2 = "N/A"

    b_1 = get_object_or_404(Bobine, pk=b1)

    comp_total = comp_dm(b_1, m_1, b_2, m_2, b_3, m_3)
    area_total = Decimal(
        comp_total) * (Decimal(bobinagem.perfil.largura_bobinagem) * Decimal(0.001))
    area_total = round(area_total, 2)

    area_bobines = []
    for b in bobines:
        area_bobines.append(
            round(Decimal(comp_total) * (Decimal(b.largura.largura) * Decimal(0.001)), 2))

    if b_3 != "N/A" and b_2 != "N/A":
        e_1 = m_1
        e_2 = int(m_1) + int(m_2)
        e_3 = int(m_1) + int(m_2) + int(m_3)
        if b_1 == b_2 == b_3:
            mr_1 = int(b_1.comp_actual) - int(m_1) - int(m_2) - int(m_3)
            mr_2 = mr_1
            mr_3 = mr_1
        elif b_1 == b_2 != b_3:
            mr_1 = int(b_1.comp_actual) - int(m_1) - int(m_2)
            mr_2 = mr_1
            mr_3 = int(b_3.comp_actual) - int(m_3)
        elif b_1 == b_3 != b_2:
            mr_1 = int(b_1.comp_actual) - int(m_1) - int(m_3)
            mr_3 = mr_1
            mr_2 = int(b_2.comp_actual) - int(m_2)
        elif b_2 == b_3 != b_1:
            mr_1 = int(b_1.comp_actual) - int(m_1)
            mr_2 = int(b_2.comp_actual) - int(m_2) - int(m_3)
            mr_3 = mr_2
        else:
            mr_1 = int(b_1.comp_actual) - int(m_1)
            mr_2 = int(b_2.comp_actual) - int(m_2)
            mr_3 = int(b_3.comp_actual) - int(m_3)
    elif b_2 != "N/A" and b_3 == "N/A":
        e_1 = m_1
        e_2 = int(m_1) + int(m_2)
        e_3 = "N/A"
        if b_1 == b_2:
            mr_1 = int(b_1.comp_actual) - int(m_1) - int(m_2)
            mr_2 = mr_1
            mr_3 = "N/A"
        else:
            mr_1 = int(b_1.comp_actual) - int(m_1)
            mr_2 = int(b_2.comp_actual) - int(m_2)
            mr_3 = "N/A"

    else:
        e_1 = m_1
        mr_1 = int(b_1.comp_actual) - int(m_1)
        e_2 = "N/A"
        e_3 = "N/A"
        mr_2 = "N/A"
        mr_3 = "N/A"

    if form.is_valid():
        recycle_1 = form.cleaned_data['recycle_1']
        recycle_2 = form.cleaned_data['recycle_2']
        recycle_3 = form.cleaned_data['recycle_3']

        if b_1 != "N/A" and b_3 != "N/A" and b_2 != "N/A":

            emenda_1 = Emenda.objects.create(
                bobinagem=bobinagem, bobine=b_1, num_emenda=1, emenda=e_1, metros=m_1)
            emenda_2 = Emenda.objects.create(
                bobinagem=bobinagem, bobine=b_2, num_emenda=2, emenda=e_2, metros=m_2)
            emenda_3 = Emenda.objects.create(
                bobinagem=bobinagem, bobine=b_3, num_emenda=3, emenda=e_3, metros=m_3)

            bobinagem.num_emendas = 3
            bobinagem.comp = e_3
            bobinagem.comp_par = e_3
            bobinagem.comp_cli = e_3
            bobinagem.area = round(Decimal(
                e_3) * (Decimal(bobinagem.perfil.largura_bobinagem) * Decimal(0.001)), 2)
            bobinagem.area_g = bobinagem.area
            bobinagem.estado = 'G'
            bobinagem.save()
            b_1.comp_actual = mr_1
            b_2.comp_actual = mr_2
            b_3.comp_actual = mr_3
            b_1.save()
            b_2.save()
            b_3.save()
            # retrabalho_nome(bobinagem.pk, 3)
            data = bobinagem.data
            data = data.strftime('%Y%m%d')
            map(int, data)
            if bobinagem.num_bobinagem < 10:
                bobinagem.nome = '3%s-0%s' % (data[1:],
                                              bobinagem.num_bobinagem)
            else:
                bobinagem.nome = '3%s-%s' % (data[1:], bobinagem.num_bobinagem)
            bobinagem.save()
            for b in bobines:
                if b.largura.num_bobine < 10:
                    b.nome = '%s-0%s' % (bobinagem.nome, b.largura.num_bobine)
                else:
                    b.nome = '%s-%s' % (bobinagem.nome, b.largura.num_bobine)
                b.save()

        elif b_1 != "N/A" and b_2 != "N/A":
            emenda_1 = Emenda.objects.create(
                bobinagem=bobinagem, bobine=b_1, num_emenda=1, emenda=e_1, metros=m_1)
            emenda_2 = Emenda.objects.create(
                bobinagem=bobinagem, bobine=b_2, num_emenda=2, emenda=e_2, metros=m_2)
            bobinagem.num_emendas = 2
            bobinagem.comp = e_2
            bobinagem.comp_par = e_2
            bobinagem.comp_cli = e_2
            bobinagem.area = round(Decimal(
                e_2) * (Decimal(bobinagem.perfil.largura_bobinagem) * Decimal(0.001)), 2)
            bobinagem.area_g = bobinagem.area
            bobinagem.estado = 'G'
            bobinagem.save()
            b_1.comp_actual = mr_1
            b_2.comp_actual = mr_2
            b_1.save()
            b_2.save()

            # retrabalho_nome(bobinagem.pk, 2)
            data = bobinagem.data
            data = data.strftime('%Y%m%d')
            map(int, data)
            if bobinagem.num_bobinagem < 10:
                bobinagem.nome = '3%s-0%s' % (data[1:],
                                              bobinagem.num_bobinagem)
            else:
                bobinagem.nome = '3%s-%s' % (data[1:], bobinagem.num_bobinagem)
            bobinagem.save()
            for b in bobines:
                if b.largura.num_bobine < 10:
                    b.nome = '%s-0%s' % (bobinagem.nome, b.largura.num_bobine)

                else:
                    b.nome = '%s-%s' % (bobinagem.nome, b.largura.num_bobine)
                b.save()

        elif b_1 != "N/A":
            emenda_1 = Emenda.objects.create(
                bobinagem=bobinagem, bobine=b_1, num_emenda=1, emenda=e_1, metros=m_1)
            bobinagem.num_emendas = 1
            bobinagem.comp = e_1
            bobinagem.comp_par = e_1
            bobinagem.comp_cli = e_1
            bobinagem.area = round(Decimal(
                e_1) * (Decimal(bobinagem.perfil.largura_bobinagem) * Decimal(0.001)), 2)
            bobinagem.area_g = bobinagem.area
            bobinagem.estado = 'G'
            bobinagem.save()
            b_1.comp_actual = mr_1
            b_1.save()

        if recycle_1 == True and b_1 != "N/A":
            b_1.recycle = True
            a_1 = round(Decimal(b_1.comp_actual) * (Decimal(b_1.largura.largura) * Decimal(0.001)), 2)
            reciclado_status = addToReciclado({"qtd":a_1,"source":"bobine_retrabalho","timestamp":datetime.datetime.now(),"lote":b_1.nome,"unit":"m2","user_id":request.user.id},None )
            if reciclado_status == False:
                pass
            try:
                movimento_bobine = MovimentosBobines.objects.create(
                    bobine=b_1, palete=b_1.palete, timestamp=bobinagem.timestamp, destino="Reciclada")
            except:
                movimento_bobine = MovimentosBobines.objects.create(
                    bobine=b_1, timestamp=bobinagem.timestamp, destino="Reciclada")
            b_1.save()

        if recycle_2 == True and b_2 != "N/A":
            b_2.recycle = True
            a_2 = round(Decimal(b_2.comp_actual) * (Decimal(b_2.largura.largura) * Decimal(0.001)), 2)
            reciclado_status = addToReciclado({"qtd":a_2,"source":"bobine_retrabalho","timestamp":datetime.datetime.now(),"lote":b_2.nome,"unit":"m2","user_id":request.user.id},None )
            if reciclado_status == False:
                pass
            try:
                movimento_bobine = MovimentosBobines.objects.create(
                    bobine=b_2, palete=b_2.palete, timestamp=bobinagem.timestamp, destino="Reciclada")
            except:
                movimento_bobine = MovimentosBobines.objects.create(
                    bobine=b_2, timestamp=bobinagem.timestamp, destino="Reciclada")
            b_2.save()

        if recycle_3 == True and b_3 != "N/A":
            b_3.recycle = True
            a_3 = round(Decimal(b_3.comp_actual) * (Decimal(b_3.largura.largura) * Decimal(0.001)), 2)
            reciclado_status = addToReciclado({"qtd":a_3,"source":"bobine_retrabalho","timestamp":datetime.datetime.now(),"lote":b_3.nome,"unit":"m2","user_id":request.user.id},None )
            if reciclado_status == False:
                pass
            try:
                movimento_bobine = MovimentosBobines.objects.create(
                    bobine=b_3, palete=b_3.palete, timestamp=bobinagem.timestamp, destino="Reciclada")
            except:
                movimento_bobine = MovimentosBobines.objects.create(
                    bobine=b_3, timestamp=bobinagem.timestamp, destino="Reciclada")
            b_3.save()

        for bob in bobines:
            bob.comp_actual = bobinagem.comp_cli
            bob.comp = bobinagem.comp_cli
            bob.area = round(Decimal(bobinagem.comp_cli) *
                             (Decimal(bob.largura.largura) * Decimal(0.001)), 2)
            bob.estado = 'G'
            bob.save()

        return redirect('producao:finalizar_retrabalho', pk=bobinagem.pk)

    template_name = "retrabalho/retrabalho_confirmacao.html"

    context = {
        "bobinagem": bobinagem,
        "bobines": bobines,
        "b_1": b_1,
        "b_2": b_2,
        "b_3": b_3,
        "m_1": m_1,
        "m_2": m_2,
        "m_3": m_3,
        "comp_total": comp_total,
        "area_total": area_total,
        "area_bobines": area_bobines,
        "e_1": e_1,
        "e_2": e_2,
        "e_3": e_3,
        "mr_1": mr_1,
        "mr_2": mr_2,
        "mr_3": mr_3,
        "form": form,

    }
    return render(request, template_name, context)


@login_required
def palete_picagem(request, pk):
    palete = get_object_or_404(Palete, pk=pk)
    cliente = palete.cliente
    template_name = 'palete/palete_picagem_v2.html'
    num_bobines = palete.num_bobines
    PicagemBobinesFormSet = formset_factory(PicagemBobines, extra=num_bobines)

    if request.method == 'POST':
        formset = PicagemBobinesFormSet(request.POST)
        if formset.is_valid():
            cliente = 0
            count = 0
            array_bobines = []
            array_cores = []
            array_larguras = []
            array_produtos = []
            array_estados = []
            validation = True
            if palete.cliente != None:
                cliente = palete.cliente

            # Validação individual
            for f in formset:
                count += 1
                cd = f.cleaned_data
                b = cd.get('bobine')

                if b is not None:
                    try:
                        bobine = get_object_or_404(Bobine, nome=b)
                        array_bobines.append(bobine)
                        array_cores.append(bobine.bobinagem.perfil.core)
                        array_larguras.append(bobine.largura.largura)
                        array_produtos.append(bobine.designacao_prod.lower())
                        array_estados.append(bobine.estado)
                        if cliente != 0:
                            try:
                                artigo_cliente = ArtigoCliente.objects.get(
                                    cliente=cliente, artigo=bobine.artigo)
                                pass
                            except:
                                messages.error(request, '(' + str(count) + ') O artigo da bobine ' + bobine.nome + ' não esta associado ao cliente ' + cliente.nome)
                                validation = False

                        if bobine.estado != 'G' and bobine.estado != 'LAB' and bobine.estado != 'SC':
                            messages.error(request, '(' + str(count) + ') A bobine ' +
                                           bobine.nome + ' não está em estado GOOD, LAB ou SC.')
                            validation = False

                        if bobine.palete:
                            p_b = get_object_or_404(
                                Palete, pk=bobine.palete.pk)
                            if p_b.estado == 'G':
                                messages.error(request, '(' + str(count) + ') A bobine ' +
                                               bobine.nome + ' encontra-se noutra palete GOOD.')
                                validation = False

                        if cliente != 0:
                            if bobine.diam > cliente.limsup:
                                messages.error(request, '(' + str(count) + ') O diâmetro da bobine ' + bobine.nome + ' é superior ao limite máximo aceite pelo cliente ' + cliente.nome + '.')
                                validation = False
                            elif bobine.diam > cliente.limsup:
                                messages.error(request, '(' + str(count) + ') O diâmetro da bobine ' + bobine.nome + ' é inferior ao limite mínimo aceite pelo cliente ' + cliente.nome + '.')
                                validation = False

                        if bobine.largura.largura != palete.largura_bobines:
                            messages.error(request, '(' + str(count) + ') A largura de ' + bobine.nome + ' (' + str(bobine.largura.largura) + ' mm) não corresponde a largura definida na palete de ' + str(palete.largura_bobines) + ' mm.')
                            validation = False

                        if bobine.bobinagem.perfil.core != palete.core_bobines:
                            messages.error(request, '(' + str(count) + ') O core da ' + bobine.nome + ' (' + str(bobine.bobinagem.perfil.core) + '") não corresponde ao core definido na palete de ' + str(palete.core_bobines) + '".')
                            validation = False

                    except:
                        messages.error(
                            request, '(' + str(count) + ') A bobine ' + b + ' não existe.')
                        validation = False

                else:
                    messages.error(
                        request, '(' + str(count) + ') Por favor preencha a campo nº ' + str(count) + '.')
                    validation = False

            validation_estados = True
            for estado in array_estados:
                if estado == 'SC':
                    validation_estados = False
                    pass

            # Validação global -> remover validação global de core e largura
            if len(array_bobines) == palete.num_bobines and validation == True:
                if len(array_bobines) > len(set(array_bobines)):
                    messages.error(
                        request, 'A picagem contem bobines repetidas.')
                elif len(set(array_produtos)) != 1:
                    messages.error(
                        request, 'As bobines picadas são produtos diferentes. Para que a palete seja válida todas as bobines têm de ser o mesmo produto.')
                elif len(set(array_estados)) != 1 and validation_estados == False:
                    messages.error(
                        request, 'Todas as bobines tem de ser classificadas como SC antes de validar esta palete.')
                else:
                    c = 0
                    area_sum = 0
                    comp_total = 0

                    for y in array_bobines:
                        y_b = get_object_or_404(Bobine, nome=y)
                        if y_b.palete:
                            y_p = get_object_or_404(Palete, pk=y_b.palete.pk)
                            if y_p.estado == 'DM':
                                y_p.area -= Decimal(y_b.area)
                                y_p.comp_total -= Decimal(
                                    y_b.bobinagem.comp_cli)
                                y_p.num_bobines -= 1
                                y_p.num_bobines_act -= 1
                                y_p.save()

                    for ab in array_bobines:
                        c += 1
                        bob = get_object_or_404(Bobine, nome=ab)
                        bob.palete = palete
                        bob.posicao_palete = c
                        area_sum += bob.area
                        comp_total += bob.comp
                        bob.save()
                        movimento_bobine = MovimentosBobines.objects.create(
                            bobine=bob, palete=palete, timestamp=palete.timestamp, destino=bob.destino)

                    e_p = EtiquetaPalete.objects.get(palete=palete)
                    for x in array_bobines:
                        bobine = Bobine.objects.get(nome=x)
                        if bobine.bobinagem.perfil.retrabalho == True:
                            ano = palete.data_pal
                            ano = ano.strftime('%Y')
                            num = palete.num
                            if num < 10:
                                palete.nome = 'R000%s-%s' % (num, ano)
                            elif num < 100:
                                palete.nome = 'R00%s-%s' % (num, ano)
                            elif num < 1000:
                                palete.nome = 'R0%s-%s' % (num, ano)
                            else:
                                palete.nome = 'R%s-%s' % (num, ano)
                            palete.save()
                            e_p.palete_nome = palete.nome
                            break

                    palete.num_bobines_act = c
                    palete.area = area_sum
                    palete.comp_total = comp_total
                    palete.save()

                    bobines = Bobine.objects.filter(palete=palete)
                    bobines_nome = []
                    d_min = 0
                    d_max = 0
                    bobine_posicao = [None] * 61

                    for bn in bobines:
                        bobines_nome.append(bn.nome)
                        d = bn.diam
                        pos = bn.posicao_palete
                        bobine_posicao[pos] = bn.nome
                        if d_min == 0 and d_max == 0:
                            d_min = d
                            d_max = d
                        elif d <= d_min:
                            d_min = d
                        elif d >= d_max:
                            d_max = d

                    bobine_produto = Bobine.objects.get(nome=bobines_nome[0])
                    e_p.produto = bobine_produto.designacao_prod
                    e_p.artigo = bobine_produto.artigo.des
                    if cliente != 0:
                        artigo_cliente = ArtigoCliente.objects.get(
                            cliente=cliente, artigo=bobine_produto.artigo)
                        e_p.cod_cliente = artigo_cliente.cod_client

                    e_p.bobine1 = bobine_posicao[1]
                    e_p.bobine2 = bobine_posicao[2]
                    e_p.bobine3 = bobine_posicao[3]
                    e_p.bobine4 = bobine_posicao[4]
                    e_p.bobine5 = bobine_posicao[5]
                    e_p.bobine6 = bobine_posicao[6]
                    e_p.bobine7 = bobine_posicao[7]
                    e_p.bobine8 = bobine_posicao[8]
                    e_p.bobine9 = bobine_posicao[9]
                    e_p.bobine10 = bobine_posicao[10]
                    e_p.bobine11 = bobine_posicao[11]
                    e_p.bobine12 = bobine_posicao[12]
                    e_p.bobine13 = bobine_posicao[13]
                    e_p.bobine14 = bobine_posicao[14]
                    e_p.bobine15 = bobine_posicao[15]
                    e_p.bobine16 = bobine_posicao[16]
                    e_p.bobine17 = bobine_posicao[17]
                    e_p.bobine18 = bobine_posicao[18]
                    e_p.bobine19 = bobine_posicao[19]
                    e_p.bobine20 = bobine_posicao[20]
                    e_p.bobine21 = bobine_posicao[21]
                    e_p.bobine22 = bobine_posicao[22]
                    e_p.bobine23 = bobine_posicao[23]
                    e_p.bobine24 = bobine_posicao[24]
                    e_p.bobine25 = bobine_posicao[25]
                    e_p.bobine26 = bobine_posicao[26]
                    e_p.bobine27 = bobine_posicao[27]
                    e_p.bobine28 = bobine_posicao[28]
                    e_p.bobine29 = bobine_posicao[29]
                    e_p.bobine30 = bobine_posicao[30]
                    e_p.bobine31 = bobine_posicao[31]
                    e_p.bobine32 = bobine_posicao[32]
                    e_p.bobine33 = bobine_posicao[33]
                    e_p.bobine34 = bobine_posicao[34]
                    e_p.bobine35 = bobine_posicao[35]
                    e_p.bobine36 = bobine_posicao[36]
                    e_p.bobine37 = bobine_posicao[37]
                    e_p.bobine38 = bobine_posicao[38]
                    e_p.bobine39 = bobine_posicao[39]
                    e_p.bobine40 = bobine_posicao[40]
                    e_p.bobine41 = bobine_posicao[41]
                    e_p.bobine42 = bobine_posicao[42]
                    e_p.bobine43 = bobine_posicao[43]
                    e_p.bobine44 = bobine_posicao[44]
                    e_p.bobine45 = bobine_posicao[45]
                    e_p.bobine46 = bobine_posicao[46]
                    e_p.bobine47 = bobine_posicao[47]
                    e_p.bobine48 = bobine_posicao[48]
                    e_p.bobine49 = bobine_posicao[49]
                    e_p.bobine50 = bobine_posicao[50]
                    e_p.bobine51 = bobine_posicao[51]
                    e_p.bobine52 = bobine_posicao[52]
                    e_p.bobine53 = bobine_posicao[53]
                    e_p.bobine54 = bobine_posicao[54]
                    e_p.bobine55 = bobine_posicao[55]
                    e_p.bobine56 = bobine_posicao[56]
                    e_p.bobine57 = bobine_posicao[57]
                    e_p.bobine58 = bobine_posicao[58]
                    e_p.bobine59 = bobine_posicao[59]
                    e_p.bobine60 = bobine_posicao[60]
                    e_p.diam_min = d_min
                    e_p.diam_max = d_max
                    e_p.save()

                    return redirect('producao:palete_pesagem', pk=palete.pk)

    else:
        formset = PicagemBobinesFormSet()

    context = {
        "palete": palete,
        "formset": formset
    }
    return render(request, template_name, context)


@login_required
def classificacao_bobines_v2(request, pk):
    bobinagem = get_object_or_404(Bobinagem, pk=pk)
    bobines = Bobine.objects.filter(bobinagem=bobinagem)
    template_name = 'producao/classificacao_bobines_v2.html'

    # ClassificacaoBobinesFormSet = modelformset_factory(Bobine, fields=('estado', ))
    ClassificacaoBobinesFormSet = inlineformset_factory(Bobinagem, Bobine, fields=(
        'nome', 'estado', 'con', 'descen', 'presa', 'diam_insuf', 'furos', 'estado', 'buraco', 'esp', 'troca_nw', 'outros', 'obs', 'l_real', ), extra=0, can_delete=False)

    if request.method == 'POST':
        # formset = ClassificacaoBobinesFormSet(request.POST, queryset=Bobine.objects.filter(bobinagem=bobinagem),)
        formset = ClassificacaoBobinesFormSet(request.POST, instance=bobinagem)
        if formset.is_valid():
            #    instances = formset.save(commit=False)
            #    for instance in instances:
            #        instance.save()
            formset.save()

    formset = ClassificacaoBobinesFormSet(instance=bobinagem)

    context = {
        "bobinagem": bobinagem,
        "bobines": bobines,
        "formset": formset
    }
    return render(request, template_name, context)


@login_required
def bobinagem_list_v2(request):
    bobinagens = Bobinagem.objects.filter(
        perfil__retrabalho=0, data__gt='2018-12-31')
    bobine = Bobine.objects.filter(bobinagem__perfil__retrabalho=0)
    template_name = 'producao/bobinagem_list_v2.html'
    query = request.GET.get("q")
    if query:
        bobinagens = bobinagens.filter(nome__icontains=query)
    context = {
        "bobinagens": bobinagens,
        "bobine": bobine

    }
    return render(request, template_name, context)


@login_required
def perfil_list_v2(request):
    perfil_list = Perfil.objects.filter(obsoleto=False).order_by('-timestamp')
    template_name = 'perfil/perfil_list_v2.html'
    # form = SearchPerfil(request.POST or None)

    query = ""
    if request.GET:
        query = request.GET.get('q', '')
        # print(query)
        perfil_list = perfil_list.filter(nome__icontains=query).order_by('-timestamp')

    paginator = Paginator(perfil_list, 20)
    page = request.GET.get('page')

    try:
        perfil = paginator.page(page)
    except PageNotAnInteger:
        perfil = paginator.page(1)
    except EmptyPage:
        perfil = paginator.page(paginator.num_pages)

    context = {
        "perfil": perfil,
        "query": query,

    }
    return render(request, template_name, context)

    # if form.is_valid():
    #     cd = form.cleaned_data
    #     nome = cd.get('nome')
    #     num_bobines = cd.get('num_bobines')
    #     core = cd.get('core')
    #     largura_bobinagem = cd.get('largura_bobinagem')
    #     retrabalho = cd.get('retrabalho')

    #     if retrabalho == True:
    #         if nome and num_bobines and core and largura_bobinagem:
    #             perfil = Perfil.objects.filter(Q(nome__icontains=nome) & Q(num_bobines__iexact=num_bobines) & Q(core__iexact=core) & Q(largura_bobinagem__iexact=largura_bobinagem) & Q(retrabalho=True) & Q(obsoleto=False))
    #         elif nome and num_bobines and core:
    #             perfil = Perfil.objects.filter(Q(nome__icontains=nome) & Q(num_bobines__iexact=num_bobines) & Q(core__iexact=core) & Q(retrabalho=True) & Q(obsoleto=False))
    #         elif nome and num_bobines and largura_bobinagem:
    #             perfil = Perfil.objects.filter(Q(nome__icontains=nome) & Q(num_bobines__iexact=num_bobines) & Q(largura_bobinagem__iexact=largura_bobinagem) & Q(retrabalho=True) & Q(obsoleto=False))
    #         elif nome and core and largura_bobinagem:
    #             perfil = Perfil.objects.filter(Q(nome__icontains=nome) & Q(core__iexact=core) & Q(largura_bobinagem__iexact=largura_bobinagem) & Q(retrabalho=True) & Q(obsoleto=False))
    #         elif num_bobines and core and largura_bobinagem:
    #             perfil = Perfil.objects.filter(Q(num_bobines__iexact=num_bobines) & Q(core__iexact=core) & Q(largura_bobinagem__iexact=largura_bobinagem) & Q(retrabalho=True) & Q(obsoleto=False))
    #         elif nome and num_bobines:
    #             perfil = Perfil.objects.filter(Q(nome__icontains=nome) & Q(num_bobines__iexact=num_bobines) & Q(retrabalho=True) & Q(obsoleto=False))
    #         elif nome and largura_bobinagem:
    #             perfil = Perfil.objects.filter(Q(nome__icontains=nome) & Q(largura_bobinagem__iexact=largura_bobinagem) & Q(retrabalho=True) & Q(obsoleto=False))
    #         elif num_bobines and largura_bobinagem:
    #             perfil = Perfil.objects.filter(Q(num_bobines__iexact=num_bobines) & Q(largura_bobinagem__iexact=largura_bobinagem) & Q(retrabalho=True) & Q(obsoleto=False))
    #         elif core and largura_bobinagem:
    #             perfil = Perfil.objects.filter(Q(core__iexact=core) & Q(largura_bobinagem__iexact=largura_bobinagem) & Q(retrabalho=True) & Q(obsoleto=False))
    #         elif nome and core:
    #             perfil = Perfil.objects.filter(Q(nome__icontains=nome) & Q(core__iexact=core) & Q(retrabalho=True) & Q(obsoleto=False))
    #         elif num_bobines and core:
    #             perfil = Perfil.objects.filter(Q(num_bobines__iexact=num_bobines) & Q(core__iexact=core) & Q(retrabalho=True) & Q(obsoleto=False))
    #         elif nome:
    #             perfil = Perfil.objects.filter(Q(nome__icontains=nome) & Q(retrabalho=True) & Q(obsoleto=False))
    #         elif num_bobines:
    #             perfil = Perfil.objects.filter(Q(num_bobines__iexact=num_bobines) & Q(retrabalho=True)) & Q(obsoleto=False)
    #         elif core:
    #             perfil = Perfil.objects.filter(Q(core__iexact=core) & Q(retrabalho=True) & Q(obsoleto=False))
    #         elif largura_bobinagem:
    #             perfil = Perfil.objects.filter(Q(largura_bobinagem__iexact=largura_bobinagem) & Q(retrabalho=True) & Q(obsoleto=False))
    #         else:
    #             perfil = Perfil.objects.filter(Q(retrabalho=True) & Q(obsoleto=False))

    #     if retrabalho == False:
    #         if nome and num_bobines and core and largura_bobinagem:
    #             perfil = Perfil.objects.filter(Q(nome__icontains=nome) & Q(num_bobines__iexact=num_bobines) & Q(core__iexact=core) & Q(largura_bobinagem__iexact=largura_bobinagem) & Q(retrabalho=False) & Q(obsoleto=False))
    #         elif nome and num_bobines and core:
    #             perfil = Perfil.objects.filter(Q(nome__icontains=nome) & Q(num_bobines__iexact=num_bobines) & Q(core__iexact=core) & Q(retrabalho=False) & Q(obsoleto=False))
    #         elif nome and num_bobines and largura_bobinagem:
    #             perfil = Perfil.objects.filter(Q(nome__icontains=nome) & Q(num_bobines__iexact=num_bobines) & Q(largura_bobinagem__iexact=largura_bobinagem) & Q(retrabalho=False) & Q(obsoleto=False))
    #         elif nome and core and largura_bobinagem:
    #             perfil = Perfil.objects.filter(Q(nome__icontains=nome) & Q(core__iexact=core) & Q(largura_bobinagem__iexact=largura_bobinagem) & Q(retrabalho=False) & Q(obsoleto=False))
    #         elif num_bobines and core and largura_bobinagem:
    #             perfil = Perfil.objects.filter(Q(num_bobines__iexact=num_bobines) & Q(core__iexact=core) & Q(largura_bobinagem__iexact=largura_bobinagem) & Q(retrabalho=False) & Q(obsoleto=False))
    #         elif nome and num_bobines:
    #             perfil = Perfil.objects.filter(Q(nome__icontains=nome) & Q(num_bobines__iexact=num_bobines) & Q(retrabalho=False) & Q(obsoleto=False))
    #         elif nome and largura_bobinagem:
    #             perfil = Perfil.objects.filter(Q(nome__icontains=nome) & Q(largura_bobinagem__iexact=largura_bobinagem) & Q(retrabalho=False) & Q(obsoleto=False))
    #         elif num_bobines and largura_bobinagem:
    #             perfil = Perfil.objects.filter(Q(num_bobines__iexact=num_bobines) & Q(largura_bobinagem__iexact=largura_bobinagem) & Q(retrabalho=False)& Q(obsoleto=False))
    #         elif core and largura_bobinagem:
    #             perfil = Perfil.objects.filter(Q(core__iexact=core) & Q(largura_bobinagem__iexact=largura_bobinagem) & Q(retrabalho=False) & Q(obsoleto=False))
    #         elif nome and core:
    #             perfil = Perfil.objects.filter(Q(nome__icontains=nome) & Q(core__iexact=core) & Q(retrabalho=False) & Q(obsoleto=False))
    #         elif num_bobines and core:
    #             perfil = Perfil.objects.filter(Q(num_bobines__iexact=num_bobines) & Q(core__iexact=core) & Q(retrabalho=False) & Q(obsoleto=False))
    #         elif nome:
    #             perfil = Perfil.objects.filter(Q(nome__icontains=nome) & Q(retrabalho=False) & Q(obsoleto=False))
    #         elif num_bobines:
    #             perfil = Perfil.objects.filter(Q(num_bobines__iexact=num_bobines) & Q(retrabalho=False) & Q(obsoleto=False))
    #         elif core:
    #             perfil = Perfil.objects.filter(Q(core__iexact=core) & Q(retrabalho=False) & Q(obsoleto=False))
    #         elif largura_bobinagem:
    #             perfil = Perfil.objects.filter(Q(largura_bobinagem__iexact=largura_bobinagem) & Q(retrabalho=False) & Q(obsoleto=False))
    #         else:
    #             perfil = Perfil.objects.filter(Q(retrabalho=False) & Q(obsoleto=False))


@login_required
def perfil_details_v2(request, pk):
    template_name = 'perfil/perfil_details_v2.html'
    perfil = get_object_or_404(Perfil, pk=pk)
    largura = Largura.objects.filter(perfil=pk)

    context = {
        "perfil": perfil,
        "largura": largura,
    }
    return render(request, template_name, context)


@login_required
def perfil_create_linha_v2(request):
    template_name = 'perfil/perfil_create_linha_v2.html'
    form = PerfilLinhaForm(request.POST or None)

    if form.is_valid():
        instance = form.save(commit=False)
        cd = form.cleaned_data
        largura_bobines = cd.get('largura_bobines')

        instance.user = request.user
        instance.retrabalho = False
        instance.save()

        for i in range(instance.num_bobines):
            if largura_bobines and instance.gramagem:
                lar = Largura.objects.create(
                    perfil=instance, num_bobine=i+1, designacao_prod=instance.produto, largura=largura_bobines, gsm=instance.gramagem)
                lar.save()
            elif largura_bobines:
                lar = Largura.objects.create(
                    perfil=instance, num_bobine=i+1, designacao_prod=instance.produto, largura=largura_bobines)
            elif instance.gramagem:
                lar = Largura.objects.create(
                    perfil=instance, num_bobine=i+1, designacao_prod=instance.produto, gsm=instance.gramagem)
            else:
                lar = Largura.objects.create(
                    perfil=instance, num_bobine=i+1, designacao_prod=instance.produto)

        return redirect('producao:perfil_larguras_v2', pk=instance.pk)

    context = {
        "form": form,
    }
    return render(request, template_name, context)


@login_required
def perfil_create_dm_v2(request):
    template_name = 'perfil/perfil_create_dm_v2.html'
    form = PerfilDMForm(request.POST or None)

    if form.is_valid():
        instance = form.save(commit=False)
        cd = form.cleaned_data
        largura_bobines = cd.get('largura_bobines')
        largura_original = cd.get('largura_original')
        core_original = cd.get('core_original')
        instance.user = request.user
        instance.retrabalho = True
        instance.core_original = core_original
        instance.largura_original = largura_original

        instance.save()

        for i in range(instance.num_bobines):
            if largura_bobines and instance.gramagem:
                lar = Largura.objects.create(
                    perfil=instance, num_bobine=i+1, designacao_prod=instance.produto, largura=largura_bobines, gsm=instance.gramagem)
                lar.save()
            elif largura_bobines:
                lar = Largura.objects.create(
                    perfil=instance, num_bobine=i+1, designacao_prod=instance.produto, largura=largura_bobines)
            elif instance.gramagem:
                lar = Largura.objects.create(
                    perfil=instance, num_bobine=i+1, designacao_prod=instance.produto, gsm=instance.gramagem)
            else:
                lar = Largura.objects.create(
                    perfil=instance, num_bobine=i+1, designacao_prod=instance.produto)

        return redirect('producao:perfil_larguras_v2', pk=instance.pk)

    context = {
        "form": form,
    }
    return render(request, template_name, context)


@login_required
def perfil_larguras_v2(request, pk):
    perfil = get_object_or_404(Perfil, pk=pk)
    template_name = 'perfil/perfil_larguras_v2.html'
    # LargurasPerfilFormSet = modelformset_factory(Largura, fields=('designacao_prod', 'largura', 'gsm'), extra=0)
    LargurasPerfilFormSet = modelformset_factory(Largura, fields=(
        'designacao_prod', 'largura', 'gsm', 'cliente', 'artigo'), extra=0)
    largura_total = 0
    larguras = []
    produtos = []
    clientes = []
    artigos = []
    gsms = []
    cont = []
    nome_largura = ''
    bobinagem = Bobinagem.objects.filter(perfil=perfil)
    can_edit = True
    valid = True
    if bobinagem.exists():
        can_edit = False

    if request.method == 'POST':
        formset = LargurasPerfilFormSet(
            request.POST, queryset=Largura.objects.filter(perfil=perfil))
        if formset.is_valid():
            for f in formset:
                cd = f.cleaned_data
                designacao_prod = cd.get('designacao_prod')
                largura = cd.get('largura')
                gsm = cd.get('gsm')
                cliente = cd.get('cliente')
                artigo = cd.get('artigo')
                largura_total += int(largura)
                larguras.append(largura)
                produtos.append(designacao_prod)
                gsms.append(gsm)
                clientes.append(cliente)
                artigos.append(artigo)
                artigo_obj = Artigo.objects.get(id=artigo.id)

                if artigo_obj.cod == 'EEEEFTACPAAR000181' or 'EEEEFTACPAAR000042' or 'EEEEFTACPAAR000228':
                    pass
                elif designacao_prod != artigo_obj.produto or gsm != artigo_obj.gsm or largura != artigo_obj.lar:
                    valid = False

            if valid == False:
                messages.error(
                    request, 'O artigo selecionado não condiz com as caracteristicas submetidas. Verifique o artigo selecionado.')
            else:
                if perfil.retrabalho == True and perfil.largura_original < largura_total:
                    messages.error(
                        request, 'Não é possivel validar as larguras inseridas. A largura da bobine original é inferior ao total da bobine final')
                else:

                    lar_s_dupli = list(dict.fromkeys(larguras))

                    for i in lar_s_dupli:
                        num = larguras.count(i)
                        cont.append(num)

                    larguras_dict = dict(zip(lar_s_dupli, cont))
                    for ld in larguras_dict:
                        nome_largura += str(larguras_dict[ld]) + \
                            'x' + str(ld) + '+'

                    nome_largura = nome_largura[:-1]
                    if perfil.retrabalho == False:
                        nome_parcial = 'L1 ' + perfil.produto + \
                            ' [' + nome_largura + '=' + \
                            str(largura_total) + '] ' + perfil.core + '"'
                    if perfil.retrabalho == True:
                        nome_parcial = 'DM ' + perfil.produto + ' [' + nome_largura + '=' + str(
                            largura_total) + '] ' + perfil.core + '"' + ' CO:' + str(perfil.core_original) + '" LO:' + str(perfil.largura_original)

                    if Perfil.objects.filter(nome__icontains=nome_parcial).exists():
                        perfis = Perfil.objects.filter(
                            nome__icontains=nome_parcial).count()
                        perfis += 1
                        nome = nome_parcial + ' ' + str(perfis)

                    else:
                        nome = nome_parcial + ' 1'

                    if perfil.retrabalho == True:
                        token = create_perfil_token(perfil.num_bobines, perfil.produto, perfil.core, larguras, produtos,
                                                    gsms, perfil.retrabalho, perfil.core_original, perfil.largura_original, clientes, artigos)
                    else:
                        token = create_perfil_token(perfil.num_bobines, perfil.produto, perfil.core,
                                                    larguras, produtos, gsms, perfil.retrabalho, None, None, clientes, artigos)

                    if Perfil.objects.filter(token=token).exists():
                        perfil_2 = get_object_or_404(Perfil, token=token)
                        messages.error(
                            request, 'O perfil que deseja criar já existe. Verifique as larguras, produtos e gramagens atribuidas.')
                        messages.error(
                            request, 'O perfil que procura é ' + perfil_2.nome)
                    else:
                        # print(largura_total, larguras, larguras_dict, nome_parcial, produtos, gsms)
                        # print(token)
                        perfil.token = token
                        perfil.largura_bobinagem = largura_total
                        perfil.nome = nome
                        perfil.save()
                        formset.save()
                        return redirect('producao:perfil_details_v2', pk=perfil.pk)

    else:
        formset = LargurasPerfilFormSet(
            queryset=Largura.objects.filter(perfil=perfil))

    context = {
        "formset": formset,
        "perfil": perfil,
        "can_edit": can_edit
    }
    return render(request, template_name, context)


@login_required
def perfil_edit_larguras_v2(request, pk):
    perfil = get_object_or_404(Perfil, pk=pk)
    template_name = 'perfil/perfil_edit_larguras_v2.html'
    # LargurasPerfilFormSet = modelformset_factory(Largura, fields=('designacao_prod', 'largura', 'gsm'), extra=0)
    LargurasPerfilFormSet = modelformset_factory(Largura, fields=(
        'designacao_prod', 'largura', 'gsm', 'cliente', 'artigo'), extra=0)
    largura_total = 0
    larguras = []
    produtos = []
    clientes = []
    artigos = []
    gsms = []
    cont = []
    nome_largura = ''
    bobinagem = Bobinagem.objects.filter(perfil=perfil)
    can_edit = True
    if bobinagem.exists():
        can_edit = False

    if request.method == 'POST':
        formset = LargurasPerfilFormSet(
            request.POST, queryset=Largura.objects.filter(perfil=perfil))
        if formset.is_valid():
            for f in formset:
                cd = f.cleaned_data
                designacao_prod = cd.get('designacao_prod')
                largura = cd.get('largura')
                gsm = cd.get('gsm')
                cliente = cd.get('cliente')
                artigo = cd.get('artigo')
                largura_total += int(largura)
                larguras.append(largura)
                produtos.append(designacao_prod)
                gsms.append(gsm)
                clientes.append(cliente)
                artigos.append(artigo)

            lar_s_dupli = list(dict.fromkeys(larguras))

            for i in lar_s_dupli:
                num = larguras.count(i)
                cont.append(num)

            larguras_dict = dict(zip(lar_s_dupli, cont))
            for ld in larguras_dict:
                nome_largura += str(larguras_dict[ld]) + 'x' + str(ld) + '+'

            nome_largura = nome_largura[:-1]
            if perfil.retrabalho == False:
                nome_parcial = 'L1 ' + perfil.produto + \
                    ' [' + nome_largura + '=' + \
                    str(largura_total) + '] ' + perfil.core + '"'
            if perfil.retrabalho == True:
                nome_parcial = 'DM ' + perfil.produto + ' [' + nome_largura + '=' + str(
                    largura_total) + '] ' + perfil.core + '"' + ' CO:' + str(perfil.core_original) + '" LO:' + str(perfil.largura_original)

            if Perfil.objects.filter(nome__icontains=nome_parcial).exists():
                perfis = Perfil.objects.filter(
                    nome__icontains=nome_parcial).count()
                perfis += 1
                nome = nome_parcial + ' ' + str(perfis)
            else:
                nome = nome_parcial + ' 1'

            if perfil.retrabalho == True:
                token = create_perfil_token(perfil.num_bobines, perfil.produto, perfil.core, larguras, produtos,
                                            gsms, perfil.retrabalho, perfil.core_original, perfil.largura_original, clientes, artigos)
            else:
                token = create_perfil_token(perfil.num_bobines, perfil.produto, perfil.core,
                                            larguras, produtos, gsms, perfil.retrabalho, None, None, clientes, artigos)

            if Perfil.objects.filter(token=token).exists():
                perfil_2 = get_object_or_404(Perfil, token=token)
                messages.error(
                    request, 'O perfil que deseja criar já existe. Verifique as larguras, produtos e gramagens atribuidas.')
                messages.error(
                    request, 'O perfil que procura é ' + perfil_2.nome)
            else:
                # print(largura_total, larguras, larguras_dict, nome_parcial, produtos, gsms)
                # print(token)
                perfil.token = token
                perfil.largura_bobinagem = largura_total
                perfil.nome = nome
                perfil.save()
                formset.save()
                return redirect('producao:perfil_details_v2', pk=perfil.pk)

    else:
        formset = LargurasPerfilFormSet(
            queryset=Largura.objects.filter(perfil=perfil))

    context = {
        "formset": formset,
        "perfil": perfil,
        "can_edit": can_edit
    }
    return render(request, template_name, context)


@login_required
def perfil_delete_v2(request, pk):
    perfil = get_object_or_404(Perfil, pk=pk)
    template_name = 'perfil/perfil_delete_v2.html'
    can_delete = True
    bobinagem = Bobinagem.objects.filter(perfil=perfil)
    if bobinagem.exists():
        can_delete = False

    if request.method == 'POST':
        perfil.delete()
        return redirect('producao:perfil_list_v2')

    context = {
        "can_delete": can_delete,
        "perfil": perfil,
        "bobinagem": bobinagem

    }
    return render(request, template_name, context)


@login_required
def bobinagem_retrabalho_list_v2(request):
    bobinagem_list = Bobinagem.objects.filter(
        perfil__retrabalho=True).order_by('-data', '-num_bobinagem')
    template_name = 'retrabalho/bobinagem_retrabalho_list_v2.html'

    query = ""
    if request.GET:
        query = request.GET.get('q', '')
        # print(query)
        bobinagem_list = Bobinagem.objects.filter(
            nome__icontains=query, perfil__retrabalho=True).order_by('-data', '-num_bobinagem')

    paginator = Paginator(bobinagem_list, 10)
    page = request.GET.get('page')

    try:
        bobinagem = paginator.page(page)
    except PageNotAnInteger:
        bobinagem = paginator.page(1)
    except EmptyPage:
        bobinagem = paginator.page(paginator.num_pages)

    context = {
        "bobinagem": bobinagem,
        "query": query
    }
    return render(request, template_name, context)


@login_required
def inventario_bobines_list(request):
    bobines_inventario = InventarioBobinesDM.objects.all()
    template_name = 'inventario/inventario_bobines_list.html'
    bobines = []
    for bi in bobines_inventario:
        b = get_object_or_404(Bobine, pk=bi.bobine.pk)
        bobines.append(b)

    context = {

        "bobines": bobines

    }

    return render(request, template_name, context)


@login_required
def inventario_bobines_dm_insert(request):
    template_name = 'inventario/inventario_bobines_dm_insert.html'
    form = InventarioBobineDMInsert(request.POST or None)

    if form.is_valid():
        cd = form.cleaned_data
        bobine_picada = cd.get('bobine')
        bobines = Bobine.objects.filter(nome=bobine_picada)
        user = request.user.username
        if user != 'elsa.rodrigues':
            messages.error(
                request, 'Não tem permissão para introduzir bobines no inventário. Por favor, fale com o administrador.')
        else:
            if bobines.exists():
                bobine = InventarioBobinesDM.objects.filter(nome=bobine_picada)
                if bobine.exists():
                    messages.warning(
                        request, 'A bobine ' + bobine_picada + ' já se encontra no Inventário.')
                    form = InventarioBobineDMInsert()
                else:
                    bobine = get_object_or_404(Bobine, nome=bobine_picada)
                    bobine_inv = InventarioBobinesDM.objects.create(
                        user=request.user, bobine=bobine, nome=bobine_picada)
                    bobine_inv.save()
                    messages.success(request, 'A bobine ' +
                                     bobine_picada + ' foi inserida com sucesso.')
                    form = InventarioBobineDMInsert()

            else:
                messages.error(
                    request, 'A bobine que inseriu não existe. Tente de novo.')
                form = InventarioBobineDMInsert()

    context = {
        "form": form,
    }
    return render(request, template_name, context)


@login_required
def inventario_bobines_dm_remover(request, pk):
    bobine = get_object_or_404(Bobine, pk=pk)
    bobine_inv = get_object_or_404(InventarioBobinesDM, nome=bobine.nome)
    template_name = 'inventario/inventario_bobines_dm_remover.html'

    if request.method == 'POST':
        bobine_inv.delete()
        return redirect('producao:inventario_bobines_list')

    context = {
        "bobine": bobine,
        "bobine_inv": bobine_inv
    }

    return render(request, template_name, context)


@login_required
def inventario_paletes_cliente_list(request):
    paletes_inventario = InventarioPaletesCliente.objects.all()
    template_name = 'inventario/inventario_paletes_cliente_list.html'
    # paletes = []
    # for pi in paletes_inventario:
    #     p = get_object_or_404(Palete, pk=pi.palete.pk)
    #     paletes.append(p)

    context = {
        # "paletes": paletes,
        "paletes_inventario": paletes_inventario
    }

    return render(request, template_name, context)


@login_required
def inventario_palete_cliente_insert(request):
    template_name = 'inventario/inventario_palete_cliente_insert.html'
    form = InventarioPaleteClienteInsert(request.POST or None)

    if form.is_valid():
        cd = form.cleaned_data
        palete_picada = cd.get('palete')
        paletes = Palete.objects.filter(nome=palete_picada)
        user = request.user.username
        # if user != 'elsa.rodrigues':
        #     messages.error(request, 'Não tem permissão para introduzir Paletes no inventário. Por favor, fale com o administrador.')
        # else:
        if paletes.exists():
            palete = InventarioPaletesCliente.objects.filter(
                nome=palete_picada)
            if palete.exists():
                messages.warning(request, 'A palete ' +
                                 palete_picada + ' já se encontra no Inventário.')
                form = InventarioPaleteClienteInsert()
            else:
                palete = get_object_or_404(Palete, nome=palete_picada)
                palete_etiqueta = get_object_or_404(
                    EtiquetaPalete, palete=palete)
                bobines = Bobine.objects.filter(palete=palete)
                artigo = bobines[0].artigo
                palete_inv = InventarioPaletesCliente.objects.create(user=request.user, palete=palete, nome=palete_picada, cliente=palete_etiqueta.cliente, comp_total=palete.comp_total, area=palete.area,
                                                                     largura_bobine=palete_etiqueta.largura_bobine, diam_max=palete_etiqueta.diam_max, diam_min=palete_etiqueta.diam_min, core_bobines=palete.core_bobines, artigo=artigo)
                palete_inv.save()
                messages.success(request, 'A palete ' +
                                 palete_picada + ' foi inserida com sucesso.')
                form = InventarioPaleteClienteInsert()

        else:
            messages.error(
                request, 'A palete que inseriu não existe. Tente de novo.')
            form = InventarioPaleteClienteInsert()

    context = {
        "form": form,

    }
    return render(request, template_name, context)


@login_required
def bobines_larguras_reais(request, pk):
    bobinagem = get_object_or_404(Bobinagem, pk=pk)
    template_name = 'producao/bobines_larguras_reais.html'
    LargurasReaisFormSet = modelformset_factory(
        Bobine, fields=('l_real',), extra=0)
    larguras_validator = []

    if request.method == 'POST':
        formset = LargurasReaisFormSet(
            request.POST, queryset=Bobine.objects.filter(bobinagem=bobinagem))
        if formset.is_valid():
            for f in formset:
                bobine = f.save(commit=False)
                cd = f.cleaned_data
                l_real = cd.get('l_real')
                if l_real == None:
                    larguras_validator.append(1)
                else:
                    bobine.save()

            print(larguras_validator)
            if len(larguras_validator) == 0:
                return redirect('producao:etiqueta_retrabalho', pk=bobinagem.pk)
            else:
                messages.error(
                    request, 'Todos os campos de larguras reais são de preenchimento obrigatório.')

    else:
        formset = LargurasReaisFormSet(
            queryset=Bobine.objects.filter(bobinagem=bobinagem))

    context = {
        "formset": formset,
        "bobinagem": bobinagem
    }

    return render(request, template_name, context)


@login_required
def bobinagem_list_v3(request):
    bobinagem_list = Bobinagem.objects.filter(
        perfil__retrabalho=False).order_by('-data', '-num_bobinagem')
    template_name = 'producao/bobinagem_list_v3.html'

    query = ""
    if request.GET:
        query = request.GET.get('q', '')
        # print(query)
        bobinagem_list = Bobinagem.objects.filter(
            nome__icontains=query, perfil__retrabalho=False).order_by('-data', '-num_bobinagem')

    paginator = Paginator(bobinagem_list, 12)
    page = request.GET.get('page')

    try:
        bobinagem = paginator.page(page)
    except PageNotAnInteger:
        bobinagem = paginator.page(1)
    except EmptyPage:
        bobinagem = paginator.page(paginator.num_pages)

    context = {
        "bobinagem": bobinagem,
        "query": query,

    }
    return render(request, template_name, context)


@login_required
def cliente_details(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)
    artigos_cliente = ArtigoCliente.objects.filter(cliente=cliente)
    template_name = 'cliente/cliente_details.html'

    context = {
        "cliente": cliente,
        "artigos_cliente": artigos_cliente
    }

    return render(request, template_name, context)


@login_required
def cliente_add_artigo(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)
    artigo_cliente = ArtigoCliente.objects.filter(cliente=cliente)
    template_name = 'cliente/cliente_add_artigo.html'
    form = ArtigoClientInsert(request.POST or None)
    # form_embal = EmbalamentoCreateForm(request.POST or None)
    artigo_valido = True
    # if form.is_valid() and form_embal.is_valid():
    #     instance = form.save(commit=False)
    #     instance_embal = form_embal.save(commit=False)
    #     instance.user = request.user
    #     instance_embal.user = request.user
    #     instance.cliente = cliente

    #     for ac in artigo_cliente:
    #         if ac.artigo == instance.artigo:
    #             artigo_valido = False
    #             break

    #     if artigo_valido == True:
    #         instance_embal.save()
    #         instance.embalamento = instance_embal
    #         instance.save()
    #         return redirect('producao:cliente_details', pk=cliente.pk)
    #     else:
    #         messages.error(request, 'O Artigo ' + instance.artigo.des + ' já está associado ao cliente ' + cliente.nome + '.')
    #         form = ArtigoClientInsert()

    if form.is_valid():
        instance = form.save(commit=False)
        instance.user = request.user
        instance.cliente = cliente

        for ac in artigo_cliente:
            if ac.artigo == instance.artigo:
                artigo_valido = False
                break

        if artigo_valido == True:
            instance.save()
            return redirect('producao:cliente_details', pk=cliente.pk)
        else:
            messages.error(request, 'O Artigo ' + instance.artigo.des + ' já está associado ao cliente ' + cliente.nome + '.')
            form = ArtigoClientInsert()

    context = {
        "cliente": cliente,
        "form": form,
        # "form_embal": form_embal,
    }

    return render(request, template_name, context)


@login_required
def cliente_remover_artigo(request, pk_cliente, pk_artigo):
    cliente = get_object_or_404(Cliente, pk=pk_cliente)
    artigo = get_object_or_404(Artigo, pk=pk_artigo)
    artigo_cliente = ArtigoCliente.objects.get(cliente=cliente, artigo=artigo)
    template_name = 'cliente/cliente_remover_artigo.html'

    if request.method == "POST":
        artigo_cliente.delete()
        return redirect('producao:cliente_details', pk=cliente.pk)

    context = {
        "cliente": cliente,
        "artigo:": artigo,
        "artigo_cliente": artigo_cliente
    }

    return render(request, template_name, context)


@login_required
def load_artigos_cliente(request):
    cliente = request.GET.get('cliente')
    largura = request.GET.get('largura')
    produto = request.GET.get('produto')
    gsm = request.GET.get('gsm')
    cliente_obj = get_object_or_404(Cliente, pk=cliente)
    print("aaaaaaaaaaaaaaaaaaaaaaa")
    print(cliente)
    print(largura)
    print(produto)
    print(gsm)
    print(cliente_obj.cod)
    if cliente_obj.cod == 0:
        artigos_cliente = ArtigoCliente.objects.filter(
            cliente=cliente_obj).order_by('artigo')
    else:
        artigos_cliente = ArtigoCliente.objects.filter(
            cliente=cliente_obj, artigo__lar=largura, artigo__gsm=gsm).order_by('artigo')
    return render(request, 'perfil/dropdown_options.html', {'artigos_cliente': artigos_cliente})


@login_required
def fornecedor_list(request):
    fornecedor_list = Fornecedor.objects.all().order_by('-timestamp')
    template_name = 'fornecedor/fornecedor_list.html'

    query = ""
    if request.GET:
        query = request.GET.get('q', '')
        # print(query)
        fornecedor_list = Fornecedor.objects.filter(
            designacao__icontains=query).order_by('-timestamp')

    paginator = Paginator(fornecedor_list, 12)
    page = request.GET.get('page')

    try:
        fornecedor = paginator.page(page)
    except PageNotAnInteger:
        fornecedor = paginator.page(1)
    except EmptyPage:
        fornecedor = paginator.page(paginator.num_pages)

    context = {
        "fornecedor": fornecedor,
        "query": query,

    }
    return render(request, template_name, context)


@login_required
def fornecedor_create(request):
    template_name = 'fornecedor/fornecedor_create.html'
    form = FornecedorCreateForm(request.POST or None)

    if form.is_valid():
        instance = form.save(commit=False)
        instance.user = request.user
        instance.save()
        return redirect('producao:fornecedor_list')

    context = {
        "form": form,
    }
    return render(request, template_name, context)


@login_required
def fornecedor_delete(request, pk):
    fornecedor = get_object_or_404(Fornecedor, pk=pk)
    template_name = 'fornecedor/fornecedor_delete.html'
    can_delete = True
    artigonw = ArtigoNW.objects.filter(fornecedor=fornecedor)
    if artigonw.exists():
        can_delete = False

    if request.method == 'POST':
        fornecedor.delete()
        return redirect('producao:fornecedor_list')

    context = {
        "can_delete": can_delete,
        "fornecedor": fornecedor,
        "artigonw": artigonw

    }
    return render(request, template_name, context)


@login_required
def fornecedor_edit(request, pk):
    template_name = 'fornecedor/fornecedor_edit.html'
    fornecedor = get_object_or_404(Fornecedor, pk=pk)
    form = FornecedorEditForm(request.POST or None, instance=fornecedor)

    if form.is_valid():
        instance = form.save(commit=False)
        instance.user = request.user
        instance.save()
        return redirect('producao:fornecedor_list')

    context = {
        "form": form,
        "fornecedor": fornecedor
    }
    return render(request, template_name, context)


@login_required
def rececao_list(request):
    rececao_list = Rececao.objects.all().order_by('estado', '-num')
    template_name = 'rececao/rececao_list.html'

    query = ""
    if request.GET:
        query = request.GET.get('q', '')
        # print(query)
        rececao_list = Rececao.objects.filter(
            rececao__icontains=query).order_by('estado', '-num')

    paginator = Paginator(rececao_list, 14)
    page = request.GET.get('page')

    try:
        rececao = paginator.page(page)
    except PageNotAnInteger:
        rececao = paginator.page(1)
    except EmptyPage:
        rececao = paginator.page(paginator.num_pages)

    context = {
        "rececao": rececao,
        "query": query,

    }
    return render(request, template_name, context)


@login_required
def rececao_create(request):
    template_name = 'rececao/rececao_create.html'
    form = RececaoCreateForm(request.POST or None)
    rececoes = Rececao.objects.all().order_by('-num')

    if form.is_valid():
        instance = form.save(commit=False)
        instance.user = request.user
        instance.quantidade = 0
        instance.estado = 'A'
        ult_rec = rececoes.latest('num')
        instance.num = ult_rec.num + 1
        date = datetime.now()
        mes = date.strftime("%m")
        ano = date.strftime("%y")
        str_num = ''
        if instance.num < 10:
            str_num = '0000'
        elif instance.num < 100:
            str_num = '000'
        elif instance.num < 1000:
            str_num = '00'
        elif instance.num < 10000:
            str_num = '0'
        elif instance.num < 100000:
            str_num = ''

        instance.rececao = 'R-' + mes + ano + '/' + str_num + str(instance.num)
        instance.save()
        return redirect('producao:rececao_insert_nw', pk=instance.pk)

    context = {
        "form": form,
    }
    return render(request, template_name, context)


@login_required
def artigonw_list(request):
    artigonw_list = ArtigoNW.objects.all().order_by('-timestamp')
    template_name = 'artigonw/artigonw_list.html'

    query = ""
    if request.GET:
        query = request.GET.get('q', '')
        # print(query)
        artigonw_list = ArtigoNW.objects.filter(
            designacao__icontains=query).order_by('-timestamp')

    paginator = Paginator(artigonw_list, 14)
    page = request.GET.get('page')

    try:
        artigonw = paginator.page(page)
    except PageNotAnInteger:
        artigonw = paginator.page(1)
    except EmptyPage:
        artigonw = paginator.page(paginator.num_pages)

    context = {
        "artigonw": artigonw,
        "query": query,

    }
    return render(request, template_name, context)


@login_required
def artigonw_create(request):
    template_name = 'artigonw/artigonw_create.html'
    form = ArtigoNWCreateForm(request.POST or None)

    if form.is_valid():
        instance = form.save(commit=False)
        instance.user = request.user

        instance.save()
        return redirect('producao:artigonw_list')

    context = {
        "form": form,
    }
    return render(request, template_name, context)


@login_required
def rececao_insert_nw(request, pk):
    template_name = 'rececao/rececao_insert_nw.html'
    rececao = get_object_or_404(Rececao, pk=pk)
    fornecedor = get_object_or_404(Fornecedor, pk=rececao.fornecedor.pk)
    form = RececaoInsertNW(request.POST or None)

    if form.is_valid():
        cd = form.cleaned_data
        cod_artigo = cd.get('artigo_nw')
        sqm = cd.get('sqm')
        lote = cd.get('lote')
        prod = cd.get('prod')
        stack_num = cd.get('stack_num')
        # print(artigo_nw)
        # print(sqm)
        # print(lote)
        # print(prod)
        # print(stack_num)
        artigo_nw = ''

        try:
            artigo_nw = get_object_or_404(
                ArtigoNW, cod=cod_artigo, fornecedor=fornecedor)
            form = RececaoInsertNW()

        except:
            messages.error(request, 'O Artigo com o código ' + cod_artigo +
                           ' não existe ou  não corresponde ao fornecedor da receção.')
            form = RececaoInsertNW()

        try:
            sqm = sqm[:-3]
            sqm = Decimal(sqm.replace(',', '.'))
            form = RececaoInsertNW()
        except:
            messages.error(request, 'Metros quadrados inválidos')
            form = RececaoInsertNW()

        try:
            nw = get_object_or_404(Nonwoven, stack_num=stack_num)
            messages.error(
                request, 'O Nonwoven com a Stack Number ' + stack_num + ' já foi inserido.')
        except:
            pass

        try:
            comp = sqm / (artigo_nw.largura / 1000)
            nw = Nonwoven.objects.create(user=request.user, artigo_nw=artigo_nw, rececao=rececao,
                                         stack_num=stack_num, sqm=sqm, lote=lote, prod=prod, comp_total=comp, comp_actual=comp)
            cont = Nonwoven.objects.filter(rececao=rececao).count()

            if cont < 10:
                cont = '00' + str(cont)
            elif cont < 100:
                cont = '0' + str(cont)
            else:
                cont = '' + str(cont)

            cod_nw = 'NW' + rececao.encomenda + cont

            etiquetanw = EtiquetaNonwoven.objects.create(nonwoven=nw, rececao=rececao, cod_nw=cod_nw, rececao_rec=rececao.rececao,
                                                         encomenda=rececao.encomenda, data_rec=rececao.timestamp, nw_des=nw.artigo_nw.designacao)
            nw.cod_nw = cod_nw
            nw.save()
            rececao.quantidade += 1
            rececao.save()
            messages.success(request, 'Nonwoven inserido com successo')
        except:
            pass

    context = {
        "form": form,
        "rececao": rececao
    }
    return render(request, template_name, context)


@login_required
def rececao_close(request, pk):
    rececao = get_object_or_404(Rececao, pk=pk)
    nonwovens = Nonwoven.objects.filter(rececao=rececao)
    template_name = 'rececao/rececao_close.html'

    if request.method == "POST":
        rececao.estado = 'F'
        rececao.save()
        return redirect('producao:rececao_list')

    context = {
        "rececao": rececao,
        "nonwovens": nonwovens
    }
    return render(request, template_name, context)


@login_required
def rececao_open(request, pk):
    rececao = get_object_or_404(Rececao, pk=pk)
    nonwovens = Nonwoven.objects.filter(rececao=rececao)
    template_name = 'rececao/rececao_open.html'

    if request.method == "POST":
        rececao.estado = 'A'
        rececao.save()
        return redirect('producao:rececao_insert_nw', pk=rececao.pk)

    context = {
        "rececao": rececao,
        "nonwovens": nonwovens
    }
    return render(request, template_name, context)


@login_required
def fornecedor_details(request, pk):
    fornecedor = get_object_or_404(Fornecedor, pk=pk)
    template_name = 'fornecedor/fornecedor_details.html'

    artigonw = ArtigoNW.objects.filter(fornecedor=fornecedor)

    context = {
        "fornecedor": fornecedor,
        "artigonw": artigonw
    }

    return render(request, template_name, context)


@login_required
def artigonw_details(request, pk):
    artigonw = get_object_or_404(ArtigoNW, pk=pk)
    nonwoven = Nonwoven.objects.filter(artigo_nw=artigonw)
    template_name = 'artigonw/artigonw_details.html'

    context = {
        "artigonw": artigonw,
        "nonwoven": nonwoven
    }

    return render(request, template_name, context)


@login_required
def artigonw_edit(request, pk):
    artigonw = get_object_or_404(ArtigoNW, pk=pk)
    template_name = 'artigonw/artigonw_edit.html'
    form = ArtigoNWCreateForm(request.POST or None, instance=artigonw)
    can_edit = True
    if Nonwoven.objects.filter(artigo_nw=artigonw).exists():
        can_edit = False

    if form.is_valid():
        instance = form.save(commit=False)
        instance.user = request.user

        instance.save()
        return redirect('producao:artigonw_list')

    context = {
        "form": form,
        "can_edit": can_edit
    }
    return render(request, template_name, context)


@login_required
def artigonw_delete(request, pk):
    artigonw = get_object_or_404(ArtigoNW, pk=pk)
    template_name = 'artigonw/artigonw_delete.html'
    can_delete = True
    if Nonwoven.objects.filter(artigo_nw=artigonw).exists():
        can_delete = False

    if request.method == 'POST':
        artigonw.delete()
        return redirect('producao:artigonw_list')

    context = {
        "can_delete": can_delete,
        "artigonw": artigonw

    }
    return render(request, template_name, context)


@login_required
def rececao_details(request, pk):
    rececao = get_object_or_404(Rececao, pk=pk)
    template_name = 'rececao/rececao_details.html'

    nonwoven = Nonwoven.objects.filter(rececao=rececao)

    context = {
        "rececao": rececao,
        "nonwoven": nonwoven
    }

    return render(request, template_name, context)


@login_required
def bobinagem_create_v2(request):

    template_name = 'producao/bobinagem_create_v2.html'
    form = BobinagemCreateFormV2(request.POST or None)

    if form.is_valid():
        instance = form.save(commit=False)
        instance.user = request.user

        cd = form.cleaned_data
        nonwoven_sup = cd.get('nonwoven_sup')
        consumo_sup = cd.get('consumo_sup')
        nonwoven_inf = cd.get('nonwoven_inf')
        consumo_inf = cd.get('consumo_inf')

        try:
            nonwoven_sup = get_object_or_404(Nonwoven, cod_nw=nonwoven_sup)
            nonwoven_inf = get_object_or_404(Nonwoven, cod_nw=nonwoven_inf)
            print(nonwoven_sup.cod_nw + '->' + str(consumo_sup))
            print(nonwoven_inf.cod_nw + '->' + str(consumo_inf))

        except:
            messages.error(request, 'O Nonwoven ' +
                           nonwoven_sup + ' não existe.')

        # sup = instance.lotenwsup
        # inf = instance.lotenwinf

        # metros_nwsup = instance.nwsup
        # metros_nwinf = instance.nwinf

        # instance.lotenwsup = sup.replace(" ", "")
        # instance.lotenwinf = inf.replace(" ", "")

        # if (instance.estado == 'R' and instance.obs == ''):
        #     messages.error(request, 'Para Rejeitar a bobinagem é necessario indicar motivo nas observações.')
        # else:
        #     nonwovensup = Bobinagem.objects.filter(lotenwsup=sup)
        #     nonwoveninf = Bobinagem.objects.filter(lotenwinf=inf)

        #     total_sup = instance.nwsup
        #     total_inf = instance.nwinf
        #     for ns in nonwovensup:
        #         total_sup += ns.nwsup
        #     for ni in nonwoveninf:
        #         total_inf += ni.nwinf

        #     # if (total_sup > 7500):
        #     #     messages.error(request, 'A soms total de metros do lote de Nonwoven superior "' + sup + '" excede o limite establecido de 7500. Por favor verifique o valor introduzido.')
        #     # if (total_inf > 7500):

        #     if (total_inf > 7500 or total_sup > 7500):
        #         messages.error(request, 'A soma total de metros dos lotes de Nonwoven excedem o limite establecido de 7500. Por favor verifique os valores introduzidos.')
        #     else:
        #         instance.save()
        #         bobinagem_create(instance.pk)

        #         if not instance.estado == 'LAB' or instance.estado == 'HOLD':
        #             areas(instance.pk)

        #         if instance.estado == 'BA':
        #             return redirect('producao:bobines_larguras_reais', pk=instance.pk)
        #         else:
        #             return redirect('producao:etiqueta_retrabalho', pk=instance.pk)

    context = {
        "form": form
    }

    return render(request, template_name, context)


@login_required
def carga_etiqueta_nonwoven(request, pk):
    nonwoven = get_object_or_404(Nonwoven, pk=pk)
    etiqueta = EtiquetaNonwoven.objects.get(nonwoven=nonwoven)

    etiqueta.impressora = 'ARMAZEM_CAB_SQUIX_6.3_200'
    etiqueta.num_copias = 1
    etiqueta.estado_impressao = True
    etiqueta.save()

    context = {}

    return redirect('producao:rececao_details', pk=nonwoven.rececao.pk)


@login_required
def carga_etiqueta_nonwoven_rececao(request, pk):
    rececao = get_object_or_404(Rececao, pk=pk)
    nonwovens = Nonwoven.objects.filter(rececao=rececao)
    for nw in nonwovens:
        etiqueta = EtiquetaNonwoven.objects.get(nonwoven=nw)
        etiqueta.impressora = 'ARMAZEM_CAB_SQUIX_6.3_200'
        etiqueta.num_copias = 1
        etiqueta.estado_impressao = True
        etiqueta.save()

    context = {}

    return redirect('producao:rececao_list')


@login_required
def bobinagem_classificacao(request, pk):


       
    BobineClassificacaoFormSet = inlineformset_factory(Bobinagem, Bobine,  fields=('estado', 'l_real', 'nok', 'con', 'descen', 'presa', 'diam_insuf', 'suj', 'car',  'lac', 'ncore', 'sbrt', 'fc',  'fc_diam_ini', 'fc_diam_fim', 'ff', 'ff_m_ini', 'ff_m_fim', 'fmp',  'furos', 'buraco', 'esp', 'prop', 'prop_obs', 'outros', 'obs', 'troca_nw'), extra=0, can_delete=False)
    bobinagem = get_object_or_404(Bobinagem, pk=pk)
    bobines = Bobine.objects.filter(bobinagem=bobinagem)
    template_name = 'bobine/bobinagem_classificacao.html'
    valid = True
    user = request.user
    for b in bobines:
        if b.estado == 'HOLD':
            valid = False
        else:
            pass

    if request.method == 'POST':

        formset = BobineClassificacaoFormSet(request.POST, instance=bobinagem)
        if formset.is_valid():
            index = 0
            messages_count = 0
            for form in formset:
                bobine = Bobine.objects.get(
                    bobinagem=bobinagem, largura__num_bobine=index+1)
                if not user.groups.filter(Q(name='Qualidade Supervisor') | Q(name='Qualidade Tecnico')).exists() and bobine.estado == 'HOLD':
                    index += 1
                    messages.error(
                        request, 'Não tem permissões para modificar o estado de uma bobine em HOLD.')
                else:
                    cd = form.cleaned_data
                    estado = cd.get('estado')
                    fc_diam_ini = cd.get('fc_diam_ini')
                    fc_diam_fim = cd.get('fc_diam_fim')
                    ff_m_ini = cd.get('ff_m_ini')
                    ff_m_fim = cd.get('ff_m_fim')
                    obs = cd.get('obs')
                    prop_obs = cd.get('prop_obs')
                    if not user.groups.filter(Q(name='Qualidade Supervisor') | Q(name='Qualidade Tecnico')).exists() and estado == 'HOLD':
                        index += 1
                        messages.error(
                            request, 'Não tem permissões para modificar o estado de uma bobine para HOLD')
                    else:
                        defeitos = [cd.get('nok'), cd.get('con'), cd.get('descen'), cd.get('presa'), cd.get('diam_insuf'), cd.get('suj'), cd.get('car'), cd.get('lac'), cd.get('ncore'), cd.get(
                            'sbrt'), cd.get('fc'), cd.get('ff'), cd.get('fmp'), cd.get('furos'), cd.get('buraco'), cd.get('esp'), cd.get('prop'), cd.get('outros'), cd.get('troca_nw')]
                        defeitos_validation = not any(defeitos)
                        index += 1
                        if (estado == 'DM' or estado == 'R') and defeitos_validation == True:
                            messages.error(request, 'Bobine nº' + str(
                                index) + ' - Para classificar a bobine como DM ou R é necessário atribuir pelo menos um defeito.')
                            messages_count += 1
                        elif (estado == 'DM' or estado == 'R') and defeitos[10] == True and (fc_diam_ini == None or fc_diam_fim == None):
                            messages.error(
                                request, 'Bobine nº' + str(index) + ' - Preencher inicio e fim da falha de corte.')
                            messages_count += 1
                        elif (estado == 'DM' or estado == 'R') and defeitos[11] == True and (ff_m_ini == None or ff_m_fim == None):
                            messages.error(
                                request, 'Bobine nº' + str(index) + ' - Preencher inicio e fim da falha de filme.')
                            messages_count += 1
                        elif (estado == 'DM' or estado == 'R') and defeitos[12] == True and obs == '':
                            messages.error(request, 'Bobine nº' + str(index) +
                                           ' - Falha de MP: Preencher nas observações o motivo.')
                            messages_count += 1
                        elif (estado == 'DM' or estado == 'R') and defeitos[14] == True and obs == '':
                            messages.error(request, 'Bobine nº' + str(
                                index) + ' - Buracos: Preencher nas observações os metros de desbobinagem.')
                            messages_count += 1
                        elif (estado == 'DM' or estado == 'R') and defeitos[15] == True and prop_obs == '':
                            messages.error(
                                request, 'Bobine nº' + str(index) + ' - Gramagem: Preencher nas Prop. Obs.')
                            messages_count += 1
                        elif (estado == 'DM' or estado == 'R') and defeitos[16] == True and prop_obs == '':
                            messages.error(
                                request, 'Bobine nº' + str(index) + ' - Propriedades: Preencher nas Prop. Obs.')
                            messages_count += 1
                        else:
                            form.save()
                            messages.success(
                                request, 'Bobine nº' + str(index) + ' - Alterações guardadas com sucesso.')

            # if messages_count == 0:
            #     return redirect('producao:bobinestatus', pk=bobinagem.pk)

                # if estado == 'DM':
                #     messages.error(request, 'DM')

                # form.save()

            # return redirect('producao:bobinestatus', pk=bobinagem.pk)

    formset = BobineClassificacaoFormSet(instance=bobinagem)

    context = {
        "formset": formset,
        "bobinagem": bobinagem,
        "bobines": bobines,
    }
    return render(request, template_name, context)


def classificacao_bobines_all(request, operation, pk):
    bobinagem = get_object_or_404(Bobinagem, pk=pk)
    bobines = Bobine.objects.filter(bobinagem=bobinagem)
    form = ClasssificacaoBobineDm(request.POST or None)
    template_name = 'bobine/bobinagem_classificacao_all.html'
    user = request.user
    if operation == 'ap':
        for bob in bobines:
            if bob.estado == 'LAB' or (user.groups.filter(Q(name='Qualidade Supervisor') | Q(name='Qualidade Tecnico')).exists() and bob.estado == 'HOLD'):
                bob.estado = 'G'
                bob.save()
            else:
                messages.error(
                    request, 'Não tem permissões para aprovar uma bobine em HOLD')

        return redirect('producao:bobinagem_classificacao', pk=bobinagem.pk)
    elif operation == 'hold':
        for bob in bobines:
            if bob.estado == 'LAB' or bob.estado == 'DM':
                bob.estado = 'HOLD'
                bob.save()
        return redirect('producao:bobinagem_classificacao', pk=bobinagem.pk)

    if form.is_valid():
        index = 0
        if operation == 'dm' or operation == 'rej':
            cd = form.cleaned_data
            fc_diam_ini = cd.get('fc_diam_ini')
            fc_diam_fim = cd.get('fc_diam_fim')
            ff_m_ini = cd.get('ff_m_ini')
            ff_m_fim = cd.get('ff_m_fim')
            obs = cd.get('obs')
            prop_obs = cd.get('prop_obs')

            defeitos = [cd.get('nok'), cd.get('con'), cd.get('descen'), cd.get('presa'), cd.get('diam_insuf'), cd.get('suj'), cd.get('car'), cd.get('lac'), cd.get('ncore'), cd.get(
                'sbrt'), cd.get('fc'), cd.get('ff'), cd.get('fmp'), cd.get('furos'), cd.get('buraco'), cd.get('esp'), cd.get('prop'), cd.get('outros'), cd.get('troca_nw')]
            defeitos_validation = not any(defeitos)
            if defeitos_validation == True:
                messages.error(request, 'Para classificar a bobinagem como ' +
                               operation + ' é necessário atribuir pelo menos um defeito.')
            elif defeitos[10] == True and (fc_diam_ini == None or fc_diam_fim == None):
                messages.error(
                    request, 'Preencher inicio e fim da falha de corte.')
            elif defeitos[11] == True and (ff_m_ini == None or ff_m_fim == None):
                messages.error(
                    request, 'Preencher inicio e fim da falha de filme.')
            elif defeitos[12] == True and obs == '':
                messages.error(
                    request, 'Falha de MP: Preencher nas observações o motivo.')
            elif defeitos[14] == True and obs == '':
                messages.error(
                    request, 'Buracos: Preencher nas observações os metros de desbobinagem.')
            elif defeitos[15] == True and prop_obs == '':
                messages.error(request, 'Gramagem: Preencher nas Prop. Obs.')
            elif defeitos[16] == True and prop_obs == '':
                messages.error(
                    request, 'Propriedades: Preencher nas Prop. Obs.')
            else:
                for bob in bobines:
                    if bob.estado == 'LAB' and operation == 'dm':
                        bob.estado = 'DM'
                        bob.nok = cd.get('nok')
                        bob.con = cd.get('con')
                        bob.descen = cd.get('descen')
                        bob.presa = cd.get('presa')
                        bob.diam_insuf = cd.get('diam_insuf')
                        bob.suj = cd.get('suj')
                        bob.car = cd.get('car')
                        bob.lac = cd.get('lac')
                        bob.ncore = cd.get('ncore')
                        bob.sbrt = cd.get('sbrt')
                        bob.fc = cd.get('fc')
                        bob.ff = cd.get('ff')
                        bob.fmp = cd.get('fmp')
                        bob.furos = cd.get('furos')
                        bob.buraco = cd.get('buraco')
                        bob.esp = cd.get('esp')
                        bob.prop = cd.get('prop')
                        bob.outros = cd.get('outros')
                        bob.troca_nw = cd.get('troca_nw')
                        bob.fc_diam_ini = cd.get('fc_diam_ini')
                        bob.fc_diam_fim = cd.get('fc_diam_fim')
                        bob.ff_m_ini = cd.get('ff_m_ini')
                        bob.ff_m_fim = cd.get('ff_m_fim')
                        bob.obs = cd.get('obs')
                        bob.prop_obs = cd.get('prop_obs')
                        bob.save()
                    if bob.estado == 'LAB' and operation == 'rej':
                        bob.estado = 'R'
                        bob.nok = cd.get('nok')
                        bob.con = cd.get('con')
                        bob.descen = cd.get('descen')
                        bob.presa = cd.get('presa')
                        bob.diam_insuf = cd.get('diam_insuf')
                        bob.suj = cd.get('suj')
                        bob.car = cd.get('car')
                        bob.lac = cd.get('lac')
                        bob.ncore = cd.get('ncore')
                        bob.sbrt = cd.get('sbrt')
                        bob.fc = cd.get('fc')
                        bob.ff = cd.get('ff')
                        bob.fmp = cd.get('fmp')
                        bob.furos = cd.get('furos')
                        bob.buraco = cd.get('buraco')
                        bob.esp = cd.get('esp')
                        bob.prop = cd.get('prop')
                        bob.outros = cd.get('outros')
                        bob.troca_nw = cd.get('troca_nw')
                        bob.fc_diam_ini = cd.get('fc_diam_ini')
                        bob.fc_diam_fim = cd.get('fc_diam_fim')
                        bob.ff_m_ini = cd.get('ff_m_ini')
                        bob.ff_m_fim = cd.get('ff_m_fim')
                        bob.obs = cd.get('obs')
                        bob.prop_obs = cd.get('prop_obs')
                        bob.save()
                return redirect('producao:bobinestatus', pk=bobinagem.pk)

            #     form.save()
                # messages.success(request, 'Bobine nº' + str(index) + ' - Alterações guardadas com sucesso.')

    context = {
        "form": form,
        "bobinagem": bobinagem,
        "operation": operation
    }
    return render(request, template_name, context)


# def atualizar_movimentos(request):
#     paletes = Palete.objects.filter(estado='DM')
#     for pal in paletes:
#         if pal.num_bobines_act > 0:
#             bobines = Bobine.objects.filter(palete=pal)
#             for bob in bobines:
#                 movimento = MovimentosBobines.objects.create(bobine=bob, palete=pal, timestamp=pal.timestamp, destino=bob.destino)

#     return redirect('producao:producao_home')
@login_required
def reciclado_home(request):
    template_name = 'reciclado/reciclado_home.html'
    context = {}

    return render(request, template_name, context)


@login_required
def produto_granulado_list(request):
    produto_granulado_list = ProdutoGranulado.objects.all().order_by('-timestamp')
    template_name = 'reciclado/produto_granulado_list.html'

    query = ""
    if request.GET:
        query = request.GET.get('q', '')
        # print(query)
        produto_granulado_list = ProdutoGranulado.objects.filter(
            produto_granulado__icontains=query).order_by('-timestamp')

    paginator = Paginator(produto_granulado_list, 14)
    page = request.GET.get('page')

    try:
        produto_granulado = paginator.page(page)
    except PageNotAnInteger:
        produto_granulado = paginator.page(1)
    except EmptyPage:
        produto_granulado = paginator.page(paginator.num_pages)

    context = {
        "produto_granulado": produto_granulado,
        "query": query,

    }
    return render(request, template_name, context)


@login_required
def produto_granulado_create(request):
    template_name = 'reciclado/produto_granulado_create.html'
    form = ProdutoGranuladoCreateForm(request.POST or None)

    if form.is_valid():
        cd = form.cleaned_data
        produto = cd.get('produto_granulado')
        produtos = ProdutoGranulado.objects.filter(produto_granulado=produto)
        if produtos.exists():
            messages.error(request, 'O produto ' + produto + ' já existe.')
        else:
            instance = form.save(commit=False)
            instance.user = request.user
            instance.save()
            return redirect('producao:produto_granulado_list')

    context = {
        "form": form,
    }
    return render(request, template_name, context)


@login_required
def reciclado_list(request):
    reciclado_list = Reciclado.objects.all().order_by('-timestamp')
    template_name = 'reciclado/reciclado_list.html'

    query = ""
    if request.GET:
        query = request.GET.get('q', '')
        # print(query)
        reciclado_list = Reciclado.objects.filter(
            lote__icontains=query).order_by('-timestamp')

    paginator = Paginator(reciclado_list, 14)
    page = request.GET.get('page')

    try:
        reciclado = paginator.page(page)
    except PageNotAnInteger:
        reciclado = paginator.page(1)
    except EmptyPage:
        reciclado = paginator.page(paginator.num_pages)

    context = {
        "reciclado": reciclado,
        "query": query,

    }
    return render(request, template_name, context)


@login_required
def reciclado_create(request):
    template_name = 'reciclado/reciclado_create.html'
    form = RecicladoCreateForm(request.POST or None)

    if form.is_valid():
        cd = form.cleaned_data
        estado = cd.get('estado')
        instance = form.save(commit=False)
        data = datetime.datetime.today().strftime('%Y%m%d')
        print(data)
        map(int, data)
        if instance.num < 10:
            lote = '%s-0%s' % (data[0:], str(instance.num))
        elif instance.num < 100:
            lote = '%s-%s' % (data[0:], str(instance.num))

        try:
            reciclado_latest = Reciclado.objects.all().latest('timestamp')
            instance.lote = lote
            instance.estado = estado
            instance.status = 0
            instance.user = request.user
            instance.timestamp_edit = datetime.datetime.now()
            instance.timestamp = datetime.datetime.now()
            instance.timestamp_inicio = reciclado_latest.timestamp

            print(instance.peso)

            if instance.tara == '30 kg':
                instance.peso -= 30
            elif instance.tara == '15 kg':
                instance.peso -= 15

            print(instance.tara)
            print(instance.peso)

            if estado == 'R' and instance.obs == '':
                messages.error(
                    request, 'Para rejeitar um lote, é necessário escrever a causa nas observações.')
            else:
                instance.save()
                return redirect('producao:reciclado_details', pk=instance.pk)

        except:
            messages.error(request, 'O lote que pretende criar já existe.')

    context = {
        "form": form,
    }
    return render(request, template_name, context)


@login_required
def reciclado_edit(request, pk):
    template_name = 'reciclado/reciclado_edit.html'
    reciclado = get_object_or_404(Reciclado, pk=pk)
    form = RecicladoEditForm(request.POST or None, instance=reciclado)

    if form.is_valid():
        instance = form.save(commit=False)
        instance.user = request.user
        instance.timestamp_edit = datetime.datetime.now()
        if (instance.estado == 'R' and instance.obs == '') or (instance.estado == 'NOK' and instance.obs == ''):
            messages.error(
                request, 'Para rejeitar um lote, é necessário escrever a causa nas observações.')
        else:
            instance.save()
            etiqueta = get_object_or_404(EtiquetaReciclado, reciclado=instance)
            etiqueta.produto_granulado = instance.produto_granulado.produto_granulado
            etiqueta.peso = instance.peso
            etiqueta.save()
            return redirect('producao:reciclado_list')

    context = {
        "form": form,
        "reciclado": reciclado
    }
    return render(request, template_name, context)


@login_required
def reciclado_details(request, pk):
    reciclado = get_object_or_404(Reciclado, pk=pk)
    template_name = 'reciclado/reciclado_details.html'
    form = ImprimirEtiquetaReciclado(request.POST or None)
    if form.is_valid():
        try:
            num_copias = int(form['num_copias'].value())
            etiqueta = get_object_or_404(
                EtiquetaReciclado, reciclado=reciclado)
            etiqueta.delete()
            etiqueta_reciclado = EtiquetaReciclado.objects.create(user=request.user, impressora='ARMAZEM_CAB_SQUIX_6.3_200',  num_copias=num_copias, inicio=reciclado.timestamp_inicio,
                                                                  fim=reciclado.timestamp, reciclado=reciclado, lote=reciclado.lote, produto_granulado=reciclado.produto_granulado.produto_granulado, peso=reciclado.peso)
        except:
            num_copias = int(form['num_copias'].value())
            etiqueta_reciclado = EtiquetaReciclado.objects.create(user=request.user, impressora='ARMAZEM_CAB_SQUIX_6.3_200',  num_copias=num_copias, inicio=reciclado.timestamp_inicio,
                                                                  fim=reciclado.timestamp, reciclado=reciclado, lote=reciclado.lote, produto_granulado=reciclado.produto_granulado.produto_granulado, peso=reciclado.peso)

    context = {
        "reciclado": reciclado,
        "form": form
    }

    return render(request, template_name, context)


@login_required
def movimentos_list(request):
    movimentos_list = MovimentoMP.objects.all().order_by('-timestamp')
    template_name = 'reciclado/movimento_list.html'

    query = ""
    if request.GET:
        query = request.GET.get('q', '')
        # print(query)
        movimentos_list = MovimentoMP.objects.filter(
            lote__icontains=query).order_by('-timestamp')

    paginator = Paginator(movimentos_list, 14)
    page = request.GET.get('page')

    try:
        movimentos = paginator.page(page)
    except PageNotAnInteger:
        movimentos = paginator.page(1)
    except EmptyPage:
        movimentos = paginator.page(paginator.num_pages)

    context = {
        "movimentos": movimentos,
        "query": query,

    }
    return render(request, template_name, context)


@login_required
def movimento_create(request):
    template_name = 'reciclado/movimento_create.html'
    form = MovimentoCreateForm(request.POST or None)

    if form.is_valid():
        cd = form.cleaned_data
        lote = cd.get('lote')
        tipo = cd.get('tipo')
        motivo = cd.get('motivo')
        instance = form.save(commit=False)
        instance.user = request.user
        if tipo == 'NOK' and motivo == '':
            messages.error(
                request, 'Para criar um movimento NOK é necessário escrever o motivo da mesma.')
        else:
            try:
                lote = get_object_or_404(Reciclado, lote=lote)
                if (lote.estado == 'R' or lote.estado == 'NOK'):
                    messages.error(request, 'Não é possivel movimentar lote ' +
                                   lote.lote + ' porque o seu estado é R ou NOK.')
                else:
                    movimentos = MovimentoMP.objects.filter(lote=lote.lote)
                    if movimentos.exists():
                        if movimentos.count() == 2:
                            messages.error(request, 'Não é possivel movimentar o lote ' + lote.lote +
                                           ' porque o lote selecionado ja tem um movimento de Entrada e um NOK.')
                        if movimentos.count() == 1:
                            for mov in movimentos:
                                if tipo == mov.tipo:
                                    messages.error(request, 'Não é possivel movimear o lote ' + lote.lote +
                                                   ' porque o lote selecionado ja tem um movimento de ' + tipo + '.')
                                else:
                                    instance.save()
                                    return redirect('producao:movimentos_list')

                    else:
                        if tipo == 'NOK':
                            messages.error(request, 'Não é possivel passar o lote ' +
                                           lote.lote + ' a NOK porque não Entrou em Linha.')
                        else:
                            instance.save()
                            return redirect('producao:movimentos_list')

            except:
                messages.error(request, 'O lote que picou não existe.')

    context = {
        "form": form,
    }
    return render(request, template_name, context)


@login_required
def export_bobine_to_excel(request):
    template_name = 'export/export_bobine_to_excel.html'
    form = ExportBobinesToExcel(request.POST or None)
    if form.is_valid():
        cd = form.cleaned_data
        abv = cd.get('abv')
        data_inicial = cd.get('data_inicial')
        data_final = cd.get('data_final')

        if abv != '' and data_inicial != '' and data_final != '':
            try:
                cliente = get_object_or_404(Cliente, abv=abv)
                bobines = Bobine.objects.filter(largura__cliente=cliente, bobinagem__data__range=(data_inicial, data_final))
                
                output = io.BytesIO()
                workbook = xlsxwriter.Workbook(output)
                worksheet = workbook.add_worksheet()

                row = 1
                col = 0
                for bob in bobines:
                    data = bob.bobinagem.data.strftime('%d-%m-%Y')
                    worksheet.write(row, col, bob.nome)
                    worksheet.write(row, col + 1, data)
                    worksheet.write(row, col + 2, bob.largura.cliente.nome)
                    worksheet.write(row, col + 3, bob.largura.artigo.cod)
                    worksheet.write(row, col + 4, bob.area)
                    # worksheet.write(row, col + 5, bob.palete)
                    if bob.palete != None:
                        worksheet.write(row, col + 5, bob.palete.nome)
                    row += 1

                worksheet.write('A1', 'Bobine')
                worksheet.write('B1', 'Data')
                worksheet.write('C1', 'Cliente')
                worksheet.write('D1', 'Artigo')
                worksheet.write('E1', 'SQM')
                worksheet.write('F1', 'Palete')
               
                workbook.close()

                output.seek(0)

                filename = 'Bobines.xlsx'
                response = HttpResponse(
                    output,
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = 'attachment; filename=%s' % filename

                return response
            except:
                messages.error(
                    request, 'O cliente com abreviatura ' + abv + ' não existe')
        elif abv == '' and data_inicial != '' and data_final != '':
            bobines = Bobine.objects.filter(
                bobinagem__data__range=(data_inicial, data_final))

            output = io.BytesIO()
            workbook = xlsxwriter.Workbook(output)
            worksheet = workbook.add_worksheet()

            row = 1
            col = 0
            for bob in bobines:
                data = bob.bobinagem.data.strftime('%d-%m-%Y')
                worksheet.write(row, col, bob.nome)
                worksheet.write(row, col + 1, data)
                worksheet.write(row, col + 2, bob.largura.cliente.nome)
                worksheet.write(row, col + 3, bob.largura.artigo.cod)
                worksheet.write(row, col + 4, bob.area)
                if bob.palete != None:
                    worksheet.write(row, col + 5, bob.palete.nome)
                row += 1

            worksheet.write('A1', 'Bobine')
            worksheet.write('B1', 'Data')
            worksheet.write('C1', 'Cliente')
            worksheet.write('D1', 'Artigo')
            worksheet.write('E1', 'SQM')
            worksheet.write('F1', 'Palete')
            
            workbook.close()

            output.seek(0)

            filename = 'Bobines.xlsx'
            response = HttpResponse(
                output,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename=%s' % filename

            return response

    context = {
        "form": form,
    }
    return render(request, template_name, context)


def calendario_expedicoes(request):
    template_name = 'expedicoes/expedicao_calendar.html'
    ano = datetime.now().strftime('%Y')
    mes = datetime.now().strftime('%m')
    month = monthrange(int(ano), int(mes))
    data_inicial = datetime.now().strftime('%Y-%m-' + str(1))
    data_final = datetime.now().strftime('%Y-%m-' + str(month[1]))
    cargas_previstas = Carga.objects.filter(
        data_prevista__range=(data_inicial, data_final))
    cargas_expedidas = Carga.objects.filter(
        expedida=True, data_expedicao__range=(data_inicial, data_final))

    cargas_previstas_dia = defaultdict(list)
    cargas_expedidas_dia = defaultdict(list)

    for cp in cargas_previstas:
        cp_int = (int(cp.data_prevista.strftime('%d')))
        cargas_previstas_dia[cp_int].append(cp)

    for ce in cargas_expedidas:
        ce_int = (int(ce.data_expedicao.strftime('%d')))
        cargas_expedidas_dia[ce_int].append(ce)

    cargas_previstas_dia = dict(cargas_previstas_dia)
    cargas_expedidas_dia = dict(cargas_expedidas_dia)

    print(cargas_previstas_dia)
    print(cargas_expedidas_dia)

    num_dias = month[1]
    primeiro_dia = month[0]
    semana_1 = [None] * 7
    semana_2 = [None] * 7
    semana_3 = [None] * 7
    semana_4 = [None] * 7
    semana_5 = [None] * 7
    semana_6 = [None] * 7

    if primeiro_dia == 0:
        semana_1[0] = 1
        semana_1[1] = 2
        semana_1[2] = 3
        semana_1[3] = 4
        semana_1[4] = 5
        semana_1[5] = 6
        semana_1[6] = 7

        semana_2[0] = 8
        semana_2[1] = 9
        semana_2[2] = 10
        semana_2[3] = 11
        semana_2[4] = 12
        semana_2[5] = 13
        semana_2[6] = 14

        semana_3[0] = 15
        semana_3[1] = 16
        semana_3[2] = 17
        semana_3[3] = 18
        semana_3[4] = 19
        semana_3[5] = 20
        semana_3[6] = 21

        semana_4[0] = 22
        semana_4[1] = 23
        semana_4[2] = 24
        semana_4[3] = 25
        semana_4[4] = 26
        semana_4[5] = 27
        semana_4[6] = 28
        if num_dias == 29:
            semana_5[0] = 29
        elif num_dias == 30:
            semana_5[0] = 29
            semana_5[1] = 30
        elif num_dias == 31:
            semana_5[0] = 29
            semana_5[1] = 30
            semana_5[2] = 31

    elif primeiro_dia == 1:
        semana_1[1] = 1
        semana_1[2] = 2
        semana_1[3] = 3
        semana_1[4] = 4
        semana_1[5] = 5
        semana_1[6] = 5

        semana_2[0] = 7
        semana_2[1] = 8
        semana_2[2] = 9
        semana_2[3] = 10
        semana_2[4] = 11
        semana_2[5] = 12
        semana_2[6] = 13

        semana_3[0] = 14
        semana_3[1] = 15
        semana_3[2] = 16
        semana_3[3] = 17
        semana_3[4] = 18
        semana_3[5] = 19
        semana_3[6] = 20

        semana_4[0] = 21
        semana_4[1] = 22
        semana_4[2] = 23
        semana_4[3] = 24
        semana_4[4] = 25
        semana_4[5] = 26
        semana_4[6] = 27

        semana_5[0] = 28
        if num_dias == 29:
            semana_5[1] = 29
        elif num_dias == 30:
            semana_5[1] = 29
            semana_5[2] = 30
        elif num_dias == 31:
            semana_5[1] = 29
            semana_5[2] = 30
            semana_5[3] = 31

    elif primeiro_dia == 2:
        semana_1[2] = 1
        semana_1[3] = 2
        semana_1[4] = 3
        semana_1[5] = 4
        semana_1[6] = 5

        semana_2[0] = 6
        semana_2[1] = 7
        semana_2[2] = 8
        semana_2[3] = 9
        semana_2[4] = 10
        semana_2[5] = 11
        semana_2[6] = 12

        semana_3[0] = 13
        semana_3[1] = 16
        semana_3[2] = 15
        semana_3[3] = 16
        semana_3[4] = 17
        semana_3[5] = 18
        semana_3[6] = 19

        semana_4[0] = 20
        semana_4[1] = 20
        semana_4[2] = 22
        semana_4[3] = 23
        semana_4[4] = 24
        semana_4[5] = 25
        semana_4[6] = 26

        semana_5[0] = 27
        semana_5[1] = 28
        if num_dias == 29:
            semana_5[2] = 29
        elif num_dias == 30:
            semana_5[2] = 29
            semana_5[3] = 30
        elif num_dias == 31:
            semana_5[2] = 29
            semana_5[3] = 30
            semana_5[4] = 31

    elif primeiro_dia == 3:
        semana_1[3] = 1
        semana_1[4] = 2
        semana_1[5] = 3
        semana_1[6] = 4

        semana_2[0] = 5
        semana_2[1] = 6
        semana_2[2] = 7
        semana_2[3] = 8
        semana_2[4] = 9
        semana_2[5] = 10
        semana_2[6] = 11

        semana_3[0] = 12
        semana_3[1] = 13
        semana_3[2] = 14
        semana_3[3] = 15
        semana_3[4] = 16
        semana_3[5] = 17
        semana_3[6] = 18

        semana_4[0] = 19
        semana_4[1] = 20
        semana_4[2] = 21
        semana_4[3] = 22
        semana_4[4] = 23
        semana_4[5] = 24
        semana_4[6] = 25

        semana_5[0] = 26
        semana_5[1] = 27
        semana_5[2] = 28
        if num_dias == 29:
            semana_5[3] = 29
        elif num_dias == 30:
            semana_5[3] = 29
            semana_5[4] = 30
        elif num_dias == 31:
            semana_5[3] = 29
            semana_5[4] = 30
            semana_5[5] = 31

    elif primeiro_dia == 4:
        semana_1[4] = 1
        semana_1[5] = 2
        semana_1[6] = 3

        semana_2[0] = 4
        semana_2[1] = 5
        semana_2[2] = 6
        semana_2[3] = 7
        semana_2[4] = 8
        semana_2[5] = 9
        semana_2[6] = 10

        semana_3[0] = 11
        semana_3[1] = 12
        semana_3[2] = 13
        semana_3[3] = 14
        semana_3[4] = 15
        semana_3[5] = 16
        semana_3[6] = 17

        semana_4[0] = 18
        semana_4[1] = 19
        semana_4[2] = 20
        semana_4[3] = 21
        semana_4[4] = 22
        semana_4[5] = 23
        semana_4[6] = 24

        semana_5[0] = 25
        semana_5[1] = 26
        semana_5[2] = 27
        semana_5[3] = 28
        if num_dias == 29:
            semana_5[4] = 29
        elif num_dias == 30:
            semana_5[4] = 29
            semana_5[5] = 30
        elif num_dias == 31:
            semana_5[4] = 29
            semana_5[5] = 30
            semana_5[6] = 31

    elif primeiro_dia == 5:
        semana_1[5] = 1
        semana_1[6] = 2

        semana_2[0] = 3
        semana_2[1] = 4
        semana_2[2] = 5
        semana_2[3] = 6
        semana_2[4] = 7
        semana_2[5] = 8
        semana_2[6] = 9

        semana_3[0] = 10
        semana_3[1] = 11
        semana_3[2] = 12
        semana_3[3] = 13
        semana_3[4] = 14
        semana_3[5] = 15
        semana_3[6] = 16

        semana_4[0] = 17
        semana_4[1] = 18
        semana_4[2] = 19
        semana_4[3] = 20
        semana_4[4] = 21
        semana_4[5] = 22
        semana_4[6] = 23

        semana_5[0] = 24
        semana_5[1] = 25
        semana_5[2] = 26
        semana_5[3] = 27
        semana_5[4] = 28
        if num_dias == 29:
            semana_5[5] = 29
        elif num_dias == 30:
            semana_5[5] = 29
            semana_5[6] = 30
        elif num_dias == 31:
            semana_5[5] = 29
            semana_5[6] = 30
            semana_6[0] = 31

    elif primeiro_dia == 6:

        semana_1[6] = 1

        semana_2[0] = 2
        semana_2[1] = 3
        semana_2[2] = 4
        semana_2[3] = 5
        semana_2[4] = 6
        semana_2[5] = 7
        semana_2[6] = 8

        semana_3[0] = 9
        semana_3[1] = 10
        semana_3[2] = 11
        semana_3[3] = 12
        semana_3[4] = 13
        semana_3[5] = 14
        semana_3[6] = 15

        semana_4[0] = 16
        semana_4[1] = 17
        semana_4[2] = 18
        semana_4[3] = 19
        semana_4[4] = 20
        semana_4[5] = 21
        semana_4[6] = 22

        semana_5[0] = 23
        semana_5[1] = 24
        semana_5[2] = 25
        semana_5[3] = 26
        semana_5[4] = 27
        semana_5[5] = 28
        if num_dias == 29:
            semana_5[6] = 29
        elif num_dias == 30:
            semana_5[6] = 29
            semana_6[0] = 30
        elif num_dias == 31:
            semana_5[6] = 29
            semana_6[0] = 30
            semana_6[1] = 31

    context = {
        "ano": ano,
        "mes": mes,
        "num_dias": num_dias,
        "primeiro_dia": primeiro_dia,
        "semana_1": semana_1,
        "semana_2": semana_2,
        "semana_3": semana_3,
        "semana_4": semana_4,
        "semana_5": semana_5,
        "semana_6": semana_6,
        "cargas_previstas": cargas_previstas,
        "cargas_previstas_dia": cargas_previstas_dia,
        "cargas_expedidas": cargas_expedidas,
        "cargas_expedidas_dia": cargas_expedidas_dia
    }
    return render(request, template_name, context)


@login_required
def bobinagem_edit(request, pk):
    bobinagem = get_object_or_404(Bobinagem, pk=pk)
    bobines = Bobine.objects.filter(bobinagem=bobinagem)
    template_name = 'bobine/bobinagem_edit.html'

    has_palete = False
    for bob in bobines:
        if bob.palete != None:
            has_palete = True
            break
        else:
            try:
                if bob.palete.estado == 'G':
                    has_palete = True
                    break
            except:
                pass

    form = BobinagemEditForm(request.POST or None, instance=bobinagem)
    form_has_palete = BobinagemEditHasPaleteForm(
        request.POST or None, instance=bobinagem)
    metros_nw_sup_now = bobinagem.nwsup
    metros_nw_inf_now = bobinagem.nwinf
    diam = bobinagem.diam
    if form.is_valid():
        instance = form.save(commit=False)
        instance.user = request.user

        sup = instance.lotenwsup
        inf = instance.lotenwinf
        metros_nwsup = instance.nwsup
        metros_nwinf = instance.nwinf
        instance.lotenwsup = sup.replace(" ", "")
        instance.lotenwinf = inf.replace(" ", "")

        nonwovensup = Bobinagem.objects.filter(lotenwsup=sup)
        nonwoveninf = Bobinagem.objects.filter(lotenwinf=inf)

        total_sup = instance.nwsup
        total_inf = instance.nwinf
        for ns in nonwovensup:
            total_sup += ns.nwsup
        for ni in nonwoveninf:
            total_inf += ni.nwinf

        total_sup -= metros_nw_sup_now
        total_inf -= metros_nw_inf_now

        # if (total_sup > 7500):
        #     messages.error(request, 'A soms total de metros do lote de Nonwoven superior "' + sup + '" excede o limite establecido de 7500. Por favor verifique o valor introduzido.')
        # if (total_inf > 7500):

        if (total_inf > 15000 or total_sup > 15000):
        #if (total_inf > 8500 or total_sup > 8500):
            messages.error(
                request, 'A soma total de metros dos lotes de Nonwoven excedem o limite establecido. Por favor verifique os valores introduzidos.')
        else:
            instance.save()
            desperdicio(instance.pk)
            tempo_duracao(instance.pk)
            area_bobinagem(instance.pk)
            edit_bobine(instance.pk)

            if not instance.estado == 'LAB' or instance.estado == 'HOLD':
                areas(instance.pk)

            return redirect('producao:bobinestatus', pk=bobinagem.pk)

    if form_has_palete.is_valid():
        instance = form_has_palete.save(commit=False)
        instance.user = request.user
        instance.diam = diam

        sup = instance.lotenwsup
        inf = instance.lotenwinf
        metros_nwsup = instance.nwsup
        metros_nwinf = instance.nwinf
        instance.lotenwsup = sup.replace(" ", "")
        instance.lotenwinf = inf.replace(" ", "")

        nonwovensup = Bobinagem.objects.filter(lotenwsup=sup)
        nonwoveninf = Bobinagem.objects.filter(lotenwinf=inf)

        total_sup = instance.nwsup
        total_inf = instance.nwinf
        for ns in nonwovensup:
            total_sup += ns.nwsup
        for ni in nonwoveninf:
            total_inf += ni.nwinf

        total_sup -= metros_nw_sup_now
        total_inf -= metros_nw_inf_now

        # if (total_sup > 7500):
        #     messages.error(request, 'A soms total de metros do lote de Nonwoven superior "' + sup + '" excede o limite establecido de 7500. Por favor verifique o valor introduzido.')
        # if (total_inf > 7500):

        if (total_inf > 15000 or total_sup > 15000):
        #if (total_inf > 8500 or total_sup > 8500):
            messages.error(
                request, 'A soma total de metros dos lotes de Nonwoven excedem o limite establecido. Por favor verifique os valores introduzidos.')
        else:
            instance.save()
            tempo_duracao(instance.pk)

            return redirect('producao:bobinestatus', pk=bobinagem.pk)

    context = {
        "bobinagem": bobinagem,
        "form": form,
        "form_has_palete": form_has_palete,
        "has_palete": has_palete
    }
    return render(request, template_name, context)


@login_required
def bobine_edit(request, pk):
    bobine = get_object_or_404(Bobine, pk=pk)
    template_name = 'bobine/bobine_edit.html'
    has_palete = False

    if bobine.palete != None and bobine.palete.estado == 'G':
        has_palete = True

    form = BobineEditForm(request.POST or None, instance=bobine)

    if form.is_valid():
        instance = form.save(commit=False)
        instance.user = request.user
        instance.comp_actual = instance.comp
        largura = instance.largura.largura / 1000
        instance.area = largura * instance.comp
        instance.save()
        etiqueta = get_object_or_404(EtiquetaRetrabalho, bobine=instance.nome)
        etiqueta.diam = instance.diam
        etiqueta.comp_total = instance.comp
        etiqueta.area = instance.area
        etiqueta.artigo = instance.artigo.des
        etiqueta.produto = instance.designacao_prod
        try:
            artigo_cliente = ArtigoCliente.objects.get(artigo=instance.artigo, cliente__nome=instance.cliente)
            etiqueta.cod_cliente = artigo_cliente.cod_client
            etiqueta.save()
        except:
            etiqueta.cod_cliente = None
            etiqueta.save()

        return redirect('producao:bobine_details', pk=bobine.pk)
    else:
        print('erro')

    context = {
        "bobine": bobine,
        "form": form,
        "has_palete": has_palete
    }
    return render(request, template_name, context)


@login_required
def carga_carregar(request, pk):
    carga = get_object_or_404(Carga, pk=pk)
    enc = get_object_or_404(Encomenda, pk=carga.enc.pk)
    cliente = enc.cliente
    paletes_enc = Palete.objects.filter(Q(ordem__enc=enc) & ~Q(peso_liquido=0) & Q(carga__isnull=True) | (Q(stock=True) & Q(cliente=cliente)) & Q(carga__isnull=True) & ~Q(peso_liquido=0)).order_by('nome')
    paletes_carga = Palete.objects.filter(carga=carga)
    template_name = 'carga/carga_carregar.html'

    context = {
        "carga": carga,
        "enc": enc,
        "paletes_enc": paletes_enc,
        "paletes_carga": paletes_carga
    }
    return render(request, template_name, context)

    # Q(hide=False) & Q(deleted=False),
    # Q(stock=False) | Q(quantity__gte=1))


@login_required
def add_palete_carga(request, pk_carga, pk_palete):
    carga = Carga.objects.get(pk=pk_carga)
    palete = Palete.objects.get(pk=pk_palete)
    bobines = Bobine.objects.filter(palete=palete)
    valid = True
    for bobine in bobines:
        if bobine.estado != "G":
            valid = False
            
    if carga.num_paletes_actual >= carga.num_paletes:
        messages.error(request, 'Carga completa. Não é possivel adiconar mais paletes a esta carga.')
    elif valid == False:
        messages.error(request, 'A palete que deseja inserir na carga contem bobines que não estão em GOOD. Por favor, verifique com o Sep. de Qualidade.')
    else:
        carga.num_paletes_actual += 1
        palete.carga = carga
        palete.save()
        carga.save()

    return redirect('producao:carga_carregar', pk=carga.pk)


@login_required
def remove_palete_carga(request, pk_carga, pk_palete):
    carga = Carga.objects.get(pk=pk_carga)
    palete = Palete.objects.get(pk=pk_palete)
    carga.num_paletes_actual -= 1
    palete.carga = None
    palete.save()
    carga.save()

    return redirect('producao:carga_carregar', pk=carga.pk)


@login_required
def finalizar_carga(request, pk):
    carga = Carga.objects.get(pk=pk)
    paletes = Palete.objects.filter(carga=carga)
    paletes_count = Palete.objects.filter(carga=carga).count()

    sqm = 0
    metros = 0
    num_palete_carga = 0

    if paletes_count == carga.num_paletes:
        error = 0
        for pal in paletes:
            metros += pal.comp_total
            sqm += pal.area
            pal.num_palete_carga = num_palete_carga + 1
            num_palete_carga += 1
            pal.save()
            error = gerar_etiqueta_final(pal.pk)
            if error == -1:
                break
                
        if error == -1:
            messages.error(request, 'A carga contém bobines em paletes de artigos diferentes. Não é possivel finalizar carga.')
            return redirect('producao:carga_carregar', pk=carga.pk)

        enc = carga.enc
        enc.num_paletes_actual += paletes_count
        enc.save()
        carga.estado = 'C'
        carga.metros = metros
        carga.sqm = sqm
        carga.save()
        return redirect('producao:carga_detail', pk=carga.pk)
    else:
        messages.error(
            request, 'Carga incompleta. Não é possivel finalizar carga.')
        return redirect('producao:carga_carregar', pk=carga.pk)


@login_required
def carga_expedir(request, pk):
    carga = get_object_or_404(Carga, pk=pk)
    enc = get_object_or_404(Encomenda, pk=carga.enc.pk)

    hora_expedicao = timezone.now()
    hora_expedicao = timezone.localtime(
        hora_expedicao, timezone.get_fixed_timezone(60))
    carga.data_expedicao = date.today()
    carga.hora_expedicao = timezone.localtime(
        hora_expedicao, timezone.get_fixed_timezone(60))
    carga.expedida = True
    carga.save()

    return redirect('producao:carga_list_completa')


def link_callback(uri, rel):
    """
    Convert HTML URIs to absolute system paths so xhtml2pdf can access those
    resources
    """
    result = finders.find(uri)
    if result:
            if not isinstance(result, (list, tuple)):
                    result = [result]
            result = list(os.path.realpath(path) for path in result)
            path=result[0]
    else:
            sUrl = settings.STATIC_URL        # Typically /static/
            sRoot = settings.STATIC_ROOT      # Typically /home/userX/project_static/
            mUrl = settings.MEDIA_URL         # Typically /media/
            mRoot = settings.MEDIA_ROOT       # Typically /home/userX/project_static/media/

            if uri.startswith(mUrl):
                    path = os.path.join(mRoot, uri.replace(mUrl, ""))
            elif uri.startswith(sUrl):
                    path = os.path.join(sRoot, uri.replace(sUrl, ""))
            else:
                    return uri

    # make sure that file exists
    if not os.path.isfile(path):
            raise Exception(
                    'media URI must start with %s or %s' % (sUrl, mUrl)
            )
    return path

@login_required
def carga_packinglist_details(request, *args, **kwargs):
    template_path = 'carga/carga_packinglist_details.html'
    pk = kwargs.get('pk')
    carga = get_object_or_404(Carga, pk=pk)
    paletes = Palete.objects.filter(carga=carga)

    context = {'carga': carga, 'paletes': paletes}
    # Create a Django response object, and specify content_type as pdf
    response = HttpResponse(content_type='application/pdf')
    # if download run this line
    # response['Content-Disposition'] = 'attachment; filename="report.pdf"'
    # find the template and render it.
    template = get_template(template_path)
    html = template.render(context)

    # create a pdf
    pisa_status = pisa.CreatePDF(
       html, dest=response,link_callback=link_callback)
    # if error then show some funy view
    if pisa_status.err:
       return HttpResponse('We had some errors <pre>' + html + '</pre>')
    return response



@login_required
def palete_pesagem_dm(request, pk):
    palete = get_object_or_404(Palete, pk=pk)
    bobines = Bobine.objects.filter(palete=palete)
    template_name = 'palete/palete_pesagem_dm.html'

    form = PaletePesagemDMForm(request.POST or None, instance=palete)
    peso_cores = 0
    
    if form.is_valid():
        try:
            instance = form.save(commit=False)
            perfil_embalamento = get_object_or_404(PerfilEmbalamento, pk=instance.perfil_embalamento.pk)
            if bobines.count() == 0:
                messages.error(request, 'Não é possivel pesar a palete ' + palete.nome + '. Está vazia.')
            elif instance.peso_bruto == 0:
                messages.error(request, 'A palete ' + palete.nome + ' não pode ter peso bruto de 0Kg.')
            else:
                for bobine in bobines:
                    core_largura = CoreLargura.objects.get(core=bobine.largura.perfil.core, largura=bobine.largura.largura)
                    peso_cores += core_largura.peso

                instance.peso_liquido = instance.peso_bruto - int(instance.peso_palete) - perfil_embalamento.peso - peso_cores
                instance.save()
                return redirect('producao:paletes_retrabalho')
        except:
            messages.error(request, 'A palete ' + palete.nome + ' não têm perfil de embalamento atribuido. Por favor contacte o Administrador.')

    context = {
        "palete": palete,
        "bobines": bobines,
        "form": form
       
    }
    return render(request, template_name, context)
    

@login_required
def palete_picagem_v3(request, pk):
    template_name = "palete/palete_picagem_v3.html"
    palete = get_object_or_404(Palete, pk=pk)
    print(f"$$$$$$$$$$$$--->{pk}")
    cliente = palete.cliente
    print(f"$$$$$$$$$$$$--->{cliente}")
    num_bobines = palete.num_bobines
    PicagemBobinesFormSet = formset_factory(PicagemBobines, extra=num_bobines)
    form = PaletePesagemForm(request.POST or None, instance=palete)
    peso_cores = 0

    if request.method == 'POST':
        formset = PicagemBobinesFormSet(request.POST)
        
        exists = Bobine.objects.filter(palete=palete).exists()
        if exists:
            messages.error(request, 'A Palete já se encontra fechada!')
        else:
        
            if formset.is_valid() and form.is_valid():
                instance = form.save(commit=False)
                
                cliente = 0
                count = 0
                array_bobines = []
                array_cores = []
                array_larguras = []
                array_produtos = []
                array_estados = []
                array_artigos = [] #v1
                validation = True
                if palete.cliente != None:
                    cliente = palete.cliente

                # Validação individual
                for f in formset:
                    count += 1
                    cd = f.cleaned_data
                    b = cd.get('bobine')

                    if b is not None:
                        try:
                            bobine = get_object_or_404(Bobine, nome=b)
                            array_bobines.append(bobine)
                            array_cores.append(bobine.bobinagem.perfil.core)
                            array_larguras.append(bobine.largura.largura)
                            array_produtos.append(bobine.designacao_prod)
                            array_artigos.append(bobine.artigo_id) #v1
                            array_estados.append(bobine.estado)
                                
                            if cliente != 0:
                                try:
                                    artigo_cliente = ArtigoCliente.objects.get(
                                        cliente=cliente, artigo=bobine.artigo)
                                    pass
                                except:
                                    messages.error(request, '(' + str(count) + ') O artigo da bobine ' + bobine.nome + ' não esta associado ao cliente ' + cliente.nome)
                                    validation = False

                            if bobine.estado != 'G' and bobine.estado != 'LAB' and bobine.estado != 'SC':
                                messages.error(request, '(' + str(count) + ') A bobine ' +
                                            bobine.nome + ' não está em estado GOOD, LAB ou SC.')
                                validation = False

                            if bobine.palete:
                                p_b = get_object_or_404(Palete, pk=bobine.palete.pk)
                                if p_b.estado == 'G':
                                    messages.error(request, '(' + str(count) + ') A bobine ' + bobine.nome + ' encontra-se noutra palete GOOD.')
                                    validation = False

                            if cliente != 0:
                                if bobine.diam > cliente.limsup:
                                    messages.error(request, '(' + str(count) + ') O diâmetro da bobine ' + bobine.nome + ' é superior ao limite máximo aceite pelo cliente ' + cliente.nome + '.')
                                    validation = False
                                elif bobine.diam > cliente.limsup:
                                    messages.error(request, '(' + str(count) + ') O diâmetro da bobine ' + bobine.nome + ' é inferior ao limite mínimo aceite pelo cliente ' + cliente.nome + '.')
                                    validation = False

                            if bobine.largura.largura != palete.largura_bobines:
                                messages.error(request, '(' + str(count) + ') A largura de ' + bobine.nome + ' (' + str(bobine.largura.largura) + ' mm) não corresponde a largura definida na palete de ' + str(palete.largura_bobines) + ' mm.')
                                validation = False

                            if bobine.bobinagem.perfil.core != palete.core_bobines:
                                messages.error(request, '(' + str(count) + ') O core da ' + bobine.nome + ' (' + str(bobine.bobinagem.perfil.core) + '") não corresponde ao core definido na palete de ' + str(palete.core_bobines) + '".')
                                validation = False
                            
                            current_date = datetime.datetime.now().date()
                            bobine_date = bobine.bobinagem.data
                            tdate = current_date - timedelta(days=3*30)
                            expired = 1 if bobine_date < tdate else 0
                            if expired==1:
                                messages.error(request, '(' + str(count) + ') A bobine ' + bobine.nome + ' foi produzida à mais de 3 meses.')
                                validation = False

                        except Exception as e:
                            print(e)
                            messages.error(
                                request, '(' + str(count) + ') A bobine ' + b + ' não existe.')
                            validation = False

                    else:
                        messages.error(
                            request, '(' + str(count) + ') Por favor preencha a campo nº ' + str(count) + '.')
                        validation = False

                validation_estados = True
                for estado in array_estados:
                    if estado == 'SC':
                        validation_estados = False
                        pass

                # Validação global -> remover validação global de core e largura
                if len(array_bobines) == palete.num_bobines and validation == True:
                    if len(array_bobines) > len(set(array_bobines)):
                        messages.error(request, 'A picagem contem bobines repetidas.')
                    elif len(set(array_artigos)) != 1:
                        messages.error(request, 'As bobines picadas são de artigos diferentes. Para que a palete seja válida todas as bobines têm de ter o mesmo artigo.')
                    elif len(set(array_produtos)) != 1:
                        messages.error(request, 'As bobines picadas são produtos diferentes. Para que a palete seja válida todas as bobines têm de ser o mesmo produto.')
                    elif len(set(array_estados)) != 1 and validation_estados == False:
                        messages.error(request, 'Todas as bobines tem de ser classificadas como SC antes de validar esta palete.')
                    elif instance.peso_bruto <= 0 or instance.peso_bruto == None:
                        messages.error(request, 'O peso Bruto não pode ser igual ou inferior a 0.')
                    elif instance.peso_palete == None:
                        messages.error(request, 'O peso da palete é de preenchimento obrigatório.')
                    else:
                        c = 0
                        area_sum = 0
                        comp_total = 0
                        for y in array_bobines:
                            y_b = get_object_or_404(Bobine, nome=y)
                            if y_b.palete:
                                y_p = get_object_or_404(Palete, pk=y_b.palete.pk)
                                if y_p.estado == 'DM':
                                    y_p.area -= Decimal(y_b.area)
                                    y_p.comp_total -= Decimal(y_b.bobinagem.comp_cli)
                                    y_p.num_bobines -= 1
                                    y_p.num_bobines_act -= 1
                                    y_p.save()

                        for ab in array_bobines:
                            c += 1
                            bob = get_object_or_404(Bobine, nome=ab)
                            bob.palete = palete
                            bob.posicao_palete = c
                            area_sum += bob.area
                            comp_total += bob.comp
                            applyOFtoBobine(bob,palete)
                            bob.save()
                            movimento_bobine = MovimentosBobines.objects.create(bobine=bob, palete=palete, timestamp=palete.timestamp, destino=bob.destino)

                        e_p = EtiquetaPalete.objects.get(palete=palete)
                        for x in array_bobines:
                            bobine = Bobine.objects.get(nome=x)
                            if bobine.bobinagem.perfil.retrabalho == True:
                                ano = palete.data_pal
                                ano = ano.strftime('%Y')
                                num = palete.num
                                if num < 10:
                                    palete.nome = 'R000%s-%s' % (num, ano)
                                elif num < 100:
                                    palete.nome = 'R00%s-%s' % (num, ano)
                                elif num < 1000:
                                    palete.nome = 'R0%s-%s' % (num, ano)
                                else:
                                    palete.nome = 'R%s-%s' % (num, ano)
                                palete.save()
                                e_p.palete_nome = palete.nome
                                break

                        palete.num_bobines_act = c
                        palete.area = area_sum
                        palete.comp_total = comp_total
                        palete.save()

                        bobines = Bobine.objects.filter(palete=palete)
                        bobines_nome = []
                        d_min = 0
                        d_max = 0
                        bobine_posicao = [None] * 61

                        for bn in bobines:
                            bobines_nome.append(bn.nome)
                            d = bn.diam
                            pos = bn.posicao_palete
                            bobine_posicao[pos] = bn.nome
                            if d_min == 0 and d_max == 0:
                                d_min = d
                                d_max = d
                            elif d <= d_min:
                                d_min = d
                            elif d >= d_max:
                                d_max = d

                        bobine_produto = Bobine.objects.get(nome=bobines_nome[0])
                        e_p.produto = bobine_produto.designacao_prod
                        e_p.artigo = bobine_produto.artigo.des
                        if cliente != 0:
                            artigo_cliente = ArtigoCliente.objects.get(
                                cliente=cliente, artigo=bobine_produto.artigo)
                            e_p.cod_cliente = artigo_cliente.cod_client

                        e_p.bobine1 = bobine_posicao[1]
                        e_p.bobine2 = bobine_posicao[2]
                        e_p.bobine3 = bobine_posicao[3]
                        e_p.bobine4 = bobine_posicao[4]
                        e_p.bobine5 = bobine_posicao[5]
                        e_p.bobine6 = bobine_posicao[6]
                        e_p.bobine7 = bobine_posicao[7]
                        e_p.bobine8 = bobine_posicao[8]
                        e_p.bobine9 = bobine_posicao[9]
                        e_p.bobine10 = bobine_posicao[10]
                        e_p.bobine11 = bobine_posicao[11]
                        e_p.bobine12 = bobine_posicao[12]
                        e_p.bobine13 = bobine_posicao[13]
                        e_p.bobine14 = bobine_posicao[14]
                        e_p.bobine15 = bobine_posicao[15]
                        e_p.bobine16 = bobine_posicao[16]
                        e_p.bobine17 = bobine_posicao[17]
                        e_p.bobine18 = bobine_posicao[18]
                        e_p.bobine19 = bobine_posicao[19]
                        e_p.bobine20 = bobine_posicao[20]
                        e_p.bobine21 = bobine_posicao[21]
                        e_p.bobine22 = bobine_posicao[22]
                        e_p.bobine23 = bobine_posicao[23]
                        e_p.bobine24 = bobine_posicao[24]
                        e_p.bobine25 = bobine_posicao[25]
                        e_p.bobine26 = bobine_posicao[26]
                        e_p.bobine27 = bobine_posicao[27]
                        e_p.bobine28 = bobine_posicao[28]
                        e_p.bobine29 = bobine_posicao[29]
                        e_p.bobine30 = bobine_posicao[30]
                        e_p.bobine31 = bobine_posicao[31]
                        e_p.bobine32 = bobine_posicao[32]
                        e_p.bobine33 = bobine_posicao[33]
                        e_p.bobine34 = bobine_posicao[34]
                        e_p.bobine35 = bobine_posicao[35]
                        e_p.bobine36 = bobine_posicao[36]
                        e_p.bobine37 = bobine_posicao[37]
                        e_p.bobine38 = bobine_posicao[38]
                        e_p.bobine39 = bobine_posicao[39]
                        e_p.bobine40 = bobine_posicao[40]
                        e_p.bobine41 = bobine_posicao[41]
                        e_p.bobine42 = bobine_posicao[42]
                        e_p.bobine43 = bobine_posicao[43]
                        e_p.bobine44 = bobine_posicao[44]
                        e_p.bobine45 = bobine_posicao[45]
                        e_p.bobine46 = bobine_posicao[46]
                        e_p.bobine47 = bobine_posicao[47]
                        e_p.bobine48 = bobine_posicao[48]
                        e_p.bobine49 = bobine_posicao[49]
                        e_p.bobine50 = bobine_posicao[50]
                        e_p.bobine51 = bobine_posicao[51]
                        e_p.bobine52 = bobine_posicao[52]
                        e_p.bobine53 = bobine_posicao[53]
                        e_p.bobine54 = bobine_posicao[54]
                        e_p.bobine55 = bobine_posicao[55]
                        e_p.bobine56 = bobine_posicao[56]
                        e_p.bobine57 = bobine_posicao[57]
                        e_p.bobine58 = bobine_posicao[58]
                        e_p.bobine59 = bobine_posicao[59]
                        e_p.bobine60 = bobine_posicao[60]
                        e_p.diam_min = d_min
                        e_p.diam_max = d_max
                        e_p.save()

                        perfil_embalamento = get_object_or_404(PerfilEmbalamento, pk=instance.perfil_embalamento.pk)
                        core_largura = CoreLargura.objects.get(core=palete.core_bobines, largura=palete.largura_bobines)
                        peso_cores = core_largura.peso * int(palete.num_bobines)
                        instance.peso_liquido = instance.peso_bruto - int(instance.peso_palete) - perfil_embalamento.peso - peso_cores
                        instance.save()
                        
                        

                        return redirect('producao:addbobinepalete', pk=palete.pk)
                


    else:
        formset = PicagemBobinesFormSet()

    context = {
        "palete": palete,
        "formset": formset,
        "form": form
    }
    return render(request, template_name, context)


@login_required
def sql_connect(request):

    template_name = 'encomenda/teste.html'
    
    #server = 'SRV-SAGE\SAGEX3' 
    #database = 'sagex3' 
    #username = 'X3_ELASTICTEK' 
    #password = '%ElAsTicT3k@2021!RePoRt' 
    #conn = pyodbc.connect('DRIVER={sql server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)

    server = 'SRV-SAGE-V12\SAGEX3' 
    #database = 'sagex3'
    database = 'sagex3'
    username = 'sa' 
    password = '#ElasticTek#2022#' 
    #username = 'X3_ELASTICTEK' 
    #password = '%ElAsTicT3k@2021!RePoRt' 
    conn = pyodbc.connect('DRIVER={/opt/microsoft/msodbcsql17/lib64/libmsodbcsql-17.9.so.1.1};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)


    # cursor=conn.cursor()
    # cursor.execute("select distinct e.ROWID, e.SOHNUM_0, e.ORDDAT_0, e.DEMDLVDAT_0, e.SHIDAT_0, e.EXTDLVDAT_0, e.ITMREF_0, a.ITMDES1_0, e.QTY_0, e.BPCORD_0, c.BPCNAM_0, b.GROPRI_0 from sagex3.ELASTICTEK.SORDERQ as e left join sagex3.ELASTICTEK.SORDERP as b on e.SOHNUM_0 = b.SOHNUM_0 left join sagex3.ELASTICTEK.ITMMASTER as a on e.ITMREF_0 = a.ITMREF_0 left join sagex3.ELASTICTEK.BPCUSTOMER as c on e.BPCORD_0 = c.BPCNUM_0 where e.ORDDAT_0 > '2021-01-01' order by e.ROWID desc;")
    # result = cursor.fetchall()

    # for r in result:
    #     if Encomenda.objects.filter(eef=r[1]).exists():
    #         if LinhaEncomenda.objects.filter(encomenda=Encomenda.objects.get(eef=r[1])).exists():
    #             pass
    #         else:
    #             try:
    #                 linha_artigo = 1  
    #                 encomenda = Encomenda.objects.get(eef=r[1])
    #                 encomenda.sqm = 0
    #                 linhas = conn.cursor()
    #                 linhas.execute("select distinct e.ROWID, e.SOHNUM_0, e.ORDDAT_0, e.DEMDLVDAT_0, e.SHIDAT_0, e.EXTDLVDAT_0, e.ITMREF_0, a.ITMDES1_0, e.QTY_0, e.BPCORD_0, c.BPCNAM_0, b.GROPRI_0 from sagex3.ELASTICTEK.SORDERQ as e left join sagex3.ELASTICTEK.SORDERP as b on e.SOHNUM_0 = b.SOHNUM_0 left join sagex3.ELASTICTEK.ITMMASTER as a on e.ITMREF_0 = a.ITMREF_0 left join sagex3.ELASTICTEK.BPCUSTOMER as c on e.BPCORD_0 = c.BPCNUM_0 where e.SOHNUM_0 = '" + r[1] + "' order by e.ROWID desc;")
    #                 result_linhas = linhas.fetchall()
    #                 for linha in result_linhas:
    #                     artigo = get_object_or_404(Artigo, cod=linha[6])
    #                     nova_linha = LinhaEncomenda.objects.create(encomenda=encomenda, artigo=artigo, linha=linha_artigo, qtd = linha[8], prc=linha[11])
    #                     linha_artigo += 1
    #                     encomenda.sqm += linha[8]
    #                     encomenda.save()
    #             except:
    #                 pass
    #     else:
    #         try:
    #             linha_artigo = 1        
    #             cliente = get_object_or_404(Cliente, cod=r[9])
    #             nova_encomenda = Encomenda.objects.create(user=request.user, eef=r[1], data_encomenda=r[2], data_solicitada=r[3], data_expedicao=r[4], data_prevista_expedicao=r[5], cliente=cliente, sqm=0)
    #             linhas = conn.cursor()
    #             linhas.execute("select distinct e.ROWID, e.SOHNUM_0, e.ORDDAT_0, e.DEMDLVDAT_0, e.SHIDAT_0, e.EXTDLVDAT_0, e.ITMREF_0, a.ITMDES1_0, e.QTY_0, e.BPCORD_0, c.BPCNAM_0, b.GROPRI_0 from sagex3.ELASTICTEK.SORDERQ as e left join sagex3.ELASTICTEK.SORDERP as b on e.SOHNUM_0 = b.SOHNUM_0 left join sagex3.ELASTICTEK.ITMMASTER as a on e.ITMREF_0 = a.ITMREF_0 left join sagex3.ELASTICTEK.BPCUSTOMER as c on e.BPCORD_0 = c.BPCNUM_0 where e.SOHNUM_0 = '" + r[1] + "' order by e.ROWID desc;")
    #             result_linhas = linhas.fetchall()
    #             for linha in result_linhas:
    #                 artigo = get_object_or_404(Artigo, cod=linha[6])
    #                 nova_linha = LinhaEncomenda.objects.create(encomenda=nova_encomenda, artigo=artigo, linha=linha_artigo, qtd = linha[8], prc=linha[11])
    #                 linha_artigo += 1
    #                 nova_encomenda.sqm += linha[8]
    #                 nova_encomenda.save()
    #         except:
    #             pass 

    encomendas = Encomenda.objects.filter(data_encomenda__gte='2021-01-01')
    for encomenda in encomendas:

        paletes = encomenda.sqm / 5400
        paletes = round(paletes, 0) + 1
        if encomenda.num_paletes == 0:
            encomenda.num_paletes = paletes
            encomenda.save()
        

    #     cursor=conn.cursor()
    #     cursor.execute("select distinct o.SOHNUM_0, s.SOHNUM_0, s.PRFNUM_0 from sagex3.ELASTICTEK.SORDERQ as o left join sagex3.ELASTICTEK.SORDER as s on o.SOHNUM_0 = s.SOHNUM_0 where o.SOHNUM_0 = '" + encomenda.eef + "';")
    #     result = cursor.fetchall()

    #     encomenda.prf = result[0][2]
    #     encomenda.save()

    

    context = {
        # "result": result,
       
    }
    return render(request, template_name, context)


def dadosbase(request):
    template_name = 'producao/dadosbase_home.html'

    context = {

    }
    return render(request, template_name, context)


def load_perfis(request):
    clienteId = request.GET.get('clienteId')
    # produto = request.GET.get('produto')
    artigoId = request.GET.get('artigoId')
    # larguraFinal = request.GET.get('larguraFinal')
    # larguraOriginal = request.GET.get('larguraOriginal')
    # coreFinal = request.GET.get('coreFinal')
    # coreOriginal = request.GET.get('coreOriginal')
    perfis =  []
    artigo = get_object_or_404(Artigo, pk=artigoId)
    cliente = get_object_or_404(Cliente, pk=clienteId)

    larguras = Largura.objects.filter(artigo=artigo, cliente=cliente)
    
    for largura in larguras:
        perfil_id = largura.perfil.pk
        try:
            perfil = Perfil.objects.get(pk=perfil_id, retrabalho=True)
            perfis.append(perfil)
        except:
            pass    

        
    
    return render(request, 'retrabalho/dropdown_options_perfis.html', {'perfis': perfis})  

@login_required
def artigos_list(request):
    template_name = 'artigo/artigo_list.html'
    artigos = Artigo.objects.all().order_by('-cod')    

    context = {
        "artigos": artigos,
    }

    return render(request, template_name, context)

@login_required
def artigo_create(request):
    template_name = 'artigo/artigo_create.html'

    form = ArtigoCreateForm(request.POST or None)
    print(f'cccc-->{form.is_valid()} -- {form}')
    if form.is_valid():
        instance = form.save(commit=False)
        instance.save()

        return redirect('producao:artigos_list')

    context = {
        "form": form
    }

    return render(request, template_name, context)

@login_required
def embalamento_home(request):
    template_name = "embalamento/embalamento_home.html"

    context = {

    }

    return render(request, template_name, context)

@login_required
def paletes_emb_list(request):
    template_name = 'embalamento/palete_emb_list.html'
    paletes = PaleteEmb.objects.all().order_by('-cod')

    context = {
        "paletes": paletes
    }

    return render(request, template_name, context)

@login_required
def paletes_emb_create(request):
    template_name = 'embalamento/palete_emb_create.html'

    form = PaleteEmbCreateForm(request.POST or None)

    if form.is_valid():
        instance = form.save(commit=False)
        instance.user = request.user
        instance.save()

        return redirect('producao:paletes_emb_list')

    context = {
        "form": form
    }

    return render(request, template_name, context)

@login_required
def filmes_list(request):
    template_name = 'embalamento/filmes_list.html'
    filmes = Filme.objects.all().order_by('-cod')

    context = {
        "filmes": filmes
    }

    return render(request, template_name, context)

@login_required
def filme_create(request):
    template_name = 'embalamento/filme_create.html'

    form = FilmeCreateForm(request.POST or None)

    if form.is_valid():
        instance = form.save(commit=False)
        instance.user = request.user
        instance.save()

        return redirect('producao:filmes_list')

    context = {
        "form": form
    }

    return render(request, template_name, context)

@login_required
def cintas_list(request):
    template_name = 'embalamento/cintas_list.html'
    cintas = Cinta.objects.all().order_by('-cod')

    context = {
        "cintas": cintas
    }

    return render(request, template_name, context)

@login_required
def cinta_create(request):
    template_name = 'embalamento/cinta_create.html'

    form = CintaCreateForm(request.POST or None)

    if form.is_valid():
        instance = form.save(commit=False)
        instance.user = request.user
        instance.save()

        return redirect('producao:cintas_list')

    context = {
        "form": form
    }

    return render(request, template_name, context)
@login_required
def cintas_list(request):
    template_name = 'embalamento/cintas_list.html'
    cintas = Cinta.objects.all().order_by('-cod')

    context = {
        "cintas": cintas
    }

    return render(request, template_name, context)

@login_required
def cinta_create(request):
    template_name = 'embalamento/cinta_create.html'

    form = CintaCreateForm(request.POST or None)

    if form.is_valid():
        instance = form.save(commit=False)
        instance.user = request.user
        instance.save()

        return redirect('producao:cintas_list')

    context = {
        "form": form
    }

    return render(request, template_name, context)

@login_required
def cores_list(request):
    template_name = 'embalamento/cores_list.html'
    cores = Core.objects.all().order_by('-cod')

    context = {
        "cores": cores
    }

    return render(request, template_name, context)

@login_required
def core_create(request):
    template_name = 'embalamento/core_create.html'

    form = CoreCreateForm(request.POST or None)

    if form.is_valid():
        instance = form.save(commit=False)
        instance.user = request.user
        instance.save()

        return redirect('producao:cores_list')

    context = {
        "form": form
    }

    return render(request, template_name, context)

@login_required
def mdfs_list(request):
    template_name = 'embalamento/mdfs_list.html'
    mdfs = Mdf.objects.all().order_by('-cod')

    context = {
        "mdfs": mdfs
    }

    return render(request, template_name, context)

@login_required
def mdf_create(request):
    template_name = 'embalamento/mdf_create.html'

    form = MdfCreateForm(request.POST or None)

    if form.is_valid():
        instance = form.save(commit=False)
        instance.user = request.user
        instance.save()

        return redirect('producao:mdfs_list')

    context = {
        "form": form
    }

    return render(request, template_name, context)

@login_required
def cartao_list(request):
    template_name = 'embalamento/cartao_list.html'
    cartao = Cartao.objects.all().order_by('-cod')

    context = {
        "cartao": cartao
    }

    return render(request, template_name, context)

@login_required
def cartao_create(request):
    template_name = 'embalamento/cartao_create.html'

    form = CartaoCreateForm(request.POST or None)

    if form.is_valid():
        instance = form.save(commit=False)
        instance.user = request.user
        instance.save()

        return redirect('producao:cartao_list')

    context = {
        "form": form
    }

    return render(request, template_name, context)
    
@login_required
def transportes_list(request):
    template_name = 'transporte/transportes_list.html'
    transportes = Transporte.objects.all().order_by('-tipo')

    context = {
        "transportes": transportes
    }

    return render(request, template_name, context)

@login_required
def transporte_create(request):
    template_name = 'transporte/transporte_create.html'

    form = TransporteCreateForm(request.POST or None)

    if form.is_valid():
        instance = form.save(commit=False)
        instance.user = request.user
        instance.save()

        return redirect('producao:transportes_list')

    context = {
        "form": form
    }

    return render(request, template_name, context)

@login_required
def transportes_artigo_cliente(request, pk_artigo, pk_cliente):
    template_name = 'transporte/transporte_artigo_cliente.html'
    cliente = get_object_or_404(Cliente, pk=pk_cliente)
    artigo = get_object_or_404(Artigo, pk=pk_artigo)
    artigo_cliente = ArtigoCliente.objects.get(cliente=cliente, artigo=artigo)
    transportes = TrasporteArtigoCliente.objects.filter(artigocliente=artigo_cliente)

    context = {
        "transportes": transportes,
        "cliente": cliente,
        "artigo": artigo,
        "artigo_cliente": artigo_cliente
    }

    return render(request, template_name, context)

@login_required
def transportes_artigo_cliente_add(request, pk):
    template_name = 'transporte/transportes_artigo_cliente_add.html'
    artigo_cliente = get_object_or_404(ArtigoCliente, pk=pk)
    artigo = get_object_or_404(Artigo, pk=artigo_cliente.artigo.pk)
    cliente = get_object_or_404(Cliente, pk=artigo_cliente.cliente.pk)

    form = TransporteArtigoClienteAddCreateForm(request.POST or None)

    if form.is_valid():
        instance = form.save(commit=False)
        instance.user = request.user
        instance.artigocliente = artigo_cliente
        instance.save()

        return redirect('producao:transportes_artigo_cliente', pk_artigo=artigo.pk, pk_cliente=cliente.pk)

    context = {
        "form": form,
        "artigo": artigo,
        "cliente": cliente,
    }

    return render(request, template_name, context)


@login_required
def especificacoes_artigo_cliente(request, pk_artigo, pk_cliente):
    template_name = 'especificacoes/especificacoes_artigo_cliente.html'

    artigo = get_object_or_404(Artigo, pk=pk_artigo)
    cliente = get_object_or_404(Cliente, pk=pk_cliente)
    artigo_cliente = ArtigoCliente.objects.get(artigo=artigo, cliente=cliente)
    especificacoes = Especificacoes.objects.filter(artigo_cliente=artigo_cliente)

    context = {
        "artigo": artigo,
        "cliente": cliente,
        "artigo_cliente": artigo_cliente,
        "especificacoes": especificacoes,
    }

    return render(request, template_name, context)

@login_required
def especificacoes_artigo_cliente_add(request, pk):
    template_name = 'especificacoes/especificacoes_artigo_cliente_add.html'
    artigo_cliente = get_object_or_404(ArtigoCliente, pk=pk)
    artigo = get_object_or_404(Artigo, pk=artigo_cliente.artigo.pk)
    cliente = get_object_or_404(Cliente, pk=artigo_cliente.cliente.pk)

    form = EspecificacoesArtigoClienteAddCreateForm(request.POST or None)

    if form.is_valid():
        instance = form.save(commit=False)
        instance.user = request.user
        instance.artigo_cliente = artigo_cliente
        instance.save()

        return redirect('producao:especificacoes_artigo_cliente', pk_artigo=artigo.pk, pk_cliente=cliente.pk)

    context = {
        "form": form,
        "artigo": artigo,
        "cliente": cliente,
    }

    return render(request, template_name, context)

@login_required
def especificacoes_details(request, pk):
    template_name = 'especificacoes/especificacoes_details.html'
    especificacoes = get_object_or_404(Especificacoes, pk=pk)

    context = {
        "especificacoes": especificacoes,
    }

    return render(request, template_name, context)

@login_required
def encomenda_edit(request, pk):
    template_name = 'encomenda/encomenda_edit.html'
    enc = get_object_or_404(Encomenda, pk=pk)
    form = EncomendaEditForm(request.POST or None, instance=enc)

    if form.is_valid():
        instance = form.save(commit=False)
        instance.user = request.user
        instance.save()
        return redirect('producao:encomenda_detail', pk=instance.pk)

    context = {
        "form": form,
        "enc": enc
    }
    return render(request, template_name, context)
    

@login_required
def atribuir_destinos(request):
    template_name = 'producao/atribuir_destinos.html'
    # form = AtribuirDestino(request.POST or None)
    if "GET" == request.method:
        return render(request, 'producao/atribuir_destinos.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        wb = openpyxl.load_workbook(excel_file)
        worksheet = wb["Folha1"]
        # print(worksheet)
        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            bobine_row = str(row[0].value)
            try:
                bobine = Bobine.objects.get(nome=bobine_row)
                bobine.destino = row[1].value
                bobine.obs = row[2].value + ' || ' + bobine.obs
                bobine.save()
                row_data = list()                
                # row_data.append(str(cell.value))
                row_data.append(bobine.nome)
                row_data.append(bobine.destino)
                row_data.append(bobine.obs)
                excel_data.append(row_data)
            except:
                pass
    
            # row_data = list()
            # for cell in row:
            #     print(cell.value)
                # row_data.append(str(cell.value))
                # excel_data.append(row_data)

        # if form.is_valid():
        #     cd = form.cleaned_data            
        # destino = cd.get('destino')

        # ws = wb.active
        # first_column = ws['A']
        # bobines = []
            
            # Print the contents
            # for x in range(len(first_column)-1): 
            #     bobine = Bobine.objects.get(nome=first_column[x].value)
            #     bobine.destino = destino
            #     bobine.save()
                
          
            

          

        context = {
            "excel_data": excel_data, 
            # "destino": destino, 
            # "bobines": bobines, 
        }

        return render(request, 'producao/atribuir_destinos.html', context)

        
@login_required
def export_packing_list_carga_excel(request, pk):
    carga = get_object_or_404(Carga, pk=pk)
    encomenda = get_object_or_404(Encomenda, pk=carga.enc.pk)
    paletes = Palete.objects.filter(carga=carga)
        
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet()

    row = 1
    col = 0
    for p in paletes:
        bobines = Bobine.objects.filter(palete=p)
        for b in bobines: 
            worksheet.write(row, col, carga.enc.cliente.nome)
            worksheet.write(row, col + 1, b.palete.carga.carga)
            worksheet.write(row, col + 2, p.carga.enc.prf)
            worksheet.write(row, col + 3, b.palete.nome)
            worksheet.write(row, col + 4, b.nome)
            worksheet.write(row, col + 5, b.largura.largura)
            worksheet.write(row, col + 6, b.area)
            row += 1

    worksheet.write('A1', 'Client')
    worksheet.write('B1', 'Carga')
    worksheet.write('C1', 'PRF')
    worksheet.write('D1', 'Palete')
    worksheet.write('E1', 'Bobine')
    worksheet.write('F1', 'Largura')
    worksheet.write('G1', 'SQM')
    
    
    workbook.close()

    output.seek(0)

    filename = 'Bobines Originais.xlsx'
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=%s' % filename

    return response
    
   
@login_required
def export_bobines_originais(request):
    template_name = 'export/bobines_originais_export.html'
    form = ExportBobinesOriginais(request.POST, request.FILES)
    excel_data_bobines = list()
    excel_data_paletes = list()
    excel_data_bobinagens = list()

    
    if form.is_valid() and request.method == 'POST':
        cd = form.cleaned_data
        carga = cd.get('carga')
        prf = cd.get('prf')
        nwtipo = cd.get('nwtipo')
        nwlote = cd.get('nwlote')
        data_inicio = cd.get('data_inicio')
        data_fim = cd.get('data_fim')
        bobines = cd.get('bobines')
        paletes = cd.get('paletes')
        bobinagens = cd.get('bobinagens')

        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output)
        worksheet = workbook.add_worksheet()

        row = 1
        col = 0

        if carga != '':
            try:        
                carga = Carga.objects.get(carga=carga)
                paletes = Palete.objects.filter(carga=carga)
                for p in paletes:
                    bobines = Bobine.objects.filter(palete=p)
                    for b in bobines:                                                                             
                        worksheet.write(row, col, carga.enc.cliente.nome)
                        worksheet.write(row, col + 1, b.palete.carga.carga)
                        worksheet.write(row, col + 2, b.palete.carga.enc.prf)
                        worksheet.write(row, col + 3, b.palete.nome)
                        worksheet.write(row, col + 4, b.bobinagem.nome)
                        worksheet.write(row, col + 5, b.nome)
                        worksheet.write(row, col + 6, b.largura.designacao_prod)
                        worksheet.write(row, col + 7, b.largura.artigo.des)
                        worksheet.write(row, col + 8, b.largura.largura)
                        worksheet.write(row, col + 9, b.l_real)
                        worksheet.write(row, col + 10, b.area)
                        worksheet.write(row, col + 11, b.con)                            
                        worksheet.write(row, col + 12, b.descen)                            
                        worksheet.write(row, col + 13, b.presa)                            
                        worksheet.write(row, col + 14, b.diam_insuf)                            
                        worksheet.write(row, col + 15, b.furos)                            
                        worksheet.write(row, col + 16, b.esp)                            
                        worksheet.write(row, col + 17, b.troca_nw)                            
                        worksheet.write(row, col + 18, b.buraco)                            
                        worksheet.write(row, col + 19, b.nok)                            
                        worksheet.write(row, col + 20, b.fc)                            
                        worksheet.write(row, col + 21, b.fc_diam_ini)                            
                        worksheet.write(row, col + 22, b.fc_diam_fim)                            
                        worksheet.write(row, col + 23, b.ff)                            
                        worksheet.write(row, col + 24, b.ff_m_ini)                            
                        worksheet.write(row, col + 25, b.ff_m_fim)                            
                        worksheet.write(row, col + 26, b.fmp)                            
                        worksheet.write(row, col + 27, b.suj)                            
                        worksheet.write(row, col + 28, b.car)                            
                        worksheet.write(row, col + 29, b.lac)                            
                        worksheet.write(row, col + 30, b.ncore)                            
                        worksheet.write(row, col + 31, b.sbrt)                            
                        worksheet.write(row, col + 32, b.prop)                            
                        worksheet.write(row, col + 33, b.prop_obs)                            
                        worksheet.write(row, col + 34, b.obs)                            
                        worksheet.write(row, col + 35, b.bobinagem.tiponwinf)                            
                        worksheet.write(row, col + 36, b.bobinagem.tiponwsup)                            
                        worksheet.write(row, col + 37, b.bobinagem.lotenwinf)                            
                        worksheet.write(row, col + 38, b.bobinagem.lotenwsup)                            
                        worksheet.write(row, col + 39, b.obs)                            
                        worksheet.write(row, col + 40, b.destino)                            
                        row += 1

                worksheet.write('A1', 'Cliente')
                worksheet.write('B1', 'Carga')
                worksheet.write('C1', 'PRF')
                worksheet.write('D1', 'Palete')
                worksheet.write('E1', 'Bobinagem')
                worksheet.write('F1', 'Bobine')
                worksheet.write('G1', 'Produto')
                worksheet.write('H1', 'Artigo')
                worksheet.write('I1', 'Largura')
                worksheet.write('J1', 'Largura Real')
                worksheet.write('K1', 'SQM')
                worksheet.write('L1', 'Cónica')
                worksheet.write('M1', 'Descentrada')
                worksheet.write('N1', 'Presa')
                worksheet.write('O1', 'Diâm Insuficiente')
                worksheet.write('P1', 'Furos')
                worksheet.write('Q1', 'Espessura')
                worksheet.write('R1', 'Troca NW')
                worksheet.write('S1', 'Buraco')
                worksheet.write('T1', 'NOK')
                worksheet.write('U1', 'Falha Corte')
                worksheet.write('V1', 'Falha corte inicio')
                worksheet.write('W1', 'Falha corte fim')
                worksheet.write('X1', 'Falha Filme')
                worksheet.write('Y1', 'Falha Filme Inicio')
                worksheet.write('Z1', 'Falha Filme Fim')
                worksheet.write('AA1', 'Falha Matéria Prima')
                worksheet.write('AB1', 'Sujidade')
                worksheet.write('AC1', 'Caiu do carro')
                worksheet.write('AD1', 'Laçou')
                worksheet.write('AE1', 'Não Colou')
                worksheet.write('AF1', 'Sobrestiragem')
                worksheet.write('AG1', 'Propriedade')
                worksheet.write('AH1', 'Observações Propriedade')
                worksheet.write('AI1', 'Observações')
                worksheet.write('AJ1', 'Tipo de NW Inferior')
                worksheet.write('AK1', 'Tipo de NW Superior')
                worksheet.write('AL1', 'Lote NW Inferior')
                worksheet.write('AM1', 'Lote NW Superior')
                worksheet.write('AN1', 'Destino')
                            
                workbook.close()

                output.seek(0)

                filename = 'Bobines Originais.xlsx'
                response = HttpResponse(
                    output,
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = 'attachment; filename=%s' % filename       
                
                return response
            except:
                messages.error(request, 'A carga inserida não existe.')

        elif prf != '':
            try:    
                enc = Encomenda.objects.get(prf=prf)    
                cargas = Carga.objects.filter(enc=enc)
                for carga in cargas:
                    paletes = Palete.objects.filter(carga=carga)
                    for p in paletes:
                        bobines = Bobine.objects.filter(palete=p)
                        for b in bobines:                                                                             
                            worksheet.write(row, col, carga.enc.cliente.nome)
                            worksheet.write(row, col + 1, b.palete.carga.carga)
                            worksheet.write(row, col + 2, b.palete.carga.enc.prf)
                            worksheet.write(row, col + 3, b.palete.nome)
                            worksheet.write(row, col + 4, b.bobinagem.nome)
                            worksheet.write(row, col + 5, b.nome)
                            worksheet.write(row, col + 6, b.largura.designacao_prod)
                            worksheet.write(row, col + 7, b.largura.artigo.des)
                            worksheet.write(row, col + 8, b.largura.largura)
                            worksheet.write(row, col + 9, b.l_real)
                            worksheet.write(row, col + 10, b.area)
                            worksheet.write(row, col + 11, b.con)                            
                            worksheet.write(row, col + 12, b.descen)                            
                            worksheet.write(row, col + 13, b.presa)                            
                            worksheet.write(row, col + 14, b.diam_insuf)                            
                            worksheet.write(row, col + 15, b.furos)                            
                            worksheet.write(row, col + 16, b.esp)                            
                            worksheet.write(row, col + 17, b.troca_nw)                            
                            worksheet.write(row, col + 18, b.buraco)                            
                            worksheet.write(row, col + 19, b.nok)                            
                            worksheet.write(row, col + 20, b.fc)                            
                            worksheet.write(row, col + 21, b.fc_diam_ini)                            
                            worksheet.write(row, col + 22, b.fc_diam_fim)                            
                            worksheet.write(row, col + 23, b.ff)                            
                            worksheet.write(row, col + 24, b.ff_m_ini)                            
                            worksheet.write(row, col + 25, b.ff_m_fim)                            
                            worksheet.write(row, col + 26, b.fmp)                            
                            worksheet.write(row, col + 27, b.suj)                            
                            worksheet.write(row, col + 28, b.car)                            
                            worksheet.write(row, col + 29, b.lac)                            
                            worksheet.write(row, col + 30, b.ncore)                            
                            worksheet.write(row, col + 31, b.sbrt)                            
                            worksheet.write(row, col + 32, b.prop)                            
                            worksheet.write(row, col + 33, b.prop_obs)                            
                            worksheet.write(row, col + 34, b.obs)                            
                            worksheet.write(row, col + 35, b.bobinagem.tiponwinf)                            
                            worksheet.write(row, col + 36, b.bobinagem.tiponwsup)                            
                            worksheet.write(row, col + 37, b.bobinagem.lotenwinf)                            
                            worksheet.write(row, col + 38, b.bobinagem.lotenwsup)                            
                            worksheet.write(row, col + 39, b.obs)                            
                            worksheet.write(row, col + 40, b.destino)                            
                            row += 1

                    worksheet.write('A1', 'Cliente')
                    worksheet.write('B1', 'Carga')
                    worksheet.write('C1', 'PRF')
                    worksheet.write('D1', 'Palete')
                    worksheet.write('E1', 'Bobinagem')
                    worksheet.write('F1', 'Bobine')
                    worksheet.write('G1', 'Produto')
                    worksheet.write('H1', 'Artigo')
                    worksheet.write('I1', 'Largura')
                    worksheet.write('J1', 'Largura Real')
                    worksheet.write('K1', 'SQM')
                    worksheet.write('L1', 'Cónica')
                    worksheet.write('M1', 'Descentrada')
                    worksheet.write('N1', 'Presa')
                    worksheet.write('O1', 'Diâm Insuficiente')
                    worksheet.write('P1', 'Furos')
                    worksheet.write('Q1', 'Espessura')
                    worksheet.write('R1', 'Troca NW')
                    worksheet.write('S1', 'Buraco')
                    worksheet.write('T1', 'NOK')
                    worksheet.write('U1', 'Falha Corte')
                    worksheet.write('V1', 'Falha corte inicio')
                    worksheet.write('W1', 'Falha corte fim')
                    worksheet.write('X1', 'Falha Filme')
                    worksheet.write('Y1', 'Falha Filme Inicio')
                    worksheet.write('Z1', 'Falha Filme Fim')
                    worksheet.write('AA1', 'Falha Matéria Prima')
                    worksheet.write('AB1', 'Sujidade')
                    worksheet.write('AC1', 'Caiu do carro')
                    worksheet.write('AD1', 'Laçou')
                    worksheet.write('AE1', 'Não Colou')
                    worksheet.write('AF1', 'Sobrestiragem')
                    worksheet.write('AG1', 'Propriedade')
                    worksheet.write('AH1', 'Observações Propriedade')
                    worksheet.write('AI1', 'Observações')
                    worksheet.write('AJ1', 'Tipo de NW Inferior')
                    worksheet.write('AK1', 'Tipo de NW Superior')
                    worksheet.write('AL1', 'Lote NW Inferior')
                    worksheet.write('AM1', 'Lote NW Superior')
                    worksheet.write('AN1', 'Destino')
                                
                    workbook.close()

                    output.seek(0)

                    filename = 'Bobines Originais.xlsx'
                    response = HttpResponse(
                        output,
                        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    )
                    response['Content-Disposition'] = 'attachment; filename=%s' % filename       
                    
                    return response
            except:
                messages.error(request, 'A prf inserida não existe.')
                             
        elif nwlote != '':
            if Bobinagem.objects.filter(Q(lotenwsup=nwlote) | Q(lotenwinf=nwlote)).exists():  
                bobinagens_nw = Bobinagem.objects.filter(Q(lotenwsup=nwlote) | Q(lotenwinf=nwlote))
                for bobinagem in bobinagens_nw:
                    bobines = Bobine.objects.filter(bobinagem=bobinagem)
                    for b in bobines:    
                        if b.palete:
                            if b.palete.carga:                                 
                                worksheet.write(row, col, b.palete.carga.enc.cliente.nome)
                                worksheet.write(row, col + 1, b.palete.carga.carga)
                                worksheet.write(row, col + 2, b.palete.carga.enc.prf)                                   
                            worksheet.write(row, col + 3, b.palete.nome)
                        worksheet.write(row, col + 4, b.bobinagem.nome)
                        worksheet.write(row, col + 5, b.nome)
                        worksheet.write(row, col + 6, b.largura.designacao_prod)
                        worksheet.write(row, col + 7, b.largura.artigo.des)
                        worksheet.write(row, col + 8, b.largura.largura)
                        worksheet.write(row, col + 9, b.l_real)
                        worksheet.write(row, col + 10, b.area)
                        worksheet.write(row, col + 11, b.con)                            
                        worksheet.write(row, col + 12, b.descen)                            
                        worksheet.write(row, col + 13, b.presa)                            
                        worksheet.write(row, col + 14, b.diam_insuf)                            
                        worksheet.write(row, col + 15, b.furos)                            
                        worksheet.write(row, col + 16, b.esp)                            
                        worksheet.write(row, col + 17, b.troca_nw)                            
                        worksheet.write(row, col + 18, b.buraco)                            
                        worksheet.write(row, col + 19, b.nok)                            
                        worksheet.write(row, col + 20, b.fc)                            
                        worksheet.write(row, col + 21, b.fc_diam_ini)                            
                        worksheet.write(row, col + 22, b.fc_diam_fim)                            
                        worksheet.write(row, col + 23, b.ff)                            
                        worksheet.write(row, col + 24, b.ff_m_ini)                            
                        worksheet.write(row, col + 25, b.ff_m_fim)                            
                        worksheet.write(row, col + 26, b.fmp)                            
                        worksheet.write(row, col + 27, b.suj)                            
                        worksheet.write(row, col + 28, b.car)                            
                        worksheet.write(row, col + 29, b.lac)                            
                        worksheet.write(row, col + 30, b.ncore)                            
                        worksheet.write(row, col + 31, b.sbrt)                            
                        worksheet.write(row, col + 32, b.prop)                            
                        worksheet.write(row, col + 33, b.prop_obs)                            
                        worksheet.write(row, col + 34, b.obs)                            
                        worksheet.write(row, col + 35, b.bobinagem.tiponwinf)                            
                        worksheet.write(row, col + 36, b.bobinagem.tiponwsup)                            
                        worksheet.write(row, col + 37, b.bobinagem.lotenwinf)                            
                        worksheet.write(row, col + 38, b.bobinagem.lotenwsup)                            
                        worksheet.write(row, col + 39, b.obs)                            
                        worksheet.write(row, col + 40, b.destino)                            
                        row += 1

                worksheet.write('A1', 'Cliente')
                worksheet.write('B1', 'Carga')
                worksheet.write('C1', 'PRF')
                worksheet.write('D1', 'Palete')
                worksheet.write('E1', 'Bobinagem')
                worksheet.write('F1', 'Bobine')
                worksheet.write('G1', 'Produto')
                worksheet.write('H1', 'Artigo')
                worksheet.write('I1', 'Largura')
                worksheet.write('J1', 'Largura Real')
                worksheet.write('K1', 'SQM')
                worksheet.write('L1', 'Cónica')
                worksheet.write('M1', 'Descentrada')
                worksheet.write('N1', 'Presa')
                worksheet.write('O1', 'Diâm Insuficiente')
                worksheet.write('P1', 'Furos')
                worksheet.write('Q1', 'Espessura')
                worksheet.write('R1', 'Troca NW')
                worksheet.write('S1', 'Buraco')
                worksheet.write('T1', 'NOK')
                worksheet.write('U1', 'Falha Corte')
                worksheet.write('V1', 'Falha corte inicio')
                worksheet.write('W1', 'Falha corte fim')
                worksheet.write('X1', 'Falha Filme')
                worksheet.write('Y1', 'Falha Filme Inicio')
                worksheet.write('Z1', 'Falha Filme Fim')
                worksheet.write('AA1', 'Falha Matéria Prima')
                worksheet.write('AB1', 'Sujidade')
                worksheet.write('AC1', 'Caiu do carro')
                worksheet.write('AD1', 'Laçou')
                worksheet.write('AE1', 'Não Colou')
                worksheet.write('AF1', 'Sobrestiragem')
                worksheet.write('AG1', 'Propriedade')
                worksheet.write('AH1', 'Observações Propriedade')
                worksheet.write('AI1', 'Observações')
                worksheet.write('AJ1', 'Tipo de NW Inferior')
                worksheet.write('AK1', 'Tipo de NW Superior')
                worksheet.write('AL1', 'Lote NW Inferior')
                worksheet.write('AM1', 'Lote NW Superior')
                worksheet.write('AN1', 'Destino')
                            
                workbook.close()

                output.seek(0)

                filename = 'Bobines Originais.xlsx'
                response = HttpResponse(
                    output,
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = 'attachment; filename=%s' % filename       
                
                return response
            else:
                messages.error(request, 'O lote de NW inserido não existe.')
        elif nwtipo != '' and data_inicio != None and data_fim != None:    
            if Bobinagem.objects.filter(Q(data__gte=data_inicio) & Q(data__lte=data_fim) & (Q(tiponwinf=nwtipo) | Q(tiponwsup=nwtipo))).exists():  
                bobinagens_nw = Bobinagem.objects.filter(Q(data__gte=data_inicio) & Q(data__lte=data_fim) & (Q(tiponwinf=nwtipo) | Q(tiponwsup=nwtipo)))
                for bobinagem in bobinagens_nw:
                    bobines = Bobine.objects.filter(bobinagem=bobinagem)
                    for b in bobines:    
                        if b.palete:
                            if b.palete.carga:                                 
                                worksheet.write(row, col, b.palete.carga.enc.cliente.nome)
                                worksheet.write(row, col + 1, b.palete.carga.carga)
                                worksheet.write(row, col + 2, b.palete.carga.enc.prf)                                   
                            worksheet.write(row, col + 3, b.palete.nome)
                        worksheet.write(row, col + 4, b.bobinagem.nome)
                        worksheet.write(row, col + 5, b.nome)
                        worksheet.write(row, col + 6, b.largura.designacao_prod)
                        worksheet.write(row, col + 7, b.largura.artigo.des)
                        worksheet.write(row, col + 8, b.largura.largura)
                        worksheet.write(row, col + 9, b.l_real)
                        worksheet.write(row, col + 10, b.area)
                        worksheet.write(row, col + 11, b.con)                            
                        worksheet.write(row, col + 12, b.descen)                            
                        worksheet.write(row, col + 13, b.presa)                            
                        worksheet.write(row, col + 14, b.diam_insuf)                            
                        worksheet.write(row, col + 15, b.furos)                            
                        worksheet.write(row, col + 16, b.esp)                            
                        worksheet.write(row, col + 17, b.troca_nw)                            
                        worksheet.write(row, col + 18, b.buraco)                            
                        worksheet.write(row, col + 19, b.nok)                            
                        worksheet.write(row, col + 20, b.fc)                            
                        worksheet.write(row, col + 21, b.fc_diam_ini)                            
                        worksheet.write(row, col + 22, b.fc_diam_fim)                            
                        worksheet.write(row, col + 23, b.ff)                            
                        worksheet.write(row, col + 24, b.ff_m_ini)                            
                        worksheet.write(row, col + 25, b.ff_m_fim)                            
                        worksheet.write(row, col + 26, b.fmp)                            
                        worksheet.write(row, col + 27, b.suj)                            
                        worksheet.write(row, col + 28, b.car)                            
                        worksheet.write(row, col + 29, b.lac)                            
                        worksheet.write(row, col + 30, b.ncore)                            
                        worksheet.write(row, col + 31, b.sbrt)                            
                        worksheet.write(row, col + 32, b.prop)                            
                        worksheet.write(row, col + 33, b.prop_obs)                            
                        worksheet.write(row, col + 34, b.obs)                            
                        worksheet.write(row, col + 35, b.bobinagem.tiponwinf)                            
                        worksheet.write(row, col + 36, b.bobinagem.tiponwsup)                            
                        worksheet.write(row, col + 37, b.bobinagem.lotenwinf)                            
                        worksheet.write(row, col + 38, b.bobinagem.lotenwsup)                            
                        worksheet.write(row, col + 39, b.obs)                            
                        worksheet.write(row, col + 40, b.destino)                            
                        row += 1

                worksheet.write('A1', 'Cliente')
                worksheet.write('B1', 'Carga')
                worksheet.write('C1', 'PRF')
                worksheet.write('D1', 'Palete')
                worksheet.write('E1', 'Bobinagem')
                worksheet.write('F1', 'Bobine')
                worksheet.write('G1', 'Produto')
                worksheet.write('H1', 'Artigo')
                worksheet.write('I1', 'Largura')
                worksheet.write('J1', 'Largura Real')
                worksheet.write('K1', 'SQM')
                worksheet.write('L1', 'Cónica')
                worksheet.write('M1', 'Descentrada')
                worksheet.write('N1', 'Presa')
                worksheet.write('O1', 'Diâm Insuficiente')
                worksheet.write('P1', 'Furos')
                worksheet.write('Q1', 'Espessura')
                worksheet.write('R1', 'Troca NW')
                worksheet.write('S1', 'Buraco')
                worksheet.write('T1', 'NOK')
                worksheet.write('U1', 'Falha Corte')
                worksheet.write('V1', 'Falha corte inicio')
                worksheet.write('W1', 'Falha corte fim')
                worksheet.write('X1', 'Falha Filme')
                worksheet.write('Y1', 'Falha Filme Inicio')
                worksheet.write('Z1', 'Falha Filme Fim')
                worksheet.write('AA1', 'Falha Matéria Prima')
                worksheet.write('AB1', 'Sujidade')
                worksheet.write('AC1', 'Caiu do carro')
                worksheet.write('AD1', 'Laçou')
                worksheet.write('AE1', 'Não Colou')
                worksheet.write('AF1', 'Sobrestiragem')
                worksheet.write('AG1', 'Propriedade')
                worksheet.write('AH1', 'Observações Propriedade')
                worksheet.write('AI1', 'Observações')
                worksheet.write('AJ1', 'Tipo de NW Inferior')
                worksheet.write('AK1', 'Tipo de NW Superior')
                worksheet.write('AL1', 'Lote NW Inferior')
                worksheet.write('AM1', 'Lote NW Superior')
                worksheet.write('AN1', 'Destino')
                            
                workbook.close()

                output.seek(0)

                filename = 'Bobines Originais.xlsx'
                response = HttpResponse(
                    output,
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = 'attachment; filename=%s' % filename       
                
                return response
            else:
                messages.error(request, 'Por favor verifique o tipo de Nonwoven e datas inseridas(YYYY-MM-DD).')
        elif data_inicio != None and data_fim != None:
            if Bobinagem.objects.filter(Q(data__gte=data_inicio) & Q(data__lte=data_fim)).exists():  
                bobinagens_nw = Bobinagem.objects.filter(Q(data__gte=data_inicio) & Q(data__lte=data_fim))
                for bobinagem in bobinagens_nw:
                    bobines = Bobine.objects.filter(bobinagem=bobinagem)
                    for b in bobines:    
                        if b.palete:
                            if b.palete.carga:                                 
                                worksheet.write(row, col, b.palete.carga.enc.cliente.nome)
                                worksheet.write(row, col + 1, b.palete.carga.carga)
                                worksheet.write(row, col + 2, b.palete.carga.enc.prf)                                   
                            worksheet.write(row, col + 3, b.palete.nome)
                        worksheet.write(row, col + 4, b.bobinagem.nome)
                        worksheet.write(row, col + 5, b.nome)
                        worksheet.write(row, col + 6, b.largura.designacao_prod)
                        worksheet.write(row, col + 7, b.largura.artigo.des)
                        worksheet.write(row, col + 8, b.largura.largura)
                        worksheet.write(row, col + 9, b.l_real)
                        worksheet.write(row, col + 10, b.area)
                        worksheet.write(row, col + 11, b.con)                            
                        worksheet.write(row, col + 12, b.descen)                            
                        worksheet.write(row, col + 13, b.presa)                            
                        worksheet.write(row, col + 14, b.diam_insuf)                            
                        worksheet.write(row, col + 15, b.furos)                            
                        worksheet.write(row, col + 16, b.esp)                            
                        worksheet.write(row, col + 17, b.troca_nw)                            
                        worksheet.write(row, col + 18, b.buraco)                            
                        worksheet.write(row, col + 19, b.nok)                            
                        worksheet.write(row, col + 20, b.fc)                            
                        worksheet.write(row, col + 21, b.fc_diam_ini)                            
                        worksheet.write(row, col + 22, b.fc_diam_fim)                            
                        worksheet.write(row, col + 23, b.ff)                            
                        worksheet.write(row, col + 24, b.ff_m_ini)                            
                        worksheet.write(row, col + 25, b.ff_m_fim)                            
                        worksheet.write(row, col + 26, b.fmp)                            
                        worksheet.write(row, col + 27, b.suj)                            
                        worksheet.write(row, col + 28, b.car)                            
                        worksheet.write(row, col + 29, b.lac)                            
                        worksheet.write(row, col + 30, b.ncore)                            
                        worksheet.write(row, col + 31, b.sbrt)                            
                        worksheet.write(row, col + 32, b.prop)                            
                        worksheet.write(row, col + 33, b.prop_obs)                            
                        worksheet.write(row, col + 34, b.obs)                            
                        worksheet.write(row, col + 35, b.bobinagem.tiponwinf)                            
                        worksheet.write(row, col + 36, b.bobinagem.tiponwsup)                            
                        worksheet.write(row, col + 37, b.bobinagem.lotenwinf)                            
                        worksheet.write(row, col + 38, b.bobinagem.lotenwsup)                            
                        worksheet.write(row, col + 39, b.obs)                            
                        worksheet.write(row, col + 40, b.destino)                            
                        row += 1

                worksheet.write('A1', 'Cliente')
                worksheet.write('B1', 'Carga')
                worksheet.write('C1', 'PRF')
                worksheet.write('D1', 'Palete')
                worksheet.write('E1', 'Bobinagem')
                worksheet.write('F1', 'Bobine')
                worksheet.write('G1', 'Produto')
                worksheet.write('H1', 'Artigo')
                worksheet.write('I1', 'Largura')
                worksheet.write('J1', 'Largura Real')
                worksheet.write('K1', 'SQM')
                worksheet.write('L1', 'Cónica')
                worksheet.write('M1', 'Descentrada')
                worksheet.write('N1', 'Presa')
                worksheet.write('O1', 'Diâm Insuficiente')
                worksheet.write('P1', 'Furos')
                worksheet.write('Q1', 'Espessura')
                worksheet.write('R1', 'Troca NW')
                worksheet.write('S1', 'Buraco')
                worksheet.write('T1', 'NOK')
                worksheet.write('U1', 'Falha Corte')
                worksheet.write('V1', 'Falha corte inicio')
                worksheet.write('W1', 'Falha corte fim')
                worksheet.write('X1', 'Falha Filme')
                worksheet.write('Y1', 'Falha Filme Inicio')
                worksheet.write('Z1', 'Falha Filme Fim')
                worksheet.write('AA1', 'Falha Matéria Prima')
                worksheet.write('AB1', 'Sujidade')
                worksheet.write('AC1', 'Caiu do carro')
                worksheet.write('AD1', 'Laçou')
                worksheet.write('AE1', 'Não Colou')
                worksheet.write('AF1', 'Sobrestiragem')
                worksheet.write('AG1', 'Propriedade')
                worksheet.write('AH1', 'Observações Propriedade')
                worksheet.write('AI1', 'Observações')
                worksheet.write('AJ1', 'Tipo de NW Inferior')
                worksheet.write('AK1', 'Tipo de NW Superior')
                worksheet.write('AL1', 'Lote NW Inferior')
                worksheet.write('AM1', 'Lote NW Superior')
                worksheet.write('AN1', 'Destino')
                            
                workbook.close()

                output.seek(0)

                filename = 'Bobines Originais.xlsx'
                response = HttpResponse(
                    output,
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = 'attachment; filename=%s' % filename       
                
                return response
            else:
                messages.error(request, 'Por favor verifique as datas inseridas(YYYY-MM-DD).') 
        elif bobines:
            try:
                excel_file = request.FILES["bobines"]
                wb = openpyxl.load_workbook(excel_file)
                ws = wb["Folha1"]            
                for r in ws.iter_rows():
                    bobine_row = str(r[0].value)
                    if bobine_row != 'None':
                        excel_data_bobines.append(bobine_row)
                    else:
                        break               
                    
                for bobine in excel_data_bobines:
                    b = Bobine.objects.get(nome=bobine)  
                    if b.palete:
                        if b.palete.carga:                                 
                            worksheet.write(row, col, b.palete.carga.enc.cliente.nome)
                            worksheet.write(row, col + 1, b.palete.carga.carga)
                            worksheet.write(row, col + 2, b.palete.carga.enc.prf) 
                        worksheet.write(row, col + 3, b.palete.nome)
                    worksheet.write(row, col + 4, b.bobinagem.nome)
                    worksheet.write(row, col + 5, b.nome)
                    worksheet.write(row, col + 6, b.largura.designacao_prod)
                    worksheet.write(row, col + 7, b.largura.artigo.des)
                    worksheet.write(row, col + 8, b.largura.largura)
                    worksheet.write(row, col + 9, b.l_real)
                    worksheet.write(row, col + 10, b.area)
                    worksheet.write(row, col + 11, b.con)                            
                    worksheet.write(row, col + 12, b.descen)                            
                    worksheet.write(row, col + 13, b.presa)                            
                    worksheet.write(row, col + 14, b.diam_insuf)                            
                    worksheet.write(row, col + 15, b.furos)                            
                    worksheet.write(row, col + 16, b.esp)                            
                    worksheet.write(row, col + 17, b.troca_nw)                            
                    worksheet.write(row, col + 18, b.buraco)                            
                    worksheet.write(row, col + 19, b.nok)                            
                    worksheet.write(row, col + 20, b.fc)                            
                    worksheet.write(row, col + 21, b.fc_diam_ini)                            
                    worksheet.write(row, col + 22, b.fc_diam_fim)                            
                    worksheet.write(row, col + 23, b.ff)                            
                    worksheet.write(row, col + 24, b.ff_m_ini)                            
                    worksheet.write(row, col + 25, b.ff_m_fim)                            
                    worksheet.write(row, col + 26, b.fmp)                            
                    worksheet.write(row, col + 27, b.suj)                            
                    worksheet.write(row, col + 28, b.car)                            
                    worksheet.write(row, col + 29, b.lac)                            
                    worksheet.write(row, col + 30, b.ncore)                            
                    worksheet.write(row, col + 31, b.sbrt)                            
                    worksheet.write(row, col + 32, b.prop)                            
                    worksheet.write(row, col + 33, b.prop_obs)                            
                    worksheet.write(row, col + 34, b.obs)                            
                    worksheet.write(row, col + 35, b.bobinagem.tiponwinf)                            
                    worksheet.write(row, col + 36, b.bobinagem.tiponwsup)                            
                    worksheet.write(row, col + 37, b.bobinagem.lotenwinf)                            
                    worksheet.write(row, col + 38, b.bobinagem.lotenwsup)                            
                    worksheet.write(row, col + 39, b.obs)                            
                    worksheet.write(row, col + 40, b.destino)                            
                    row += 1

                worksheet.write('A1', 'Cliente')
                worksheet.write('B1', 'Carga')
                worksheet.write('C1', 'PRF')
                worksheet.write('D1', 'Palete')
                worksheet.write('E1', 'Bobinagem')
                worksheet.write('F1', 'Bobine')
                worksheet.write('G1', 'Produto')
                worksheet.write('H1', 'Artigo')
                worksheet.write('I1', 'Largura')
                worksheet.write('J1', 'Largura Real')
                worksheet.write('K1', 'SQM')
                worksheet.write('L1', 'Cónica')
                worksheet.write('M1', 'Descentrada')
                worksheet.write('N1', 'Presa')
                worksheet.write('O1', 'Diâm Insuficiente')
                worksheet.write('P1', 'Furos')
                worksheet.write('Q1', 'Espessura')
                worksheet.write('R1', 'Troca NW')
                worksheet.write('S1', 'Buraco')
                worksheet.write('T1', 'NOK')
                worksheet.write('U1', 'Falha Corte')
                worksheet.write('V1', 'Falha corte inicio')
                worksheet.write('W1', 'Falha corte fim')
                worksheet.write('X1', 'Falha Filme')
                worksheet.write('Y1', 'Falha Filme Inicio')
                worksheet.write('Z1', 'Falha Filme Fim')
                worksheet.write('AA1', 'Falha Matéria Prima')
                worksheet.write('AB1', 'Sujidade')
                worksheet.write('AC1', 'Caiu do carro')
                worksheet.write('AD1', 'Laçou')
                worksheet.write('AE1', 'Não Colou')
                worksheet.write('AF1', 'Sobrestiragem')
                worksheet.write('AG1', 'Propriedade')
                worksheet.write('AH1', 'Observações Propriedade')
                worksheet.write('AI1', 'Observações')
                worksheet.write('AJ1', 'Tipo de NW Inferior')
                worksheet.write('AK1', 'Tipo de NW Superior')
                worksheet.write('AL1', 'Lote NW Inferior')
                worksheet.write('AM1', 'Lote NW Superior')
                worksheet.write('AN1', 'Destino')
                            
                workbook.close()

                output.seek(0)

                filename = 'Bobines Originais.xlsx'
                response = HttpResponse(
                    output,
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = 'attachment; filename=%s' % filename       
                
                return response
            except:
                messages.error(request, 'Por favor verifique o ficheiro de Bobines carregado.') 
        elif paletes:
            try:
                excel_file = request.FILES["paletes"]
                wb = openpyxl.load_workbook(excel_file)
                ws = wb["Folha1"]            
                for r in ws.iter_rows():
                    palete_row = str(r[0].value)
                    if palete_row != 'None':
                        excel_data_paletes.append(palete_row)
                    else:
                        break               
                for palete in excel_data_paletes:
                    p = Palete.objects.get(nome=palete)
                    bobines = Bobine.objects.filter(palete=p) 
                    for b in bobines:
                        if b.palete:
                            if b.palete.carga:                                 
                                worksheet.write(row, col, b.palete.carga.enc.cliente.nome)
                                worksheet.write(row, col + 1, b.palete.carga.carga)
                                worksheet.write(row, col + 2, b.palete.carga.enc.prf) 
                            worksheet.write(row, col + 3, b.palete.nome)
                        worksheet.write(row, col + 4, b.bobinagem.nome)
                        worksheet.write(row, col + 5, b.nome)
                        worksheet.write(row, col + 6, b.largura.designacao_prod)
                        worksheet.write(row, col + 7, b.largura.artigo.des)
                        worksheet.write(row, col + 8, b.largura.largura)
                        worksheet.write(row, col + 9, b.l_real)
                        worksheet.write(row, col + 10, b.area)
                        worksheet.write(row, col + 11, b.con)                            
                        worksheet.write(row, col + 12, b.descen)                            
                        worksheet.write(row, col + 13, b.presa)                            
                        worksheet.write(row, col + 14, b.diam_insuf)                            
                        worksheet.write(row, col + 15, b.furos)                            
                        worksheet.write(row, col + 16, b.esp)                            
                        worksheet.write(row, col + 17, b.troca_nw)                            
                        worksheet.write(row, col + 18, b.buraco)                            
                        worksheet.write(row, col + 19, b.nok)                            
                        worksheet.write(row, col + 20, b.fc)                            
                        worksheet.write(row, col + 21, b.fc_diam_ini)                            
                        worksheet.write(row, col + 22, b.fc_diam_fim)                            
                        worksheet.write(row, col + 23, b.ff)                            
                        worksheet.write(row, col + 24, b.ff_m_ini)                            
                        worksheet.write(row, col + 25, b.ff_m_fim)                            
                        worksheet.write(row, col + 26, b.fmp)                            
                        worksheet.write(row, col + 27, b.suj)                            
                        worksheet.write(row, col + 28, b.car)                            
                        worksheet.write(row, col + 29, b.lac)                            
                        worksheet.write(row, col + 30, b.ncore)                            
                        worksheet.write(row, col + 31, b.sbrt)                            
                        worksheet.write(row, col + 32, b.prop)                            
                        worksheet.write(row, col + 33, b.prop_obs)                            
                        worksheet.write(row, col + 34, b.obs)                            
                        worksheet.write(row, col + 35, b.bobinagem.tiponwinf)                            
                        worksheet.write(row, col + 36, b.bobinagem.tiponwsup)                            
                        worksheet.write(row, col + 37, b.bobinagem.lotenwinf)                            
                        worksheet.write(row, col + 38, b.bobinagem.lotenwsup)                            
                        worksheet.write(row, col + 39, b.obs)                            
                        worksheet.write(row, col + 40, b.destino)                            
                        row += 1

                worksheet.write('A1', 'Cliente')
                worksheet.write('B1', 'Carga')
                worksheet.write('C1', 'PRF')
                worksheet.write('D1', 'Palete')
                worksheet.write('E1', 'Bobinagem')
                worksheet.write('F1', 'Bobine')
                worksheet.write('G1', 'Produto')
                worksheet.write('H1', 'Artigo')
                worksheet.write('I1', 'Largura')
                worksheet.write('J1', 'Largura Real')
                worksheet.write('K1', 'SQM')
                worksheet.write('L1', 'Cónica')
                worksheet.write('M1', 'Descentrada')
                worksheet.write('N1', 'Presa')
                worksheet.write('O1', 'Diâm Insuficiente')
                worksheet.write('P1', 'Furos')
                worksheet.write('Q1', 'Espessura')
                worksheet.write('R1', 'Troca NW')
                worksheet.write('S1', 'Buraco')
                worksheet.write('T1', 'NOK')
                worksheet.write('U1', 'Falha Corte')
                worksheet.write('V1', 'Falha corte inicio')
                worksheet.write('W1', 'Falha corte fim')
                worksheet.write('X1', 'Falha Filme')
                worksheet.write('Y1', 'Falha Filme Inicio')
                worksheet.write('Z1', 'Falha Filme Fim')
                worksheet.write('AA1', 'Falha Matéria Prima')
                worksheet.write('AB1', 'Sujidade')
                worksheet.write('AC1', 'Caiu do carro')
                worksheet.write('AD1', 'Laçou')
                worksheet.write('AE1', 'Não Colou')
                worksheet.write('AF1', 'Sobrestiragem')
                worksheet.write('AG1', 'Propriedade')
                worksheet.write('AH1', 'Observações Propriedade')
                worksheet.write('AI1', 'Observações')
                worksheet.write('AJ1', 'Tipo de NW Inferior')
                worksheet.write('AK1', 'Tipo de NW Superior')
                worksheet.write('AL1', 'Lote NW Inferior')
                worksheet.write('AM1', 'Lote NW Superior')
                worksheet.write('AN1', 'Destino')
                            
                workbook.close()

                output.seek(0)

                filename = 'Bobines Originais.xlsx'
                response = HttpResponse(
                    output,
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = 'attachment; filename=%s' % filename       
                
                return response
            except:
                messages.error(request, 'Por favor verifique o ficheiro de Paletes carregado.') 
        elif bobinagens:
            try:
                excel_file = request.FILES["bobinagens"]
                wb = openpyxl.load_workbook(excel_file)
                ws = wb["Folha1"]            
                for r in ws.iter_rows():
                    bobinagem_row = str(r[0].value)
                    if bobinagem_row != 'None':
                        excel_data_bobinagens.append(bobinagem_row)
                    else:
                        break               
                for bobinagem in excel_data_bobinagens:
                    bob = Bobinagem.objects.get(nome=bobinagem)
                    emenda = Emenda.objects.filter(bobinagem=bob)
                    for e in emenda:
                        b = Bobine.objects.get(pk=e.bobine.pk)
                        
                        worksheet.write(row, col, bob.nome)
                        worksheet.write(row, col + 1, b.nome)
                        worksheet.write(row, col + 2, b.largura.designacao_prod)
                        worksheet.write(row, col + 3, b.largura.artigo.des)
                        worksheet.write(row, col + 4, b.largura.largura)
                        worksheet.write(row, col + 5, b.l_real)
                        worksheet.write(row, col + 6, b.area)
                        worksheet.write(row, col + 7, b.con)                            
                        worksheet.write(row, col + 8, b.descen)                            
                        worksheet.write(row, col + 9, b.presa)                            
                        worksheet.write(row, col + 10, b.diam_insuf)                            
                        worksheet.write(row, col + 11, b.furos)                            
                        worksheet.write(row, col + 12, b.esp)                            
                        worksheet.write(row, col + 13, b.troca_nw)                            
                        worksheet.write(row, col + 14, b.buraco)                            
                        worksheet.write(row, col + 15, b.nok)                            
                        worksheet.write(row, col + 16, b.fc)                            
                        worksheet.write(row, col + 17, b.fc_diam_ini)                            
                        worksheet.write(row, col + 18, b.fc_diam_fim)                            
                        worksheet.write(row, col + 19, b.ff)                            
                        worksheet.write(row, col + 20, b.ff_m_ini)                            
                        worksheet.write(row, col + 21, b.ff_m_fim)                            
                        worksheet.write(row, col + 22, b.fmp)                            
                        worksheet.write(row, col + 23, b.suj)                            
                        worksheet.write(row, col + 24, b.car)                            
                        worksheet.write(row, col + 25, b.lac)                            
                        worksheet.write(row, col + 26, b.ncore)                            
                        worksheet.write(row, col + 27, b.sbrt)                            
                        worksheet.write(row, col + 28, b.prop)                            
                        worksheet.write(row, col + 29, b.prop_obs)                            
                        worksheet.write(row, col + 30, b.obs)                            
                        worksheet.write(row, col + 31, b.bobinagem.tiponwinf)                            
                        worksheet.write(row, col + 32, b.bobinagem.tiponwsup)                            
                        worksheet.write(row, col + 33, b.bobinagem.lotenwinf)                            
                        worksheet.write(row, col + 34, b.bobinagem.lotenwsup)                            
                        worksheet.write(row, col + 35, b.destino)                            
                        row += 1

                
                worksheet.write('A1', 'Bobinagem Final')
                worksheet.write('B1', 'Bobine Original')
                worksheet.write('C1', 'Produto')
                worksheet.write('D1', 'Artigo')
                worksheet.write('E1', 'Largura')
                worksheet.write('F1', 'Largura Real')
                worksheet.write('G1', 'SQM')
                worksheet.write('H1', 'Cónica')
                worksheet.write('I1', 'Descentrada')
                worksheet.write('J1', 'Presa')
                worksheet.write('K1', 'Diâm Insuficiente')
                worksheet.write('L1', 'Furos')
                worksheet.write('M1', 'Espessura')
                worksheet.write('N1', 'Troca NW')
                worksheet.write('O1', 'Buraco')
                worksheet.write('P1', 'NOK')
                worksheet.write('Q1', 'Falha Corte')
                worksheet.write('R1', 'Falha corte inicio')
                worksheet.write('S1', 'Falha corte fim')
                worksheet.write('T1', 'Falha Filme')
                worksheet.write('U1', 'Falha Filme Inicio')
                worksheet.write('V1', 'Falha Filme Fim')
                worksheet.write('W1', 'Falha Matéria Prima')
                worksheet.write('X1', 'Sujidade')
                worksheet.write('Y1', 'Caiu do carro')
                worksheet.write('Z1', 'Laçou')
                worksheet.write('AA1', 'Não Colou')
                worksheet.write('AB1', 'Sobrestiragem')
                worksheet.write('AC1', 'Propriedade')
                worksheet.write('AD1', 'Observações Propriedade')
                worksheet.write('AE1', 'Observações')
                worksheet.write('AF1', 'Tipo de NW Inferior')
                worksheet.write('AG1', 'Tipo de NW Superior')
                worksheet.write('AH1', 'Lote NW Inferior')
                worksheet.write('AI1', 'Lote NW Superior')
                worksheet.write('AJ1', 'Destino')
                            
                workbook.close()

                output.seek(0)

                filename = 'Bobines Originais.xlsx'
                response = HttpResponse(
                    output,
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = 'attachment; filename=%s' % filename       
                
                return response
            except:
                messages.error(request, 'Por favor verifique o ficheiro de Bobinagens carregado.')          
        
        
        
        
        

    context = {
        "form": form,
        
    }

    return render(request, template_name, context)